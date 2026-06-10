import csv
import io
import json
import math
import logging
import os
import re
import subprocess
import sys
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional
from unicodedata import normalize
from uuid import uuid4
from zoneinfo import ZoneInfo
from xml.etree.ElementTree import Element, SubElement, tostring

from fastapi import FastAPI, Depends, HTTPException, Query, Response
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook
from openpyxl.chart import BarChart as ExcelBarChart, LineChart as ExcelLineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.graphics import renderPDF
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image as PlatypusImage,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.graphics.shapes import Drawing, Line as DrawingLine, Rect, String
try:
    from svglib.svglib import svg2rlg
except ImportError:  # pragma: no cover - optional PDF logo renderer
    svg2rlg = None
from sqlalchemy import asc, case, desc, func, or_, text
from sqlalchemy.orm import Session

from agent import perguntar

from .database import engine, Base, get_db, run_sql_loaders
from .smtp_mailer import (
    OrderEmail,
    OrderEmailItem,
    get_smtp_settings,
    send_order_email,
)
from .models import (
    Categoria,
    Cliente,
    Contagem,
    ContagemLog,
    CriticalityReportItem,
    CriticalityReportRun,
    Estoque,
    EstoqueAtual,
    EstoqueMovimento,
    FeriadoRecife,
    Fornecedor,
    FornecedorIngrediente,
    Ingrediente,
    JobStatus,
    LogContagem,
    Pedido,
    PurchasePlan,
    PurchasePlanItem,
    PurchasePlanSupplierOption,
    Receita,
    ReceitaIngrediente,
    ResumoDiarioVenda,
    SupplierQuote,
    Venda,
    VendaDocumentoFiscal,
    VendaItem,
    VendaPagamento,
    VendaTransacao,
)
from .schemas import (
    AgentChatRequest,
    AgentChatResponse,
    ClienteCreate,
    ClienteOut,
    IngredienteOut,
    ContagemCreate,
    ContagemDetalheCategoria,
    ContagemDetalheItem,
    ContagemDetalheOut,
    ContagemListItem,
    ContagemOut,
    CriticidadeReportCategoryOut,
    CriticidadeReportItemOut,
    CriticidadeReportLatestOut,
    CriticidadeReportRunOut,
    EstoquePaginado,
    FornecedorCreate,
    FornecedorKpis,
    FornecedorListItem,
    FornecedorListResponse,
    FornecedorOrderOut,
    FornecedorOut,
    FornecedorProductOut,
    FornecedorProfileKpis,
    FornecedorProfileResponse,
    JobStatusOut,
    PedidoCreateRequest,
    PedidoCreateResponse,
    PedidoCreateItem,
    PedidoEmailResult,
    PedidoDetailItem,
    PedidoDetailResponse,
    PedidoGroupOut,
    PedidoOut,
    PedidoPaginado,
    PedidoRecommendationItem,
    PedidoRecommendationRequest,
    PedidoRecommendationResponse,
    VendaConfirmRequest,
    VendaCreateRequest,
    VendaDetailOut,
    VendaFiscalDocumentOut,
    VendaItemOut,
    VendaListItem,
    VendaPagamentoCreate,
    VendaPagamentoOut,
    VendaPaginado,
    VendaProdutoOut,
    PurchasePlanGenerateRequest,
    PurchasePlanItemOut,
    PurchasePlanItemUpdateRequest,
    PurchasePlanOut,
    PurchasePlanSimulationOut,
    PurchasePlanSimulationRequest,
    PurchasePlanSupplierOptionOut,
    RecommendedOrderGroup,
    RecommendedOrderItem,
    SupplierOption,
    SupplierQuoteOut,
    AtualizacaoLote,
    AtualizacaoIngrediente,
    LogContagemOut,
    ResultadoLote,
    DashboardAlert,
    DashboardCards,
    DashboardCategoryItem,
    DashboardFilters,
    DashboardHolidayFilter,
    DashboardHistoryPoint,
    DashboardIngredientFilter,
    DashboardKpi,
    DashboardMonthFilter,
    DashboardNamedMetric,
    DashboardRecipeItem,
    DashboardRankItem,
    DashboardResponse,
    DashboardRevenueSummary,
    DashboardUnitCategoryGroup,
    DashboardUnitRankGroup,
)


PRODUCTION_CATEGORY_ID = "CAT0015"
DASHBOARD_STOCK_UNITS = ("KG", "UND", "L")
AGENT_ROWS_PREVIEW_LIMIT = 5
RECIFE_TZ = ZoneInfo("America/Recife")
logger = logging.getLogger(__name__)


PURCHASE_NEEDED_FALLBACK_SQL = """
WITH latest_date AS (
    SELECT max("date") AS reference_date
    FROM ml.abt_reposicao
    WHERE y_comprar = 1
      AND is_compravel = 1
)
SELECT
    abt.ingredient_id,
    abt.nome_ingrediente AS ingrediente,
    abt.categoria,
    abt.unidade,
    abt.saldo_atual AS estoque_atual,
    abt.y_qtd_comprar AS qtd_sugerida,
    abt.y_nivel_criticidade AS criticidade,
    abt.criticidade_score,
    abt."date" AS data_referencia
FROM ml.abt_reposicao abt
JOIN latest_date latest ON latest.reference_date = abt."date"
WHERE abt.y_comprar = 1
  AND abt.is_compravel = 1
ORDER BY
    abt.criticidade_score DESC NULLS LAST,
    abt.y_qtd_comprar DESC NULLS LAST,
    abt.nome_ingrediente ASC
LIMIT :limit
"""

PURCHASE_NEEDED_FALLBACK_COUNT_SQL = """
WITH latest_date AS (
    SELECT max("date") AS reference_date
    FROM ml.abt_reposicao
    WHERE y_comprar = 1
      AND is_compravel = 1
)
SELECT count(1)
FROM ml.abt_reposicao abt
JOIN latest_date latest ON latest.reference_date = abt."date"
WHERE abt.y_comprar = 1
  AND abt.is_compravel = 1
"""


def _now_recife() -> datetime:
    return datetime.now(RECIFE_TZ)


@asynccontextmanager
async def lifespan(app: FastAPI):
    run_sql_loaders()
    ensure_ml_schema()
    Base.metadata.create_all(bind=engine)
    ensure_contagem_estoque_schema()
    ensure_pedidos_schema()
    ensure_purchase_plan_schema()
    ensure_sales_schema()
    yield


app = FastAPI(title="Saltim Café API", version="1.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:5174",
        "http://localhost:5175",
        "http://localhost:5176",
        "http://localhost:4173",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# Estoque
# ---------------------------------------------------------------------------


def ensure_ml_schema() -> None:
    statements = (
        "CREATE SCHEMA IF NOT EXISTS ml",
        """
        CREATE TABLE IF NOT EXISTS ml.job_status (
            id BIGSERIAL PRIMARY KEY,
            dia DATE NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            inicio_em TIMESTAMPTZ,
            fim_em TIMESTAMPTZ,
            atualizado_em TIMESTAMPTZ NOT NULL DEFAULT now(),
            error_message TEXT,
            CONSTRAINT uq_job_status_dia UNIQUE (dia),
            CONSTRAINT ck_job_status_status CHECK (status IN ('running', 'pending', 'success', 'failed'))
        )
        """,
        "ALTER TABLE ml.job_status ADD COLUMN IF NOT EXISTS inicio_em TIMESTAMPTZ",
        "ALTER TABLE ml.job_status ADD COLUMN IF NOT EXISTS fim_em TIMESTAMPTZ",
        "CREATE INDEX IF NOT EXISTS idx_job_status_dia ON ml.job_status (dia)",
    )
    raw_conn = engine.raw_connection()
    try:
        with raw_conn.cursor() as cursor:
            for statement in statements:
                cursor.execute(statement)
        raw_conn.commit()
    except Exception:
        raw_conn.rollback()
        raise
    finally:
        raw_conn.close()


def ensure_contagem_estoque_schema() -> None:
    statements = (
        "ALTER TABLE contagens ADD COLUMN IF NOT EXISTS data_contagem DATE",
        (
            "UPDATE contagens "
            "SET data_contagem = date(criada_em) "
            "WHERE data_contagem IS NULL"
        ),
        "ALTER TABLE contagens ALTER COLUMN data_contagem SET NOT NULL",
        (
            "CREATE UNIQUE INDEX IF NOT EXISTS uq_contagens_data_contagem "
            "ON contagens (data_contagem)"
        ),
        (
            "CREATE INDEX IF NOT EXISTS idx_contagens_data_contagem "
            "ON contagens (data_contagem)"
        ),
        "ALTER TABLE contagens ADD COLUMN IF NOT EXISTS estoque_snapshot_data DATE",
        "ALTER TABLE contagem_log ADD COLUMN IF NOT EXISTS estoque_id TEXT",
        "ALTER TABLE contagem_log ADD COLUMN IF NOT EXISTS estoque_data DATE",
        "ALTER TABLE contagem_log ADD COLUMN IF NOT EXISTS estoque_quantidade NUMERIC(14, 4)",
        (
            "CREATE INDEX IF NOT EXISTS idx_contagens_estoque_snapshot_data "
            "ON contagens (estoque_snapshot_data)"
        ),
        (
            "CREATE INDEX IF NOT EXISTS idx_contagem_log_estoque "
            "ON contagem_log (estoque_id)"
        ),
        (
            "UPDATE contagens "
            "SET estoque_snapshot_data = (SELECT max(date(date_time)) FROM estoques) "
            "WHERE estoque_snapshot_data IS NULL "
            "AND EXISTS (SELECT 1 FROM estoques)"
        ),
        (
            "DO $$ BEGIN "
            "IF NOT EXISTS ("
            "SELECT 1 "
            "FROM pg_constraint c "
            "JOIN pg_attribute a "
            "ON a.attrelid = c.conrelid AND a.attnum = ANY(c.conkey) "
            "WHERE c.contype = 'f' "
            "AND c.conrelid = 'contagem_log'::regclass "
            "AND c.confrelid = 'estoques'::regclass "
            "AND a.attname = 'estoque_id'"
            ") THEN "
            "ALTER TABLE contagem_log "
            "ADD CONSTRAINT fk_contagem_log_estoque "
            "FOREIGN KEY (estoque_id) REFERENCES estoques(id); "
            "END IF; "
            "END $$"
        ),
    )
    raw_conn = engine.raw_connection()
    try:
        with raw_conn.cursor() as cursor:
            for statement in statements:
                cursor.execute(statement)
        raw_conn.commit()
    except Exception:
        raw_conn.rollback()
        raise
    finally:
        raw_conn.close()


def ensure_pedidos_schema() -> None:
    statements = (
        "ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS estoque_aplicado_em TIMESTAMPTZ",
    )
    raw_conn = engine.raw_connection()
    try:
        with raw_conn.cursor() as cursor:
            for statement in statements:
                cursor.execute(statement)
        raw_conn.commit()
    except Exception:
        raw_conn.rollback()
        raise
    finally:
        raw_conn.close()


def ensure_purchase_plan_schema() -> None:
    statements = (
        """
        CREATE TABLE IF NOT EXISTS purchase_plans (
            id SERIAL PRIMARY KEY,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            status VARCHAR NOT NULL DEFAULT 'rascunho',
            source VARCHAR NOT NULL DEFAULT 'manual',
            horizon_days INTEGER NOT NULL DEFAULT 7,
            date_from DATE,
            date_to DATE,
            contagem_id INTEGER REFERENCES contagens(id),
            total_estimated NUMERIC(14,4) NOT NULL DEFAULT 0,
            approved_total NUMERIC(14,4) NOT NULL DEFAULT 0,
            critical_items_count INTEGER NOT NULL DEFAULT 0,
            avg_coverage_days DOUBLE PRECISION NOT NULL DEFAULT 0,
            savings_potential NUMERIC(14,4) NOT NULL DEFAULT 0
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS purchase_plan_items (
            id SERIAL PRIMARY KEY,
            plan_id INTEGER NOT NULL REFERENCES purchase_plans(id) ON DELETE CASCADE,
            ingredient_id VARCHAR NOT NULL REFERENCES ingredientes(id),
            ingredient_name VARCHAR NOT NULL,
            category VARCHAR,
            unit VARCHAR,
            current_qty NUMERIC(14,4) NOT NULL DEFAULT 0,
            avg_daily_usage NUMERIC(14,4) NOT NULL DEFAULT 0,
            forecast_qty NUMERIC(14,4) NOT NULL DEFAULT 0,
            in_transit_qty NUMERIC(14,4) NOT NULL DEFAULT 0,
            recommended_qty NUMERIC(14,4) NOT NULL DEFAULT 0,
            approved_qty NUMERIC(14,4) NOT NULL DEFAULT 0,
            selected_supplier_id VARCHAR,
            selected_supplier_name VARCHAR,
            estimated_unit_price NUMERIC(14,4) NOT NULL DEFAULT 0,
            estimated_total NUMERIC(14,4) NOT NULL DEFAULT 0,
            coverage_days DOUBLE PRECISION NOT NULL DEFAULT 0,
            criticality VARCHAR NOT NULL DEFAULT 'OK',
            criticality_source VARCHAR NOT NULL DEFAULT 'operational_rule',
            justification VARCHAR,
            note VARCHAR,
            CONSTRAINT uq_purchase_plan_item UNIQUE (plan_id, ingredient_id)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS purchase_plan_supplier_options (
            id SERIAL PRIMARY KEY,
            item_id INTEGER NOT NULL REFERENCES purchase_plan_items(id) ON DELETE CASCADE,
            supplier_id VARCHAR NOT NULL REFERENCES fornecedores(id),
            supplier_name VARCHAR NOT NULL,
            unit_price NUMERIC(14,4) NOT NULL DEFAULT 0,
            discount_percent NUMERIC(8,4) NOT NULL DEFAULT 0,
            min_to_discount NUMERIC(14,4) NOT NULL DEFAULT 0,
            effective_unit_price NUMERIC(14,4) NOT NULL DEFAULT 0,
            delivery_time_days INTEGER NOT NULL DEFAULT 0,
            delay_risk DOUBLE PRECISION NOT NULL DEFAULT 0,
            score DOUBLE PRECISION NOT NULL DEFAULT 0,
            recommended INTEGER NOT NULL DEFAULT 0,
            reason VARCHAR
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS supplier_quotes (
            id SERIAL PRIMARY KEY,
            plan_id INTEGER NOT NULL REFERENCES purchase_plans(id) ON DELETE CASCADE,
            supplier_id VARCHAR NOT NULL REFERENCES fornecedores(id),
            supplier_name VARCHAR NOT NULL,
            email VARCHAR,
            channel VARCHAR NOT NULL DEFAULT 'email',
            status VARCHAR NOT NULL DEFAULT 'rascunho',
            sent_at TIMESTAMPTZ,
            responded_at TIMESTAMPTZ,
            approved_at TIMESTAMPTZ,
            total_estimated NUMERIC(14,4) NOT NULL DEFAULT 0,
            notes VARCHAR,
            CONSTRAINT uq_supplier_quote_plan_supplier UNIQUE (plan_id, supplier_id)
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_purchase_plans_created_at ON purchase_plans(created_at)",
        "CREATE INDEX IF NOT EXISTS idx_purchase_plans_status ON purchase_plans(status)",
        "ALTER TABLE purchase_plan_items ADD COLUMN IF NOT EXISTS criticality_source VARCHAR NOT NULL DEFAULT 'operational_rule'",
        "CREATE INDEX IF NOT EXISTS idx_purchase_plan_items_plan ON purchase_plan_items(plan_id)",
        "CREATE INDEX IF NOT EXISTS idx_purchase_plan_supplier_options_item ON purchase_plan_supplier_options(item_id)",
        "CREATE INDEX IF NOT EXISTS idx_supplier_quotes_plan ON supplier_quotes(plan_id)",
    )
    raw_conn = engine.raw_connection()
    try:
        with raw_conn.cursor() as cursor:
            for statement in statements:
                cursor.execute(statement)
        raw_conn.commit()
    except Exception:
        raw_conn.rollback()
        raise
    finally:
        raw_conn.close()


def ensure_sales_schema() -> None:
    statements = (
        """
        CREATE TABLE IF NOT EXISTS clientes (
            id VARCHAR PRIMARY KEY,
            name VARCHAR NOT NULL,
            document VARCHAR,
            email VARCHAR,
            phone VARCHAR,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS venda_transacoes (
            id VARCHAR PRIMARY KEY,
            date_time TIMESTAMPTZ NOT NULL,
            cliente_id VARCHAR REFERENCES clientes(id),
            status VARCHAR NOT NULL DEFAULT 'aberta',
            subtotal NUMERIC(14,4) NOT NULL DEFAULT 0,
            discount_total NUMERIC(14,4) NOT NULL DEFAULT 0,
            total NUMERIC(14,4) NOT NULL DEFAULT 0,
            source VARCHAR NOT NULL DEFAULT 'balcao',
            fiscal_status VARCHAR NOT NULL DEFAULT 'pendente_preparacao',
            notes VARCHAR,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            confirmed_at TIMESTAMPTZ,
            canceled_at TIMESTAMPTZ
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS venda_itens (
            id VARCHAR PRIMARY KEY,
            venda_id VARCHAR NOT NULL REFERENCES venda_transacoes(id) ON DELETE CASCADE,
            recipe_id VARCHAR NOT NULL REFERENCES receitas(id),
            recipe_name VARCHAR NOT NULL,
            quantity NUMERIC(14,4) NOT NULL,
            unit_price NUMERIC(14,4) NOT NULL,
            discount_value NUMERIC(14,4) NOT NULL DEFAULT 0,
            total_value NUMERIC(14,4) NOT NULL,
            venda_historica_id VARCHAR REFERENCES vendas(id)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS venda_pagamentos (
            id VARCHAR PRIMARY KEY,
            venda_id VARCHAR NOT NULL REFERENCES venda_transacoes(id) ON DELETE CASCADE,
            method VARCHAR NOT NULL,
            amount NUMERIC(14,4) NOT NULL,
            status VARCHAR NOT NULL DEFAULT 'pago',
            paid_at TIMESTAMPTZ,
            change_amount NUMERIC(14,4) NOT NULL DEFAULT 0,
            external_reference VARCHAR
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS estoque_movimentos (
            id VARCHAR PRIMARY KEY,
            ingredient_id VARCHAR NOT NULL REFERENCES ingredientes(id),
            source_type VARCHAR NOT NULL,
            source_id VARCHAR NOT NULL,
            delta_qty NUMERIC(14,4) NOT NULL,
            previous_qty NUMERIC(14,4) NOT NULL,
            new_qty NUMERIC(14,4) NOT NULL,
            unit VARCHAR NOT NULL,
            reason VARCHAR NOT NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS venda_documentos_fiscais (
            id VARCHAR PRIMARY KEY,
            venda_id VARCHAR NOT NULL REFERENCES venda_transacoes(id) ON DELETE CASCADE,
            document_type VARCHAR NOT NULL DEFAULT 'NFC-e',
            status VARCHAR NOT NULL DEFAULT 'pendente_preparacao',
            provider VARCHAR,
            access_key VARCHAR,
            protocol VARCHAR,
            issued_at TIMESTAMPTZ,
            cancelled_at TIMESTAMPTZ,
            payload JSON,
            error_message VARCHAR,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_clientes_name ON clientes(name)",
        "CREATE INDEX IF NOT EXISTS idx_clientes_document ON clientes(document)",
        "CREATE INDEX IF NOT EXISTS idx_venda_transacoes_date ON venda_transacoes(date_time)",
        "CREATE INDEX IF NOT EXISTS idx_venda_transacoes_status ON venda_transacoes(status)",
        "CREATE INDEX IF NOT EXISTS idx_venda_transacoes_cliente ON venda_transacoes(cliente_id)",
        "CREATE INDEX IF NOT EXISTS idx_venda_transacoes_fiscal_status ON venda_transacoes(fiscal_status)",
        "CREATE INDEX IF NOT EXISTS idx_venda_itens_venda ON venda_itens(venda_id)",
        "CREATE INDEX IF NOT EXISTS idx_venda_itens_recipe ON venda_itens(recipe_id)",
        "CREATE INDEX IF NOT EXISTS idx_venda_itens_historical ON venda_itens(venda_historica_id)",
        "CREATE INDEX IF NOT EXISTS idx_venda_pagamentos_venda ON venda_pagamentos(venda_id)",
        "CREATE INDEX IF NOT EXISTS idx_venda_pagamentos_status ON venda_pagamentos(status)",
        "CREATE INDEX IF NOT EXISTS idx_estoque_movimentos_ingredient_date ON estoque_movimentos(ingredient_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_estoque_movimentos_source ON estoque_movimentos(source_type, source_id)",
        "CREATE INDEX IF NOT EXISTS idx_venda_documentos_fiscais_venda ON venda_documentos_fiscais(venda_id)",
        "CREATE INDEX IF NOT EXISTS idx_venda_documentos_fiscais_status ON venda_documentos_fiscais(status)",
    )
    raw_conn = engine.raw_connection()
    try:
        with raw_conn.cursor() as cursor:
            for statement in statements:
                cursor.execute(statement)
        raw_conn.commit()
    except Exception:
        raw_conn.rollback()
        raise
    finally:
        raw_conn.close()


def _countable_items_total(db: Session) -> int:
    return (
        db.query(func.count(Ingrediente.id))
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .scalar()
        or 0
    )


def _resolve_estoque_snapshot_data(db: Session, reference_date: Optional[date] = None) -> Optional[date]:
    reference_date = reference_date or _now_recife().date()
    snapshot = (
        db.query(func.max(func.date(Estoque.date_time)))
        .filter(func.date(Estoque.date_time) <= reference_date)
        .scalar()
    )
    if snapshot is None:
        snapshot = db.query(func.max(func.date(Estoque.date_time))).scalar()
    return snapshot


def _estoque_snapshot_for_ingredient(
    db: Session,
    ingredient_id: str,
    snapshot_date: Optional[date],
) -> Optional[Estoque]:
    if snapshot_date is None:
        return None
    return (
        db.query(Estoque)
        .filter(Estoque.ingredient_id == ingredient_id)
        .filter(func.date(Estoque.date_time) == snapshot_date)
        .order_by(Estoque.date_time.desc(), Estoque.id.desc())
        .first()
    )


def _latest_logs_by_ingredient(logs: list[ContagemLog]) -> dict[str, ContagemLog]:
    ordered = sorted(logs, key=lambda log: (log.criado_em, log.id))
    return {log.ingrediente_id: log for log in ordered}


def _contagem_counts(
    contagem: Contagem,
    db: Session,
    total_itens: Optional[int] = None,
) -> dict[str, int]:
    if total_itens is None:
        total_itens = _countable_items_total(db)

    logs = (
        db.query(ContagemLog)
        .join(Ingrediente, Ingrediente.id == ContagemLog.ingrediente_id)
        .filter(ContagemLog.contagem_id == contagem.id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .all()
    )
    latest_logs = _latest_logs_by_ingredient(logs)
    itens_contados = len(latest_logs)
    itens_alterados = sum(1 for log in latest_logs.values() if round(float(log.delta), 3) != 0)
    itens_sem_alteracao = itens_contados - itens_alterados

    return {
        "total_itens": total_itens,
        "itens_contados": itens_contados,
        "itens_alterados": itens_alterados,
        "itens_sem_alteracao": itens_sem_alteracao,
        "itens_nao_contados": max(total_itens - itens_contados, 0),
    }


def _contagem_list_item(
    contagem: Contagem,
    db: Session,
    total_itens: Optional[int] = None,
) -> ContagemListItem:
    return ContagemListItem(
        id=contagem.id,
        label=contagem.label,
        data_contagem=contagem.data_contagem,
        status=contagem.status,
        estoque_snapshot_data=contagem.estoque_snapshot_data,
        criada_em=contagem.criada_em,
        finalizada_em=contagem.finalizada_em,
        **_contagem_counts(contagem, db, total_itens),
    )


@app.post("/api/contagens", response_model=ContagemOut)
def create_contagem(payload: ContagemCreate, db: Session = Depends(get_db)):
    today = _now_recife().date()
    label = payload.label or f"Contagem {today.strftime('%d/%m/%Y')}"
    existing = (
        db.query(Contagem)
        .filter(Contagem.data_contagem == today)
        .order_by(Contagem.id.desc())
        .first()
    )
    if existing is not None:
        changed = False
        if existing.estoque_snapshot_data is None:
            existing.estoque_snapshot_data = _resolve_estoque_snapshot_data(db)
            changed = True
        if existing.status == "finalizada":
            existing.status = "em_andamento"
            existing.finalizada_em = None
            changed = True
        if changed:
            db.commit()
            db.refresh(existing)
        return existing

    contagem = Contagem(
        label=label,
        data_contagem=today,
        status="em_andamento",
        estoque_snapshot_data=_resolve_estoque_snapshot_data(db),
    )
    db.add(contagem)
    db.commit()
    db.refresh(contagem)
    return contagem


@app.get("/api/contagens", response_model=list[ContagemListItem])
def list_contagens(db: Session = Depends(get_db)):
    contagens = db.query(Contagem).order_by(Contagem.criada_em.desc(), Contagem.id.desc()).all()
    total_itens = _countable_items_total(db)
    return [_contagem_list_item(contagem, db, total_itens) for contagem in contagens]


@app.get("/api/contagens/{contagem_id}", response_model=ContagemOut)
def get_contagem(contagem_id: int, db: Session = Depends(get_db)):
    contagem = db.query(Contagem).filter(Contagem.id == contagem_id).first()
    if contagem is None:
        raise HTTPException(status_code=404, detail="Contagem não encontrada")
    return contagem


@app.get("/api/contagens/{contagem_id}/detalhe", response_model=ContagemDetalheOut)
def get_contagem_detalhe(contagem_id: int, db: Session = Depends(get_db)):
    contagem = db.query(Contagem).filter(Contagem.id == contagem_id).first()
    if contagem is None:
        raise HTTPException(status_code=404, detail="Contagem não encontrada")

    ingredientes = (
        db.query(Ingrediente)
        .join(Categoria, Categoria.id == Ingrediente.category_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .order_by(Categoria.name.asc(), Ingrediente.name.asc())
        .all()
    )
    logs = (
        db.query(ContagemLog)
        .join(Ingrediente, Ingrediente.id == ContagemLog.ingrediente_id)
        .filter(ContagemLog.contagem_id == contagem.id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .all()
    )
    logs_by_ingredient = _latest_logs_by_ingredient(logs)

    categorias: dict[str, dict] = {}
    for ingrediente in ingredientes:
        group = categorias.setdefault(
            ingrediente.category_id,
            {
                "category_id": ingrediente.category_id,
                "categoria": ingrediente.category,
                "items": [],
            },
        )
        log = logs_by_ingredient.get(ingrediente.id)
        snapshot = _estoque_snapshot_for_ingredient(
            db,
            ingrediente.id,
            contagem.estoque_snapshot_data,
        )
        if log is None:
            status = "nao_contado"
            item = ContagemDetalheItem(
                ingrediente_id=ingrediente.id,
                ingrediente_nome=ingrediente.name,
                unit=ingrediente.unit,
                quantidade_atual=float(ingrediente.current_qty),
                estoque_id=snapshot.id if snapshot else None,
                estoque_data=(
                    snapshot.date_time.date()
                    if snapshot and snapshot.date_time
                    else contagem.estoque_snapshot_data
                ),
                estoque_quantidade=float(snapshot.quantity) if snapshot else None,
                status=status,
            )
        else:
            delta = round(float(log.delta), 3)
            status = "sem_alteracao" if delta == 0 else "alterado"
            item = ContagemDetalheItem(
                ingrediente_id=ingrediente.id,
                ingrediente_nome=ingrediente.name,
                unit=ingrediente.unit,
                quantidade_atual=float(ingrediente.current_qty),
                estoque_id=log.estoque_id,
                estoque_data=log.estoque_data,
                estoque_quantidade=(
                    float(log.estoque_quantidade)
                    if log.estoque_quantidade is not None
                    else None
                ),
                quantidade_anterior=float(log.quantidade_anterior),
                quantidade_nova=float(log.quantidade_nova),
                delta=delta,
                status=status,
                contado_em=log.criado_em,
            )
        group["items"].append(item)

    detalhe_categorias = []
    for group in categorias.values():
        items = group["items"]
        itens_contados = sum(1 for item in items if item.status != "nao_contado")
        itens_alterados = sum(1 for item in items if item.status == "alterado")
        itens_sem_alteracao = sum(1 for item in items if item.status == "sem_alteracao")
        detalhe_categorias.append(
            ContagemDetalheCategoria(
                category_id=group["category_id"],
                categoria=group["categoria"],
                total_itens=len(items),
                itens_contados=itens_contados,
                itens_alterados=itens_alterados,
                itens_sem_alteracao=itens_sem_alteracao,
                itens_nao_contados=len(items) - itens_contados,
                items=items,
            )
        )

    counts = _contagem_counts(contagem, db, len(ingredientes))
    return ContagemDetalheOut(
        id=contagem.id,
        label=contagem.label,
        data_contagem=contagem.data_contagem,
        status=contagem.status,
        estoque_snapshot_data=contagem.estoque_snapshot_data,
        criada_em=contagem.criada_em,
        finalizada_em=contagem.finalizada_em,
        categorias=detalhe_categorias,
        **counts,
    )


@app.patch("/api/contagens/{contagem_id}/finalizar", response_model=ContagemOut)
def finalizar_contagem(contagem_id: int, db: Session = Depends(get_db)):
    contagem = db.query(Contagem).filter(Contagem.id == contagem_id).first()
    if contagem is None:
        raise HTTPException(status_code=404, detail="Contagem não encontrada")
    counts = _contagem_counts(contagem, db)
    if counts["itens_nao_contados"] > 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "A contagem só pode ser finalizada quando todos os itens "
                "compráveis forem contados."
            ),
        )
    contagem.status = "finalizada"
    contagem.finalizada_em = _now_recife()
    db.commit()
    db.refresh(contagem)
    return contagem

def _compute_status(item: Ingrediente) -> str:
    qty = float(item.current_qty)
    min_qty = float(item.min_qty)
    if qty <= 0:
        return "Esgotado"
    if qty < min_qty:
        return "Crítico"
    if qty < min_qty * 1.5:
        return "Atenção"
    return "OK"


def _latest_criticidade_run(db: Session) -> Optional[CriticalityReportRun]:
    return (
        db.query(CriticalityReportRun)
        .filter(CriticalityReportRun.status == "success")
        .filter(CriticalityReportRun.total_items > 0)
        .order_by(CriticalityReportRun.generated_at.desc(), CriticalityReportRun.id.desc())
        .first()
    )


def _latest_criticidade_by_ingredient(db: Session) -> tuple[Optional[CriticalityReportRun], dict[str, CriticalityReportItem]]:
    run = _latest_criticidade_run(db)
    if run is None:
        return None, {}
    items = (
        db.query(CriticalityReportItem)
        .filter(CriticalityReportItem.run_id == run.id)
        .all()
    )
    return run, {item.ingredient_id: item for item in items}


def _model_stock_status(ingrediente: Ingrediente, criticidade: Optional[CriticalityReportItem]) -> str:
    if float(ingrediente.current_qty) <= 0:
        return "Esgotado"
    if criticidade is not None and bool(criticidade.necessita_compra):
        return "Crítico"
    return "OK"


def _ingrediente_out(
    ingrediente: Ingrediente,
    criticidade_run: Optional[CriticalityReportRun],
    criticidade_by_ingredient: dict[str, CriticalityReportItem],
) -> IngredienteOut:
    criticidade = criticidade_by_ingredient.get(ingrediente.id)
    return IngredienteOut(
        id=ingrediente.id,
        name=ingrediente.name,
        unit=ingrediente.unit,
        category_id=ingrediente.category_id,
        price=float(ingrediente.price),
        category=ingrediente.category,
        min_qty=float(ingrediente.min_qty),
        current_qty=float(ingrediente.current_qty),
        status=_model_stock_status(ingrediente, criticidade),
        criticidade_predita=criticidade.criticidade_predita if criticidade else None,
        criticidade_report_id=criticidade_run.id if criticidade_run else None,
        criticidade_reference_date=criticidade_run.reference_date if criticidade_run else None,
    )


def _criticidade_item_out(item: CriticalityReportItem) -> CriticidadeReportItemOut:
    return CriticidadeReportItemOut(
        ingredient_id=item.ingredient_id,
        ingredient_name=item.ingredient_name,
        category_id=item.category_id,
        category=item.category,
        unit=item.unit,
        estoque_atual=_as_float(item.estoque_atual),
        stock_position=_as_float(item.stock_position),
        baseline_threshold=_as_float(item.baseline_threshold),
        cobertura_estoque_pct=_as_float(item.cobertura_estoque_pct),
        limiar_alerta_predito_pct=_as_float(item.limiar_alerta_predito_pct),
        limiar_critico_predito_pct=_as_float(item.limiar_critico_predito_pct),
        criticidade_predita=item.criticidade_predita,
        necessita_compra=bool(item.necessita_compra),
        score_alerta_compra=_as_float(item.score_alerta_compra),
        rank_position=item.rank_position,
    )


def _criticidade_report_for_date(db: Session, reference_date: date) -> CriticidadeReportLatestOut:
    run = (
        db.query(CriticalityReportRun)
        .filter(CriticalityReportRun.reference_date == reference_date)
        .order_by(CriticalityReportRun.generated_at.desc(), CriticalityReportRun.id.desc())
        .first()
    )
    if run is None:
        empty_run = CriticidadeReportRunOut(
            status="no_report",
            reference_date=reference_date,
            model_name="XGBoost Regressor",
            model_uri="runs:/58db15b4b9364e6cb1bf7d9ebe65f922/model",
            error_message="Nenhum relatório de criticidade foi gerado para hoje.",
        )
        return CriticidadeReportLatestOut(
            run=empty_run,
            distribution=[],
            categories=[],
            critical_items=[],
            zero_items=[],
            examples_critical=[],
            examples_ok=[],
        )

    items = (
        db.query(CriticalityReportItem)
        .filter(CriticalityReportItem.run_id == run.id)
        .order_by(CriticalityReportItem.rank_position.asc())
        .all()
    )
    item_outputs = [_criticidade_item_out(item) for item in items]
    zero_outputs = [item for item in item_outputs if item.estoque_atual <= 0]
    critical_outputs = [
        item for item in item_outputs if item.necessita_compra and item.estoque_atual > 0
    ]
    ok_outputs = [item for item in item_outputs if not item.necessita_compra]

    categories: list[CriticidadeReportCategoryOut] = []
    category_names = sorted({item.category or "Sem categoria" for item in item_outputs})
    for category in category_names:
        category_items = [(item) for item in item_outputs if (item.category or "Sem categoria") == category]
        total_items = len(category_items)
        alert_count = sum(1 for item in category_items if item.necessita_compra)
        ok_count = total_items - alert_count
        categories.append(
            CriticidadeReportCategoryOut(
                category=category,
                total_items=total_items,
                ok_count=ok_count,
                alert_count=alert_count,
                alert_rate=(alert_count / total_items if total_items else 0.0),
            )
        )
    categories.sort(key=lambda item: (item.alert_count, item.alert_rate), reverse=True)

    distribution = []
    if run.total_items > 0:
        distribution = [
            {"status": "OK", "count": run.ok_count, "rate": 1 - run.alert_rate},
            {"status": "Alerta de compra", "count": run.alert_count, "rate": run.alert_rate},
        ]

    return CriticidadeReportLatestOut(
        run=CriticidadeReportRunOut(
            id=run.id,
            reference_date=run.reference_date,
            generated_at=run.generated_at,
            status=run.status,
            contagem_id=run.contagem_id,
            contagem_status=run.contagem_status,
            model_name=run.model_name,
            model_uri=run.model_uri,
            model_run_id=run.model_run_id,
            total_items=run.total_items,
            ok_count=run.ok_count,
            alert_count=run.alert_count,
            alert_rate=run.alert_rate,
            metrics=run.metrics or {},
            stability=run.stability or {},
            error_message=run.error_message,
        ),
        distribution=distribution,
        categories=categories,
        critical_items=critical_outputs,
        zero_items=zero_outputs,
        examples_critical=critical_outputs[:5],
        examples_ok=sorted(ok_outputs, key=lambda item: abs(item.score_alerta_compra))[:5],
    )


def _upsert_job_status(
    db: Session,
    dia: date,
    status: str,
    error_message: Optional[str] = None,
) -> None:
    db.execute(
        text("""
            INSERT INTO ml.job_status (
                dia,
                status,
                inicio_em,
                fim_em,
                atualizado_em,
                error_message
            )
            VALUES (
                :dia,
                :status,
                CASE WHEN :status = 'running' THEN now() ELSE NULL END,
                CASE WHEN :status IN ('pending', 'success', 'failed') THEN now() ELSE NULL END,
                now(),
                :error_message
            )
            ON CONFLICT (dia)
            DO UPDATE SET
                status = EXCLUDED.status,
                inicio_em = CASE
                    WHEN EXCLUDED.status = 'running' THEN now()
                    ELSE COALESCE(ml.job_status.inicio_em, EXCLUDED.inicio_em)
                END,
                fim_em = CASE
                    WHEN EXCLUDED.status = 'running' THEN NULL
                    WHEN EXCLUDED.status IN ('pending', 'success', 'failed') THEN now()
                    ELSE ml.job_status.fim_em
                END,
                atualizado_em = now(),
                error_message = EXCLUDED.error_message
            """),
        {"dia": dia, "status": status, "error_message": error_message},
    )
    db.commit()


def _job_status_for_date(db: Session, dia: date) -> JobStatusOut:
    status = (
        db.query(JobStatus)
        .filter(JobStatus.dia == dia)
        .order_by(JobStatus.atualizado_em.desc(), JobStatus.id.desc())
        .first()
    )
    if status is None:
        return JobStatusOut(dia=dia, status="pending")
    return JobStatusOut.model_validate(status)


def _ensure_criticidade_failed_run(
    db: Session,
    reference_date: date,
    error_message: str,
) -> None:
    existing = (
        db.query(CriticalityReportRun)
        .filter(CriticalityReportRun.reference_date == reference_date)
        .order_by(CriticalityReportRun.generated_at.desc(), CriticalityReportRun.id.desc())
        .first()
    )
    if existing is not None:
        return

    db.add(
        CriticalityReportRun(
            reference_date=reference_date,
            status="failed",
            model_name="XGBoost Regressor",
            model_uri="runs:/58db15b4b9364e6cb1bf7d9ebe65f922/model",
            model_run_id="58db15b4b9364e6cb1bf7d9ebe65f922",
            metrics={
                "status": "failed",
                "source": "backend_subprocess_fallback",
            },
            stability={},
            error_message=error_message[:4000],
        )
    )
    db.commit()


@app.get("/api/ml/criticidade/job-status/latest", response_model=JobStatusOut)
def get_latest_criticidade_job_status(response: Response, db: Session = Depends(get_db)):
    response.headers["Cache-Control"] = "no-store"
    return _job_status_for_date(db, _now_recife().date())


@app.get("/api/ml/criticidade/relatorio/latest", response_model=CriticidadeReportLatestOut)
def get_latest_criticidade_report(response: Response, db: Session = Depends(get_db)):
    response.headers["Cache-Control"] = "no-store"
    return _criticidade_report_for_date(db, _now_recife().date())


@app.post("/api/ml/criticidade/relatorio/run", response_model=CriticidadeReportLatestOut)
def run_criticidade_report(response: Response, db: Session = Depends(get_db)):
    response.headers["Cache-Control"] = "no-store"
    reference_date = _now_recife().date()
    _upsert_job_status(db, reference_date, "running")
    project_root = Path(__file__).resolve().parents[2]
    script_path = project_root / "ml" / "jobs" / "generate_criticality_report.py"
    python_candidates = (
        project_root / ".venv" / "bin" / "python",
        project_root / ".venv" / "Scripts" / "python.exe",
        project_root / "backend" / ".venv" / "bin" / "python",
        project_root / "backend" / ".venv" / "Scripts" / "python.exe",
    )
    python_executable = str(next((path for path in python_candidates if path.exists()), sys.executable))
    env = os.environ.copy()
    env.setdefault("DATABASE_URL", "postgresql+psycopg://saltim:saltim123@localhost:5433/saltim_db")
    env.setdefault("MLFLOW_TRACKING_URI", "http://localhost:5000")

    result = subprocess.run(
        [python_executable, str(script_path), "--reference-date", "today"],
        cwd=str(project_root),
        env=env,
        text=True,
        capture_output=True,
        timeout=600,
        check=False,
    )
    process_output = (
        result.stderr
        or result.stdout
        or "O job terminou sem registrar uma rodada em criticidade_report_runs."
    )
    if result.returncode not in {0, 1}:
        _upsert_job_status(
            db,
            reference_date,
            "failed",
            process_output[:2000],
        )
        _ensure_criticidade_failed_run(db, reference_date, process_output)
        return _criticidade_report_for_date(db, reference_date)

    db.expire_all()
    if (
        db.query(CriticalityReportRun)
        .filter(CriticalityReportRun.reference_date == reference_date)
        .first()
        is None
    ):
        _upsert_job_status(db, reference_date, "failed", process_output[:2000])
        _ensure_criticidade_failed_run(db, reference_date, process_output)
        db.expire_all()
    return _criticidade_report_for_date(db, reference_date)


def _as_float(value) -> float:
    return float(value or 0)


EXPORT_FORMATS = {
    "csv": ("text/csv; charset=utf-8", "csv"),
    "json": ("application/json; charset=utf-8", "json"),
    "xml": ("application/xml; charset=utf-8", "xml"),
    "yaml": ("application/x-yaml; charset=utf-8", "yaml"),
    "excel": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    ),
    "pdf": ("application/pdf", "pdf"),
}

SALTIM_ORANGE = "#F07820"
SALTIM_DARK = "#232323"
SALTIM_CREAM = "#FEF4E8"
SALTIM_STONE = "#5F5E5A"
PDF_PAGE_SIZE = landscape(A4)
PDF_PROJECT_NAME = "Maestro"
PDF_PROJECT_CONTEXT = "Compras, Estoque & Operacoes"
PDF_CLIENT_NAME = "Saltim Cafe"
PDF_SIGNATURE = "Equipe Maestro"
PDF_GENERATOR_NOTE = "Relatorio oficial do projeto Maestro para Saltim Cafe"
PDF_PROJECT_LOGO_PATH = (
    Path(__file__).resolve().parents[2] / "frontend" / "public" / "images" / "maestro-logo.svg"
)
PDF_CLIENT_LOGO_PATH = (
    Path(__file__).resolve().parents[2] / "frontend" / "public" / "images" / "saltim_logo.jpg"
)
PDF_LAYOUT = {
    "page_margin_x": 12 * mm,
    "page_margin_top": 31 * mm,
    "page_margin_bottom": 18 * mm,
    "card_radius": 7,
    "card_padding": 9,
    "section_gap": 6 * mm,
}

EXPORT_THEMES = {
    "maestro-light": {
        "primary": "#1B1464",
        "accent": "#F15A24",
        "soft": "#EFEDFF",
        "surface": "#F6F5FB",
        "text": "#211F33",
        "muted": "#6F6787",
        "grid": "#E4E0EF",
    },
    "maestro-dark": {
        "primary": "#F15A24",
        "accent": "#8E7CFF",
        "soft": "#2A1931",
        "surface": "#141127",
        "text": "#F5F3FF",
        "muted": "#C9C3DC",
        "grid": "#37304F",
    },
    "saltim-light": {
        "primary": "#F07820",
        "accent": "#2D7A3A",
        "soft": "#FEF4E8",
        "surface": "#F5F4F1",
        "text": "#1C1917",
        "muted": "#78716C",
        "grid": "#E8E6E0",
    },
    "saltim-dark": {
        "primary": "#F59E42",
        "accent": "#7DD3FC",
        "soft": "#431F0B",
        "surface": "#11100F",
        "text": "#F5F5F4",
        "muted": "#D6D3D1",
        "grid": "#3A342F",
    },
    "mossy-forest-light-green": {
        "primary": "#3D5436",
        "accent": "#A46122",
        "soft": "#E8EED4",
        "surface": "#F1F5E0",
        "text": "#2C3320",
        "muted": "#6B7D52",
        "grid": "#D2DDBA",
    },
    "anime-trinity-one-piece-dark": {
        "primary": "#EF4444",
        "accent": "#FBBF24",
        "soft": "#3B1620",
        "surface": "#0F172A",
        "text": "#E2E8F0",
        "muted": "#CBD5E1",
        "grid": "#334155",
    },
}


def _export_theme(theme_id: Optional[str]) -> dict[str, str]:
    return EXPORT_THEMES.get(theme_id or "", EXPORT_THEMES["maestro-light"])


def _hex(value: str) -> str:
    return value.replace("#", "")


def _pdf_palette(theme: Optional[dict[str, str]] = None) -> dict[str, str]:
    source = theme or _export_theme(None)
    return {
        **source,
        "card": "#FFFFFF",
        "card_border": source.get("grid", "#E4E0EF"),
        "header_soft": source.get("soft", "#EFEDFF"),
        "accent_soft": "#FFF2EC",
        "table_alt": source.get("soft", "#EFEDFF"),
        "footer_text": source.get("muted", "#6F6787"),
    }


def _apply_drawing_fill(drawing, fill_color: colors.Color) -> None:
    for child in getattr(drawing, "contents", []):
        if hasattr(child, "fillColor") and child.fillColor is not None:
            child.fillColor = fill_color
        if hasattr(child, "strokeColor") and child.strokeColor is not None:
            child.strokeColor = fill_color
        _apply_drawing_fill(child, fill_color)


def _fit_svg_logo(
    path: Path,
    max_width: float,
    max_height: float,
    fill_color: Optional[colors.Color] = None,
):
    if svg2rlg is None or not path.exists():
        return None
    drawing = svg2rlg(str(path))
    if drawing is None or not drawing.width or not drawing.height:
        return None
    if fill_color is not None:
        _apply_drawing_fill(drawing, fill_color)
    scale = min(max_width / drawing.width, max_height / drawing.height)
    drawing.width *= scale
    drawing.height *= scale
    drawing.scale(scale, scale)
    return drawing


def _pdf_logo_flowable(
    path: Path,
    max_width: float,
    max_height: float,
    fallback: str,
    style: ParagraphStyle,
    fill_color: Optional[colors.Color] = None,
):
    suffix = path.suffix.lower()
    if suffix == ".svg":
        drawing = _fit_svg_logo(path, max_width, max_height, fill_color)
        if drawing is not None:
            return drawing
    if path.exists() and suffix in {".jpg", ".jpeg", ".png"}:
        image = PlatypusImage(str(path))
        image._restrictSize(max_width, max_height)
        return image
    return Paragraph(_xml_escape(fallback), style)


def _draw_pdf_logo(
    canvas,
    path: Path,
    x: float,
    y: float,
    max_width: float,
    max_height: float,
    fill_color: Optional[colors.Color] = None,
) -> None:
    suffix = path.suffix.lower()
    if suffix == ".svg":
        drawing = _fit_svg_logo(path, max_width, max_height, fill_color)
        if drawing is not None:
            renderPDF.draw(drawing, canvas, x, y)
            return
    if path.exists() and suffix in {".jpg", ".jpeg", ".png"}:
        canvas.drawImage(
            str(path),
            x,
            y,
            width=max_width,
            height=max_height,
            preserveAspectRatio=True,
            anchor="c",
            mask="auto",
        )


def _pdf_styles(theme: Optional[dict[str, str]] = None) -> dict[str, ParagraphStyle]:
    palette = _pdf_palette(theme)
    styles = getSampleStyleSheet()
    return {
        "report_title": ParagraphStyle(
            "PdfReportTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=21,
            textColor=colors.HexColor(palette["text"]),
            spaceAfter=3,
        ),
        "section_title": ParagraphStyle(
            "PdfSectionTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=14,
            leading=17,
            textColor=colors.HexColor(palette["primary"]),
            spaceAfter=3,
        ),
        "chart_title": ParagraphStyle(
            "PdfChartTitle",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=10.5,
            leading=13,
            textColor=colors.HexColor(palette["primary"]),
            spaceAfter=4,
        ),
        "meta": ParagraphStyle(
            "PdfMeta",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor(palette["muted"]),
        ),
        "body": ParagraphStyle(
            "PdfBody",
            parent=styles["Normal"],
            fontSize=8.2,
            leading=11,
            textColor=colors.HexColor(palette["text"]),
        ),
        "table_header": ParagraphStyle(
            "PdfTableHeader",
            parent=styles["Normal"],
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            fontSize=7,
            leading=8.5,
            textColor=colors.white,
        ),
        "table_cell": ParagraphStyle(
            "PdfTableCell",
            parent=styles["Normal"],
            fontSize=6.5,
            leading=8,
            textColor=colors.HexColor(palette["text"]),
        ),
        "table_number": ParagraphStyle(
            "PdfTableNumber",
            parent=styles["Normal"],
            alignment=TA_RIGHT,
            fontSize=6.5,
            leading=8,
            textColor=colors.HexColor(palette["text"]),
        ),
        "table_date": ParagraphStyle(
            "PdfTableDate",
            parent=styles["Normal"],
            alignment=TA_CENTER,
            fontSize=6.5,
            leading=8,
            textColor=colors.HexColor(palette["text"]),
        ),
    }


def _rounded_table(data, style_commands, **kwargs) -> Table:
    try:
        table = Table(data, cornerRadii=[PDF_LAYOUT["card_radius"]] * 4, **kwargs)
    except TypeError:
        table = Table(data, **kwargs)
    table.setStyle(TableStyle(style_commands))
    return table

EXPORT_COLUMN_LABELS = {
    "id": "ID",
    "data_hora": "Data/hora",
    "ingrediente_id": "ID ingrediente",
    "ingrediente": "Ingrediente",
    "categoria": "Categoria",
    "unidade": "Unidade",
    "quantidade": "Quantidade",
    "nome": "Nome",
    "cnpj": "CNPJ",
    "email": "Email",
    "telefone": "Telefone",
    "prazo_medio_entrega_dias": "Prazo medio entrega (dias)",
    "itens_fornecidos": "Itens fornecidos",
    "preco_medio": "Preco medio",
    "pedido_id": "ID pedido",
    "data_pedido": "Data pedido",
    "fornecedor_id": "ID fornecedor",
    "fornecedor": "Fornecedor",
    "valor_total": "Valor total",
    "status": "Status",
    "data_prevista": "Data prevista",
    "indicador": "Indicador",
    "detalhe": "Detalhe",
    "comparacao": "Comparacao",
    "direcao": "Direcao",
    "estoque_atual": "Estoque atual",
    "uso_dia": "Uso/dia",
    "cobertura_dias": "Cobertura (dias)",
    "sugestao_compra": "Sugestao de compra",
    "posicao": "#",
    "receita": "Receita",
    "unidades_vendidas": "Unidades vendidas",
    "faturamento": "Faturamento",
    "periodo": "Periodo",
    "valor": "Valor",
    "data": "Data",
    "estoque": "Estoque",
    "vendas": "Vendas",
}


def _normalize_export_format(format_value: str) -> str:
    normalized = format_value.strip().lower()
    normalized = {"xlsx": "excel", "xls": "excel", "yml": "yaml"}.get(
        normalized,
        normalized,
    )
    if normalized not in EXPORT_FORMATS:
        raise HTTPException(
            status_code=400,
            detail="Formato invalido. Use pdf, excel, csv, json, xml ou yaml.",
        )
    return normalized


def _stringify_export_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, float):
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value)


def _export_response(
    rows: list[dict],
    filename: str,
    format_value: str,
    title: str,
    columns: Optional[list[str]] = None,
    theme_id: Optional[str] = None,
) -> Response:
    export_format = _normalize_export_format(format_value)
    media_type, extension = EXPORT_FORMATS[export_format]
    theme = _export_theme(theme_id)
    content = _serialize_export(rows, export_format, title, columns, theme)
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}.{extension}"'},
    )


def _serialize_export(
    rows: list[dict],
    export_format: str,
    title: str,
    columns: Optional[list[str]] = None,
    theme: Optional[dict[str, str]] = None,
):
    theme = theme or _export_theme(None)
    if export_format == "csv":
        return _serialize_csv(rows, columns)
    if export_format == "json":
        return json.dumps(rows, ensure_ascii=False, indent=2, default=str)
    if export_format == "xml":
        return _serialize_xml(rows)
    if export_format == "yaml":
        return _serialize_yaml(rows)
    if export_format == "excel":
        return _serialize_excel(rows, title, columns, theme)
    if export_format == "pdf":
        return _serialize_pdf(rows, title, columns, theme)
    raise AssertionError(export_format)


def _export_columns(rows: list[dict], columns: Optional[list[str]]) -> list[str]:
    return columns or (list(rows[0].keys()) if rows else [])


def _export_headers(columns: list[str]) -> list[str]:
    return [EXPORT_COLUMN_LABELS.get(column, column.replace("_", " ").title()) for column in columns]


def _export_cell_value(value):
    if value is None:
        return None
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _serialize_csv(rows: list[dict], columns: Optional[list[str]]) -> str:
    output = io.StringIO()
    export_columns = _export_columns(rows, columns)
    writer = csv.DictWriter(output, fieldnames=export_columns)
    writer.writeheader()
    for row in rows:
        writer.writerow({key: _stringify_export_value(value) for key, value in row.items()})
    return output.getvalue()


def _serialize_xml(rows: list[dict]) -> str:
    root = Element("export")
    for row in rows:
        row_el = SubElement(root, "row")
        for key, value in row.items():
            field = SubElement(row_el, key)
            field.text = _stringify_export_value(value)
    return tostring(root, encoding="unicode", xml_declaration=True)


def _serialize_yaml(rows: list[dict]) -> str:
    if not rows:
        return "[]\n"
    lines: list[str] = []
    for row in rows:
        lines.append("-")
        for key, value in row.items():
            text = _stringify_export_value(value).replace("\\", "\\\\").replace('"', '\\"')
            lines.append(f'  {key}: "{text}"')
    return "\n".join(lines) + "\n"


def _xml_escape(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _numeric_export_columns(rows: list[dict], columns: list[str]) -> list[str]:
    return [
        column
        for column in columns
        if any(isinstance(row.get(column), (int, float)) for row in rows)
    ]


def _style_export_worksheet(worksheet, columns: list[str], theme: dict[str, str]) -> None:
    header_fill = PatternFill("solid", fgColor=_hex(theme["primary"]))
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color=_hex(theme["grid"]))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    worksheet.freeze_panes = "A2"
    if columns:
        last_column = get_column_letter(len(columns))
        worksheet.auto_filter.ref = f"A1:{last_column}{max(1, worksheet.max_row)}"

    for row_index, row_cells in enumerate(worksheet.iter_rows(min_row=2), start=2):
        fill = PatternFill("solid", fgColor=_hex(theme["soft"] if row_index % 2 == 0 else "#FFFFFF"))
        for cell in row_cells:
            cell.fill = fill
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top")
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="top")
                cell.number_format = '#,##0.00'

    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else _stringify_export_value(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 54)


def _add_cover_sheet(workbook: Workbook, title: str, row_count: int, theme: dict[str, str]) -> None:
    worksheet = workbook.create_sheet("Capa", 0)
    worksheet.sheet_view.showGridLines = False
    worksheet["B2"] = "Maestro"
    worksheet["B2"].font = Font(size=14, bold=True, color=_hex(theme["accent"]))
    worksheet["B3"] = title
    worksheet["B3"].font = Font(size=24, bold=True, color=_hex(theme["primary"]))
    worksheet["B5"] = "Gerado em"
    worksheet["C5"] = _now_recife().strftime("%d/%m/%Y %H:%M")
    worksheet["B6"] = "Registros"
    worksheet["C6"] = row_count
    worksheet["B8"] = "Observacao"
    worksheet["C8"] = "Use os filtros da aba Dados e confira o Resumo para uma leitura executiva."
    for cell in ("B5", "B6", "B8"):
        worksheet[cell].font = Font(bold=True, color=_hex(theme["muted"]))
    for column in ("B", "C", "D"):
        worksheet.column_dimensions[column].width = 24
    worksheet["C8"].alignment = Alignment(wrap_text=True)


def _add_summary_sheet(
    workbook: Workbook,
    rows: list[dict],
    columns: list[str],
    theme: dict[str, str],
) -> None:
    worksheet = workbook.create_sheet("Resumo")
    worksheet.sheet_view.showGridLines = False
    numeric_columns = _numeric_export_columns(rows, columns)
    worksheet.append(["Metrica", "Valor"])
    worksheet.append(["Registros", len(rows)])
    worksheet.append(["Colunas", len(columns)])
    for column in numeric_columns:
        values = [_as_float(row.get(column)) for row in rows if row.get(column) is not None]
        worksheet.append([f"Soma - {_export_headers([column])[0]}", sum(values)])
        worksheet.append([f"Media - {_export_headers([column])[0]}", sum(values) / max(1, len(values))])
    _style_export_worksheet(worksheet, ["metrica", "valor"], theme)
    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 18


def _add_numeric_chart(worksheet, rows: list[dict], columns: list[str], theme: dict[str, str]) -> None:
    numeric_columns = _numeric_export_columns(rows, columns)
    if not numeric_columns or not rows:
        return
    numeric_index = columns.index(numeric_columns[0]) + 1
    label_index = 1
    for candidate in ("ingrediente", "fornecedor", "nome", "categoria", "receita", "data_pedido", "data_hora"):
        if candidate in columns:
            label_index = columns.index(candidate) + 1
            break
    chart = ExcelBarChart()
    chart.title = f"{_export_headers([numeric_columns[0]])[0]} por registro"
    chart.style = 10
    chart.height = 8
    chart.width = 18
    max_row = min(worksheet.max_row, 16)
    data = Reference(worksheet, min_col=numeric_index, min_row=1, max_row=max_row)
    cats = Reference(worksheet, min_col=label_index, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.majorGridlines = None
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = _hex(theme["primary"])
        chart.series[0].graphicalProperties.line.solidFill = _hex(theme["primary"])
    worksheet.add_chart(chart, f"{get_column_letter(len(columns) + 2)}2")


def _serialize_excel(rows: list[dict], title: str, columns: Optional[list[str]], theme: dict[str, str]) -> bytes:
    export_columns = _export_columns(rows, columns)
    headers = _export_headers(export_columns)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Maestro"
    workbook.properties.title = title
    _add_cover_sheet(workbook, title, len(rows), theme)
    _add_summary_sheet(workbook, rows, export_columns, theme)

    worksheet = workbook.create_sheet(_safe_sheet_title(title))
    worksheet.append(headers)
    for row in rows:
        worksheet.append([_export_cell_value(row.get(column)) for column in export_columns])
    _style_export_worksheet(worksheet, export_columns, theme)
    _add_numeric_chart(worksheet, rows, export_columns, theme)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_header_block(
    title: str,
    title_style: ParagraphStyle,
    meta_style: ParagraphStyle,
    theme: Optional[dict[str, str]] = None,
) -> Table:
    theme = theme or _export_theme(None)
    palette = _pdf_palette(theme)
    generated_at = _now_recife().strftime("%d/%m/%Y %H:%M")
    title_cell = [
        Paragraph(PDF_PROJECT_NAME, meta_style),
        Paragraph(_xml_escape(title), title_style),
        Paragraph(f"{PDF_PROJECT_CONTEXT} | Gerado em {generated_at}", meta_style),
    ]
    logo_style = ParagraphStyle(
        "PdfFallbackLogo",
        parent=meta_style,
        fontName="Helvetica-Bold",
        fontSize=10,
        textColor=colors.HexColor(palette["primary"]),
    )
    project_logo = _pdf_logo_flowable(
        PDF_PROJECT_LOGO_PATH,
        14 * mm,
        14 * mm,
        PDF_PROJECT_NAME,
        logo_style,
        fill_color=colors.HexColor(palette["primary"]),
    )
    client_logo = _pdf_logo_flowable(
        PDF_CLIENT_LOGO_PATH,
        16 * mm,
        16 * mm,
        PDF_CLIENT_NAME,
        logo_style,
    )

    return _rounded_table(
        [[project_logo, title_cell, client_logo]],
        [
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["card"])),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor(palette["card_border"])),
            ("LINEBELOW", (0, 0), (-1, -1), 1.3, colors.HexColor(palette["primary"])),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (2, 0), (2, 0), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ],
        colWidths=[21 * mm, 215 * mm, 22 * mm],
        hAlign="LEFT",
    )


def _pdf_table(
    rows: list[dict],
    columns: list[str],
    header_style: ParagraphStyle,
    cell_style: ParagraphStyle,
    number_style: ParagraphStyle,
    date_style: ParagraphStyle,
    available_width: float,
    theme: Optional[dict[str, str]] = None,
) -> Table:
    theme = theme or _export_theme(None)
    palette = _pdf_palette(theme)
    headers = [Paragraph(_xml_escape(header), header_style) for header in _export_headers(columns)]
    data = [headers]
    numeric_columns = {
        "quantidade",
        "valor_total",
        "preco_medio",
        "itens_fornecidos",
        "prazo_medio_entrega_dias",
        "estoque_atual",
        "uso_dia",
        "cobertura_dias",
        "sugestao_compra",
        "posicao",
        "unidades_vendidas",
        "faturamento",
        "valor",
        "estoque",
        "vendas",
    } | {
        column
        for column in columns
        if any(isinstance(row.get(column), (int, float)) for row in rows)
    }
    date_columns = {
        column
        for column in columns
        if "data" in column or any(isinstance(row.get(column), (date, datetime)) for row in rows)
    }

    if rows:
        for row in rows:
            data.append(
                [
                    Paragraph(
                        _xml_escape(_stringify_export_value(row.get(column))),
                        (
                            number_style
                            if column in numeric_columns
                            else date_style
                            if column in date_columns
                            else cell_style
                        ),
                    )
                    for column in columns
                ]
            )
    else:
        data.append([Paragraph("Nenhum registro encontrado.", cell_style)] + [""] * (len(columns) - 1))

    highlight_rows = []
    for row_index, row in enumerate(rows, start=1):
        values = " ".join(_stringify_export_value(value).lower() for value in row.values())
        if any(token in values for token in ("total", "media", "média")):
            highlight_rows.append(row_index)

    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(palette["primary"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, 0), 0.9, colors.HexColor(palette["accent"])),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(palette["table_alt"])]),
        ("BOX", (0, 0), (-1, -1), 0.55, colors.HexColor(palette["card_border"])),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor(palette["grid"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for row_index in highlight_rows:
        style_commands.extend(
            [
                ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor(palette["accent_soft"])),
                ("LINEABOVE", (0, row_index), (-1, row_index), 0.7, colors.HexColor(palette["accent"])),
            ]
        )
    if not rows and columns:
        style_commands.append(("SPAN", (0, 1), (-1, 1)))

    return _rounded_table(
        data,
        style_commands,
        colWidths=_pdf_column_widths(rows, columns, available_width),
        repeatRows=1,
        splitByRow=1,
        hAlign="LEFT",
    )


def _pdf_column_widths(rows: list[dict], columns: list[str], available_width: float) -> list[float]:
    if not columns:
        return []
    weights: list[float] = []
    for column, header in zip(columns, _export_headers(columns)):
        sample_lengths = [
            len(_stringify_export_value(row.get(column)))
            for row in rows[:80]
        ]
        content_length = max(sample_lengths, default=0)
        weights.append(max(len(header), min(content_length, 26), 5))
    total_weight = sum(weights) or 1
    min_width = 20 * mm
    widths = [max(min_width, available_width * weight / total_weight) for weight in weights]
    width_total = sum(widths)
    if width_total > available_width:
        scale = available_width / width_total
        widths = [width * scale for width in widths]
    return widths


def _draw_pdf_page_frame(canvas, document):
    theme = getattr(document, "export_theme", _export_theme(None))
    palette = _pdf_palette(theme)
    page_width, page_height = document.pagesize
    generated_at = getattr(document, "generated_at_text", _now_recife().strftime("%d/%m/%Y %H:%M"))
    report_title = getattr(document, "report_title", "Relatorio")
    canvas.saveState()

    header_top = page_height - 7 * mm
    header_bottom = page_height - 25 * mm
    canvas.setFillColor(colors.HexColor(palette["card"]))
    canvas.rect(0, header_bottom, page_width, header_top - header_bottom, stroke=0, fill=1)
    canvas.setFillColor(colors.HexColor(palette["primary"]))
    canvas.rect(0, header_bottom - 1.5 * mm, page_width, 1.5 * mm, stroke=0, fill=1)
    canvas.setStrokeColor(colors.HexColor(palette["grid"]))
    canvas.setLineWidth(0.35)
    canvas.line(document.leftMargin, header_bottom, page_width - document.rightMargin, header_bottom)

    _draw_pdf_logo(
        canvas,
        PDF_PROJECT_LOGO_PATH,
        document.leftMargin,
        page_height - 21 * mm,
        13 * mm,
        13 * mm,
        fill_color=colors.HexColor(palette["primary"]),
    )
    _draw_pdf_logo(
        canvas,
        PDF_CLIENT_LOGO_PATH,
        page_width - document.rightMargin - 17 * mm,
        page_height - 22 * mm,
        16 * mm,
        16 * mm,
    )

    text_x = document.leftMargin + 19 * mm
    canvas.setFont("Helvetica-Bold", 8.5)
    canvas.setFillColor(colors.HexColor(palette["primary"]))
    canvas.drawString(text_x, page_height - 13 * mm, PDF_PROJECT_NAME)
    canvas.setFont("Helvetica-Bold", 7)
    canvas.setFillColor(colors.HexColor(palette["accent"]))
    canvas.drawString(text_x, page_height - 17 * mm, PDF_PROJECT_CONTEXT)
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor(palette["muted"]))
    canvas.drawString(text_x, page_height - 21 * mm, str(report_title)[:95])

    footer_y = 11 * mm
    canvas.setStrokeColor(colors.HexColor(palette["grid"]))
    canvas.setLineWidth(0.45)
    canvas.line(document.leftMargin, footer_y, page_width - document.rightMargin, footer_y)
    canvas.setFillColor(colors.HexColor(palette["accent"]))
    canvas.roundRect(document.leftMargin, footer_y - 4.8 * mm, 4 * mm, 4 * mm, 1.4 * mm, stroke=0, fill=1)
    canvas.setFont("Helvetica-Bold", 7)
    canvas.setFillColor(colors.HexColor(palette["primary"]))
    canvas.drawString(document.leftMargin + 6 * mm, footer_y - 3.4 * mm, PDF_SIGNATURE)
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor(palette["footer_text"]))
    canvas.drawString(
        document.leftMargin + 36 * mm,
        footer_y - 3.4 * mm,
        f"{PDF_GENERATOR_NOTE} | Gerado em {generated_at}",
    )
    canvas.drawRightString(
        page_width - document.rightMargin,
        footer_y - 3.4 * mm,
        f"Pagina {canvas.getPageNumber()}",
    )
    canvas.restoreState()


def _serialize_pdf(rows: list[dict], title: str, columns: Optional[list[str]], theme: dict[str, str]) -> bytes:
    export_columns = _export_columns(rows, columns)
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=PDF_PAGE_SIZE,
        leftMargin=PDF_LAYOUT["page_margin_x"],
        rightMargin=PDF_LAYOUT["page_margin_x"],
        topMargin=PDF_LAYOUT["page_margin_top"],
        bottomMargin=PDF_LAYOUT["page_margin_bottom"],
        title=title,
    )
    document.export_theme = theme
    document.report_title = title
    document.generated_at_text = _now_recife().strftime("%d/%m/%Y %H:%M")
    styles = _pdf_styles(theme)

    story = [
        _pdf_header_block(title, styles["report_title"], styles["meta"], theme),
        Spacer(1, PDF_LAYOUT["section_gap"]),
    ]
    if export_columns:
        story.append(
            _pdf_table(
                rows,
                export_columns,
                styles["table_header"],
                styles["table_cell"],
                styles["table_number"],
                styles["table_date"],
                document.width,
                theme,
            )
        )
    else:
        story.append(Paragraph("Nenhum registro encontrado.", styles["body"]))

    document.build(story, onFirstPage=_draw_pdf_page_frame, onLaterPages=_draw_pdf_page_frame)
    return output.getvalue()


DASHBOARD_PDF_TABLE_ROWS = 18


def _safe_sheet_title(title: str) -> str:
    invalid = "[]:*?/\\"
    cleaned = "".join("_" if char in invalid else char for char in title)
    return cleaned[:31] or "Dados"


def _format_export_number(value: float, digits: int = 2) -> float:
    return round(_as_float(value), digits)


def _dashboard_metric_rows(metrics: list[DashboardNamedMetric]) -> list[dict]:
    return [
        {
            "periodo": metric.label,
            "valor": _format_export_number(metric.value),
        }
        for metric in metrics
    ]


def _dashboard_history_rows(
    stock: list[DashboardHistoryPoint],
    sales: list[DashboardHistoryPoint],
) -> list[dict]:
    monthly: dict[str, dict] = {}
    for item in stock:
        key = item.date.strftime("%Y-%m")
        month = monthly.setdefault(key, {"data": key, "estoque_values": [], "vendas": 0.0})
        month["estoque_values"].append(_as_float(item.value))
    for item in sales:
        key = item.date.strftime("%Y-%m")
        month = monthly.setdefault(key, {"data": key, "estoque_values": [], "vendas": 0.0})
        month["vendas"] += _as_float(item.value)

    return [
        {
            "data": key,
            "periodo": _month_year_label(date(int(key[:4]), int(key[5:7]), 1)),
            "estoque": _format_export_number(
                sum(item["estoque_values"]) / max(1, len(item["estoque_values"]))
            ),
            "vendas": _format_export_number(item["vendas"]),
        }
        for key, item in sorted(monthly.items())
    ]


def _dashboard_alert_rows(alerts: list[DashboardAlert]) -> list[dict]:
    return [
        {
            "ingrediente_id": alert.ingredient_id,
            "ingrediente": alert.name,
            "categoria": alert.category,
            "unidade": alert.unit,
            "estoque_atual": _format_export_number(alert.current_qty),
            "uso_dia": _format_export_number(alert.avg_daily_output),
            "cobertura_dias": _format_export_number(alert.coverage_days),
            "sugestao_compra": _format_export_number(alert.suggested_qty),
            "status": alert.severity,
        }
        for alert in alerts
    ]


def _dashboard_rank_rows(groups, label: str) -> list[dict]:
    rows: list[dict] = []
    for group in groups:
        for index, item in enumerate(group.items, start=1):
            rows.append(
                {
                    "posicao": index,
                    "tipo_ranking": label,
                    "nome": item.name,
                    "categoria": getattr(item, "category", None),
                    "unidade": item.unit or group.unit,
                    "valor": _format_export_number(item.value),
                }
            )
    return rows


def _dashboard_recipe_rows(recipes: list[DashboardRecipeItem]) -> list[dict]:
    return [
        {
            "posicao": index,
            "receita": recipe.name,
            "unidades_vendidas": _format_export_number(recipe.quantity),
            "faturamento": _format_export_number(recipe.revenue),
        }
        for index, recipe in enumerate(recipes, start=1)
    ]


def _dashboard_kpi_rows(kpis: list[DashboardKpi]) -> list[dict]:
    return [
        {
            "indicador": item.label,
            "valor": item.value,
            "detalhe": item.detail,
            "comparacao": item.trend_label,
            "direcao": item.trend_direction,
        }
        for item in kpis
    ]


def _dashboard_export_tables(
    db: Session,
    category_ids: Optional[list[str]] = None,
    days: int = 90,
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = None,
    years: Optional[list[int]] = None,
    month_numbers: Optional[list[int]] = None,
    event_types: Optional[list[str]] = None,
) -> tuple[list[dict], dict]:
    event_dates = _event_dates_for_types(db, event_types)
    common_kwargs = {
        "category_ids": category_ids,
        "days": days,
        "all_period": all_period,
        "date_from": date_from,
        "date_to": date_to,
        "month_keys": month_keys,
        "years": years,
        "months": month_numbers,
        "event_dates": event_dates,
    }

    kpis = _dashboard_kpis(db, **common_kwargs)
    alerts = _dashboard_alerts(db, limit=10000, **common_kwargs)
    stock_history = get_dashboard_estoque_historico(
        category_ids=category_ids,
        days=days,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        month_numbers=month_numbers,
        event_types=event_types,
        db=db,
    )
    sales_history = get_dashboard_vendas_historico(
        category_ids=category_ids,
        days=days,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        month_numbers=month_numbers,
        event_types=event_types,
        db=db,
    )
    revenue = get_dashboard_faturamento_resumo(
        months=12,
        category_ids=category_ids,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        month_numbers=month_numbers,
        event_types=event_types,
        db=db,
    )
    recipes = get_dashboard_receitas_ranking(
        category_ids=category_ids,
        days=days,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        month_numbers=month_numbers,
        event_types=event_types,
        limit=10000,
        db=db,
    )
    output_category_groups = _output_category_groups(db, desc, limit=10000, **common_kwargs)
    stock_product_groups = _stock_product_groups(db, desc, limit=10000, category_ids=category_ids)
    output_product_groups = _output_product_groups(db, desc, limit=10000, **common_kwargs)

    tables = [
        {
            "title": "Indicadores",
            "rows": _dashboard_kpi_rows(kpis),
            "columns": ["indicador", "valor", "detalhe", "comparacao", "direcao"],
        },
        {
            "title": "Alertas operacionais",
            "rows": _dashboard_alert_rows(alerts),
            "columns": [
                "ingrediente",
                "categoria",
                "unidade",
                "estoque_atual",
                "uso_dia",
                "cobertura_dias",
                "sugestao_compra",
                "status",
            ],
        },
        {
            "title": "Ranking de estoque",
            "rows": _dashboard_rank_rows(stock_product_groups, "Estoque por ingrediente"),
            "columns": ["posicao", "nome", "categoria", "unidade", "valor"],
        },
        {
            "title": "Ranking de uso",
            "rows": _dashboard_rank_rows(output_product_groups, "Uso por ingrediente"),
            "columns": ["posicao", "nome", "categoria", "unidade", "valor"],
        },
        {
            "title": "Receitas mais vendidas",
            "rows": _dashboard_recipe_rows(recipes),
            "columns": ["posicao", "receita", "unidades_vendidas", "faturamento"],
        },
    ]
    chart_data = {
        "history": _dashboard_history_rows(stock_history, sales_history),
        "revenue": _dashboard_metric_rows(revenue.monthly),
        "categories": _dashboard_rank_rows(output_category_groups, "Uso por categoria"),
    }
    return tables, chart_data


def _sample_chart_rows(rows: list[dict], limit: int) -> list[dict]:
    if len(rows) <= limit:
        return rows
    step = (len(rows) - 1) / max(1, limit - 1)
    return [rows[round(index * step)] for index in range(limit)]


def _line_chart_drawing(rows: list[dict], width: float, height: float, theme: Optional[dict[str, str]] = None) -> Drawing:
    theme = theme or _export_theme(None)
    palette = _pdf_palette(theme)
    drawing = Drawing(width, height)
    left = 44
    right = 44
    bottom = 31
    top = 25
    chart_width = width - left - right
    chart_height = height - bottom - top
    sampled = _sample_chart_rows(rows, 18)
    if not sampled:
        drawing.add(String(width / 2 - 45, height / 2, "Sem dados para exibir", fontSize=9, fillColor=colors.HexColor(palette["muted"])))
        return drawing

    series = [
        ("estoque", "Estoque medio", colors.HexColor(palette["accent"]), "left"),
        ("vendas", "Vendas totais", colors.HexColor(palette["primary"]), "right"),
    ]
    drawing.add(DrawingLine(left, bottom, left, bottom + chart_height, strokeColor=colors.HexColor(palette["grid"]), strokeWidth=0.65))
    drawing.add(DrawingLine(left, bottom, left + chart_width, bottom, strokeColor=colors.HexColor(palette["grid"]), strokeWidth=0.65))
    drawing.add(DrawingLine(left + chart_width, bottom, left + chart_width, bottom + chart_height, strokeColor=colors.HexColor(palette["grid"]), strokeWidth=0.65))

    max_values = {
        key: max((_as_float(row.get(key)) for row in sampled), default=0) or 1
        for key, _label, _color, _axis in series
    }
    for tick in range(5):
        y = bottom + chart_height * tick / 4
        drawing.add(DrawingLine(left, y, left + chart_width, y, strokeColor=colors.HexColor(palette["grid"]), strokeWidth=0.22))
        left_label = _stringify_export_value(max_values["estoque"] * tick / 4)
        right_label = _stringify_export_value(max_values["vendas"] * tick / 4)
        drawing.add(String(left - 35, y - 2, left_label, fontSize=6.2, fillColor=colors.HexColor(palette["muted"])))
        drawing.add(String(left + chart_width + 5, y - 2, right_label, fontSize=6.2, fillColor=colors.HexColor(palette["muted"])))

    totals = {
        "estoque": sum(_as_float(row.get("estoque")) for row in sampled) / max(1, len(sampled)),
        "vendas": sum(_as_float(row.get("vendas")) for row in sampled),
    }
    legend_x = left + chart_width - 190
    for index, (key, label, color, _) in enumerate(series):
        x = legend_x + index * 98
        drawing.add(Rect(x, height - 14, 8, 8, fillColor=color, strokeColor=color))
        drawing.add(
            String(
                x + 12,
                height - 12,
                f"{label}: {_stringify_export_value(totals[key])}",
                fontSize=7.2,
                fillColor=color,
            )
        )

    for key, _label, color, _axis in series:
        values = [_as_float(row.get(key)) for row in sampled]
        max_value = max_values[key]
        points = []
        for index, value in enumerate(values):
            x = left + (chart_width * index / max(1, len(sampled) - 1))
            y = bottom + (value / max_value) * chart_height
            points.append((x, y))
        for start, end in zip(points, points[1:]):
            drawing.add(DrawingLine(start[0], start[1], end[0], end[1], strokeColor=color, strokeWidth=2.2))
        for x, y in points:
            drawing.add(Rect(x - 1.5, y - 1.5, 3, 3, fillColor=color, strokeColor=color))

    for index, row in enumerate(sampled):
        label = str(row.get("periodo") or row.get("data", ""))
        x = left + (chart_width * index / max(1, len(sampled) - 1))
        drawing.add(String(x - 10, bottom - 14, label, fontSize=6.3, fillColor=colors.HexColor(palette["muted"])))
    return drawing


def _bar_chart_drawing(rows: list[dict], width: float, height: float, value_key: str = "valor", theme: Optional[dict[str, str]] = None) -> Drawing:
    theme = theme or _export_theme(None)
    palette = _pdf_palette(theme)
    drawing = Drawing(width, height)
    left = 34
    bottom = 29
    chart_width = width - 56
    chart_height = height - 50
    items = rows[:10]
    if not items:
        drawing.add(String(width / 2 - 45, height / 2, "Sem dados para exibir", fontSize=9, fillColor=colors.HexColor(palette["muted"])))
        return drawing
    max_value = max(_as_float(item.get(value_key)) for item in items) or 1
    gap = 4
    bar_width = max(8, (chart_width - gap * (len(items) - 1)) / len(items))
    drawing.add(DrawingLine(left, bottom, left, bottom + chart_height, strokeColor=colors.HexColor(palette["grid"]), strokeWidth=0.65))
    drawing.add(DrawingLine(left, bottom, left + chart_width, bottom, strokeColor=colors.HexColor(palette["grid"]), strokeWidth=0.65))
    for tick in range(1, 5):
        y = bottom + chart_height * tick / 4
        drawing.add(DrawingLine(left, y, left + chart_width, y, strokeColor=colors.HexColor(palette["grid"]), strokeWidth=0.22))
    for index, item in enumerate(items):
        value = _as_float(item.get(value_key))
        bar_height = (value / max_value) * chart_height
        x = left + index * (bar_width + gap)
        bar_color = colors.HexColor(palette["primary"] if index % 2 == 0 else palette["accent"])
        drawing.add(Rect(x, bottom, bar_width, bar_height, fillColor=bar_color, strokeColor=bar_color))
        drawing.add(String(x, bottom + bar_height + 4, _stringify_export_value(value), fontSize=6.1, fillColor=colors.HexColor(palette["muted"])))
        label = str(item.get("periodo") or item.get("nome") or item.get("receita") or "")[:12]
        drawing.add(String(x, bottom - 13, label, fontSize=6.2, fillColor=colors.HexColor(palette["muted"])))
    return drawing


def _dashboard_pdf_paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(_xml_escape(text), style)


def _pdf_card(
    content: list,
    available_width: float,
    theme: Optional[dict[str, str]] = None,
    padding: Optional[float] = None,
) -> Table:
    palette = _pdf_palette(theme)
    inner_padding = PDF_LAYOUT["card_padding"] if padding is None else padding
    return _rounded_table(
        [[content]],
        [
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["card"])),
            ("BOX", (0, 0), (-1, -1), 0.65, colors.HexColor(palette["card_border"])),
            ("LINEBEFORE", (0, 0), (0, -1), 2.3, colors.HexColor(palette["accent"])),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), inner_padding),
            ("RIGHTPADDING", (0, 0), (-1, -1), inner_padding),
            ("TOPPADDING", (0, 0), (-1, -1), inner_padding),
            ("BOTTOMPADDING", (0, 0), (-1, -1), inner_padding),
        ],
        colWidths=[available_width],
        hAlign="LEFT",
    )


def _pdf_chart_card(
    title: str,
    description: str,
    drawing: Drawing,
    available_width: float,
    styles: dict[str, ParagraphStyle],
    theme: Optional[dict[str, str]] = None,
) -> Table:
    return _pdf_card(
        [
            Paragraph(_xml_escape(title), styles["chart_title"]),
            drawing,
            Spacer(1, 3 * mm),
            Paragraph(_xml_escape(description), styles["body"]),
        ],
        available_width,
        theme,
    )


def _add_dashboard_chart_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict],
    columns: list[str],
    label_column: str,
    value_columns: list[str],
    theme: dict[str, str],
    chart_kind: str = "bar",
) -> None:
    worksheet = workbook.create_sheet(_safe_sheet_title(title))
    worksheet.append(_export_headers(columns))
    for row in rows:
        worksheet.append([_export_cell_value(row.get(column)) for column in columns])
    _style_export_worksheet(worksheet, columns, theme)
    if not rows or not value_columns:
        return
    chart = ExcelLineChart() if chart_kind == "line" else ExcelBarChart()
    chart.title = title
    chart.style = 10
    chart.height = 9
    chart.width = 22
    max_row = min(worksheet.max_row, 22)
    for value_column in value_columns:
        if value_column not in columns:
            continue
        col_index = columns.index(value_column) + 1
        data = Reference(worksheet, min_col=col_index, min_row=1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
    label_index = columns.index(label_column) + 1
    chart.set_categories(Reference(worksheet, min_col=label_index, min_row=2, max_row=max_row))
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = _hex(theme["primary"])
        chart.series[0].graphicalProperties.line.solidFill = _hex(theme["primary"])
    if len(chart.series) > 1:
        chart.series[1].graphicalProperties.solidFill = _hex(theme["accent"])
        chart.series[1].graphicalProperties.line.solidFill = _hex(theme["accent"])
    worksheet.add_chart(chart, f"{get_column_letter(len(columns) + 2)}2")


def _serialize_dashboard_excel(tables: list[dict], chart_data: dict, theme: dict[str, str]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Maestro"
    workbook.properties.title = "Dashboard Maestro"
    _add_cover_sheet(workbook, "Dashboard Maestro", sum(len(table["rows"]) for table in tables), theme)

    for table in tables:
        worksheet = workbook.create_sheet(_safe_sheet_title(table["title"]))
        columns = table["columns"]
        worksheet.append(_export_headers(columns))
        for row in table["rows"]:
            worksheet.append([_export_cell_value(row.get(column)) for column in columns])
        _style_export_worksheet(worksheet, columns, theme)
        _add_numeric_chart(worksheet, table["rows"], columns, theme)

    _add_dashboard_chart_sheet(
        workbook,
        "Grafico estoque vendas",
        chart_data["history"],
        ["periodo", "estoque", "vendas"],
        "periodo",
        ["estoque", "vendas"],
        theme,
        "line",
    )
    _add_dashboard_chart_sheet(
        workbook,
        "Grafico faturamento",
        chart_data["revenue"],
        ["periodo", "valor"],
        "periodo",
        ["valor"],
        theme,
        "bar",
    )
    _add_dashboard_chart_sheet(
        workbook,
        "Grafico categorias",
        chart_data["categories"],
        ["nome", "valor"],
        "nome",
        ["valor"],
        theme,
        "bar",
    )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _serialize_dashboard_pdf(tables: list[dict], chart_data: dict, theme: dict[str, str]) -> bytes:
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=PDF_PAGE_SIZE,
        leftMargin=PDF_LAYOUT["page_margin_x"],
        rightMargin=PDF_LAYOUT["page_margin_x"],
        topMargin=PDF_LAYOUT["page_margin_top"],
        bottomMargin=PDF_LAYOUT["page_margin_bottom"],
        title="Dashboard Maestro",
    )
    document.export_theme = theme
    document.report_title = "Dashboard Maestro"
    document.generated_at_text = _now_recife().strftime("%d/%m/%Y %H:%M")
    styles = _pdf_styles(theme)

    story = [
        _pdf_header_block("Dashboard", styles["report_title"], styles["meta"], theme),
        Spacer(1, PDF_LAYOUT["section_gap"]),
        _pdf_chart_card(
            "Estoque medio x vendas totais",
            "O grafico agrupa o periodo filtrado por mes e compara estoque medio mensal com vendas totais mensais. "
            "A linha secundaria usa o eixo da esquerda para estoque, enquanto a linha principal usa o eixo da direita para vendas. "
            "A leitura conjunta ajuda a identificar meses em que as saidas cresceram mais rapido do que a reposicao.",
            _line_chart_drawing(chart_data["history"], document.width - 18 * mm, 105 * mm, theme),
            document.width,
            styles,
            theme,
        ),
        PageBreak(),
        _pdf_header_block("Graficos secundarios", styles["report_title"], styles["meta"], theme),
        Spacer(1, PDF_LAYOUT["section_gap"]),
        Table(
            [
                [
                    _pdf_chart_card(
                        "Faturamento",
                        "Mostra a receita estimada por periodo, facilitando a comparacao entre meses.",
                        _bar_chart_drawing(chart_data["revenue"], document.width / 2 - 19 * mm, 78 * mm, theme=theme),
                        document.width / 2 - 5 * mm,
                        styles,
                        theme,
                    ),
                    _pdf_chart_card(
                        "Categorias mais vendidas",
                        "Apresenta as categorias com maior uso no periodo, indicando concentracao de demanda.",
                        _bar_chart_drawing(chart_data["categories"], document.width / 2 - 19 * mm, 78 * mm, theme=theme),
                        document.width / 2 - 5 * mm,
                        styles,
                        theme,
                    ),
                ],
            ],
            colWidths=[document.width / 2 - 2.5 * mm, document.width / 2 - 2.5 * mm],
            hAlign="LEFT",
            style=[
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ],
        ),
    ]

    for index, table in enumerate(tables):
        story.append(PageBreak())
        story.append(_pdf_header_block(table["title"], styles["report_title"], styles["meta"], theme))
        story.append(Spacer(1, PDF_LAYOUT["section_gap"]))
        rows = table["rows"]
        visible_rows = rows[:DASHBOARD_PDF_TABLE_ROWS]
        story.append(
            _pdf_table(
                visible_rows,
                table["columns"],
                styles["table_header"],
                styles["table_cell"],
                styles["table_number"],
                styles["table_date"],
                document.width,
                theme,
            )
        )
        if len(rows) > len(visible_rows):
            story.append(Spacer(1, 4 * mm))
            story.append(
                _dashboard_pdf_paragraph(
                    "Visualizacao limitada para apresentacao em PDF. Os dados completos estao disponiveis na exportacao em Excel.",
                    styles["body"],
                )
            )

    document.build(story, onFirstPage=_draw_pdf_page_frame, onLaterPages=_draw_pdf_page_frame)
    return output.getvalue()


def _dashboard_export_response(
    tables: list[dict],
    chart_data: dict,
    format_value: str,
    theme_id: Optional[str] = None,
) -> Response:
    export_format = _normalize_export_format(format_value)
    if export_format not in {"pdf", "excel"}:
        raise HTTPException(status_code=400, detail="A exportacao do dashboard aceita apenas pdf ou excel.")
    media_type, extension = EXPORT_FORMATS[export_format]
    theme = _export_theme(theme_id)
    content = (
        _serialize_dashboard_pdf(tables, chart_data, theme)
        if export_format == "pdf"
        else _serialize_dashboard_excel(tables, chart_data, theme)
    )
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="dashboard.{extension}"'},
    )


def _percent_change(current: float, previous: float) -> Optional[float]:
    if previous == 0:
        return None
    return round(((current - previous) / previous) * 100, 1)


def _trend_direction(value: Optional[float], lower_is_better: bool = False) -> str:
    if value is None or abs(value) < 0.05:
        return "neutral"
    is_positive = value > 0
    if lower_is_better:
        is_positive = not is_positive
    return "up" if is_positive else "down"


def _format_percent(value: Optional[float], fallback: str) -> str:
    if value is None:
        return fallback
    sign = "+" if value > 0 else ""
    return f"{sign}{value:.1f}%"


MONTH_ABBR = (
    "jan",
    "fev",
    "mar",
    "abr",
    "mai",
    "jun",
    "jul",
    "ago",
    "set",
    "out",
    "nov",
    "dez",
)


def _month_year_label(value: date) -> str:
    return f"{MONTH_ABBR[value.month - 1]}/{value.year % 100:02d}"


def _week_year_label(value: date) -> str:
    iso_year, iso_week, _ = value.isocalendar()
    return f"S{iso_week:02d}/{iso_year % 100:02d}"


def _month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def _shift_month_start(value: date, delta_months: int) -> date:
    month_index = value.year * 12 + value.month - 1 + delta_months
    return date(month_index // 12, month_index % 12 + 1, 1)


def _quarter_label(year: int, quarter: int) -> str:
    return f"Q{quarter}/{year % 100:02d}"


def _parse_month_key(value: str) -> Optional[tuple[int, int]]:
    try:
        year, month = value.split("-", 1)
        parsed_year = int(year)
        parsed_month = int(month)
    except ValueError:
        return None
    if parsed_month < 1 or parsed_month > 12:
        return None
    return parsed_year, parsed_month


def _apply_ingredient_category_filters(query, category_ids: Optional[list[str]]):
    if category_ids:
        return query.filter(Ingrediente.category_id.in_(category_ids))
    return query


def _apply_date_filters(
    query,
    column,
    db: Session,
    days: Optional[int] = None,
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = None,
    years: Optional[list[int]] = None,
    months: Optional[list[int]] = None,
    event_dates: Optional[list[date]] = None,
    default_reference_date: Optional[date] = None,
):
    date_column = func.date(column)
    has_explicit_filter = bool(date_from or date_to or month_keys or years or months or event_dates or all_period)

    if not has_explicit_filter and days:
        reference_date = default_reference_date or db.query(func.max(date_column)).scalar()
        if reference_date is not None:
            query = query.filter(date_column >= reference_date - timedelta(days=days - 1))
            if default_reference_date is not None:
                query = query.filter(date_column <= reference_date)

    if date_from:
        query = query.filter(date_column >= date_from)
    if date_to:
        query = query.filter(date_column <= date_to)

    parsed_months = [
        parsed for parsed in (_parse_month_key(value) for value in (month_keys or []))
        if parsed is not None
    ]
    if parsed_months:
        query = query.filter(
            or_(
                *[
                    (func.extract("year", column) == year)
                    & (func.extract("month", column) == month)
                    for year, month in parsed_months
                ]
            )
        )

    if years:
        query = query.filter(func.extract("year", column).in_(years))

    if months:
        query = query.filter(func.extract("month", column).in_(months))

    if event_dates:
        query = query.filter(date_column.in_(event_dates))

    return query


def _dashboard_history_reference_date(db: Session) -> Optional[date]:
    stock_date = db.query(func.max(func.date(Estoque.date_time))).scalar()
    sales_date = db.query(func.max(func.date(Venda.date_time))).scalar()
    available_dates = [value for value in (stock_date, sales_date) if value is not None]
    return min(available_dates) if available_dates else None


def _event_dates_for_types(db: Session, event_types: Optional[list[str]]) -> list[date]:
    if not event_types:
        return []

    predicates = []
    if "holiday" in event_types:
        predicates.append(ResumoDiarioVenda.is_holiday == 1)
    if "rain" in event_types:
        predicates.append(ResumoDiarioVenda.is_rain_event == 1)
    if "promo" in event_types:
        predicates.append(ResumoDiarioVenda.is_promo_day == 1)

    if not predicates:
        return []

    rows = (
        db.query(ResumoDiarioVenda.date)
        .filter(or_(*predicates))
        .order_by(ResumoDiarioVenda.date)
        .all()
    )
    return [row.date for row in rows]


def _resolve_date_window(
    db: Session,
    column,
    days: int,
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = None,
    years: Optional[list[int]] = None,
    months: Optional[list[int]] = None,
    event_dates: Optional[list[date]] = None,
) -> tuple[Optional[date], Optional[date]]:
    date_column = func.date(column)
    reference_date = db.query(func.max(date_column)).scalar()
    if reference_date is None:
        return None, None

    if event_dates:
        return min(event_dates), max(event_dates)

    parsed_months = [
        parsed for parsed in (_parse_month_key(value) for value in (month_keys or []))
        if parsed is not None
    ]
    if parsed_months:
        first_year, first_month = min(parsed_months)
        last_year, last_month = max(parsed_months)
        start = date(first_year, first_month, 1)
        end = _shift_month_start(date(last_year, last_month, 1), 1) - timedelta(days=1)
        return start, end

    if years:
        start_month = min(months or [1])
        end_month = max(months or [12])
        start = date(min(years), start_month, 1)
        end = _shift_month_start(date(max(years), end_month, 1), 1) - timedelta(days=1)
        return start, end

    if months:
        start = date(reference_date.year, min(months), 1)
        end = _shift_month_start(date(reference_date.year, max(months), 1), 1) - timedelta(days=1)
        return start, end

    if date_from or date_to:
        return date_from or reference_date - timedelta(days=days - 1), date_to or reference_date

    if all_period:
        start_date = db.query(func.min(date_column)).scalar()
        return start_date, reference_date

    return reference_date - timedelta(days=days - 1), reference_date


def _stock_total_at_or_before(
    db: Session,
    target_date: date,
    category_ids: Optional[list[str]] = None,
) -> float:
    stock_date = (
        db.query(func.max(func.date(Estoque.date_time)))
        .filter(func.date(Estoque.date_time) <= target_date)
        .scalar()
    )
    if stock_date is None:
        return 0

    query = (
        db.query(func.coalesce(func.sum(Estoque.quantity), 0))
        .join(Ingrediente, Estoque.ingredient_id == Ingrediente.id)
        .filter(func.date(Estoque.date_time) == stock_date)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
    )
    query = _apply_ingredient_category_filters(query, category_ids)
    total = query.scalar()
    return _as_float(total)


def _coverage_rows(
    db: Session,
    start_date: date,
    end_date: date,
    current_stock: bool = True,
    category_ids: Optional[list[str]] = None,
) -> list[dict[str, float | str]]:
    days = max(1, (end_date - start_date).days + 1)
    output_rows = (
        db.query(
            Ingrediente.id,
            Ingrediente.name,
            Ingrediente.unit,
            Categoria.name.label("category"),
            func.sum(Venda.quantity * ReceitaIngrediente.qty).label("output"),
        )
        .join(Categoria, Ingrediente.category_id == Categoria.id)
        .join(ReceitaIngrediente, ReceitaIngrediente.ingredient_id == Ingrediente.id)
        .join(Venda, Venda.recipe_id == ReceitaIngrediente.recipe_id)
        .filter(func.date(Venda.date_time).between(start_date, end_date))
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
    )
    output_query = _apply_ingredient_category_filters(output_rows, category_ids)
    output_rows = (
        output_query
        .group_by(Ingrediente.id, Ingrediente.name, Ingrediente.unit, Categoria.name)
        .all()
    )

    if current_stock:
        stock_query = (
            db.query(EstoqueAtual.ingrediente, EstoqueAtual.qtd)
            .join(Ingrediente, EstoqueAtual.ingrediente == Ingrediente.id)
            .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        )
        stock_query = _apply_ingredient_category_filters(stock_query, category_ids)
        stock_rows = stock_query.all()
        stock_by_id = {row.ingrediente: _as_float(row.qtd) for row in stock_rows}
    else:
        stock_date = (
            db.query(func.max(func.date(Estoque.date_time)))
            .filter(func.date(Estoque.date_time) <= end_date)
            .scalar()
        )
        stock_rows = []
        if stock_date is not None:
            stock_query = (
                db.query(Estoque.ingredient_id, Estoque.quantity)
                .join(Ingrediente, Estoque.ingredient_id == Ingrediente.id)
                .filter(func.date(Estoque.date_time) == stock_date)
                .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
            )
            stock_query = _apply_ingredient_category_filters(stock_query, category_ids)
            stock_rows = stock_query.all()
        stock_by_id = {row.ingredient_id: _as_float(row.quantity) for row in stock_rows}

    coverage = []
    for row in output_rows:
        total_output = _as_float(row.output)
        if total_output <= 0:
            continue
        avg_daily_output = total_output / days
        stock_qty = stock_by_id.get(row.id, 0)
        coverage_days = stock_qty / avg_daily_output if avg_daily_output else 999
        coverage.append(
            {
                "ingredient_id": row.id,
                "name": row.name,
                "category": row.category,
                "unit": row.unit,
                "coverage_days": coverage_days,
                "current_qty": stock_qty,
                "avg_daily_output": avg_daily_output,
            }
        )
    return coverage


def _top_recipe_for_period(
    db: Session,
    start_date: date,
    end_date: date,
    category_ids: Optional[list[str]] = None,
):
    quantity_expr = func.coalesce(func.sum(Venda.quantity), 0).label("quantity")
    query = (
        db.query(Receita.id, Receita.name, quantity_expr)
        .join(Venda, Venda.recipe_id == Receita.id)
        .filter(func.date(Venda.date_time).between(start_date, end_date))
    )
    if category_ids:
        query = query.filter(
            Receita.id.in_(_filtered_recipe_ids_query(db, category_ids=category_ids))
        )
    return (
        query.group_by(Receita.id, Receita.name)
        .order_by(desc(quantity_expr), Receita.name)
        .first()
    )


def _recipe_quantity_for_period(db: Session, recipe_id: str, start_date: date, end_date: date) -> float:
    value = (
        db.query(func.coalesce(func.sum(Venda.quantity), 0))
        .filter(Venda.recipe_id == recipe_id)
        .filter(func.date(Venda.date_time).between(start_date, end_date))
        .scalar()
    )
    return _as_float(value)


def _dashboard_kpis(
    db: Session,
    category_ids: Optional[list[str]] = None,
    days: int = 28,
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = None,
    years: Optional[list[int]] = None,
    months: Optional[list[int]] = None,
    event_dates: Optional[list[date]] = None,
) -> list[DashboardKpi]:
    sales_reference = db.query(func.max(func.date(Venda.date_time))).scalar()
    stock_reference = db.query(func.max(func.date(Estoque.date_time))).scalar()
    current_items_query = (
        db.query(func.count(Ingrediente.id))
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
    )
    current_items = _apply_ingredient_category_filters(current_items_query, category_ids).scalar() or 0
    current_stock_query = (
        db.query(func.coalesce(func.sum(EstoqueAtual.qtd), 0))
        .join(Ingrediente, EstoqueAtual.ingrediente == Ingrediente.id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
    )
    current_stock_qty = _apply_ingredient_category_filters(current_stock_query, category_ids).scalar() or 0
    previous_stock_reference = (stock_reference or date.today()) - timedelta(days=7)
    previous_stock_qty = _stock_total_at_or_before(db, previous_stock_reference, category_ids)
    stock_delta = _percent_change(_as_float(current_stock_qty), previous_stock_qty)

    empty_kpis = [
        DashboardKpi(
            id="ingredients",
            label="Estoque atual",
            value=f"{_as_float(current_stock_qty):,.1f}".replace(",", "X").replace(".", ",").replace("X", "."),
            detail=f"{current_items:,}".replace(",", ".") + " ingredientes cadastrados",
            trend_value=stock_delta,
            trend_label=f"{_format_percent(stock_delta, 'sem histórico')} vs {_week_year_label(previous_stock_reference)}",
            trend_direction=_trend_direction(stock_delta),
        )
    ]
    if sales_reference is None:
        return empty_kpis

    current_start, current_end = _resolve_date_window(
        db,
        Venda.date_time,
        days=max(1, days),
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        months=months,
        event_dates=event_dates,
    )
    if current_start is None or current_end is None:
        return empty_kpis
    window_days = max(1, (current_end - current_start).days + 1)
    previous_end = current_start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=window_days - 1)

    current_coverage = _coverage_rows(db, current_start, current_end, current_stock=True, category_ids=category_ids)
    previous_coverage = _coverage_rows(db, previous_start, previous_end, current_stock=False, category_ids=category_ids)
    current_avg_coverage = (
        sum(float(item["coverage_days"]) for item in current_coverage) / len(current_coverage)
        if current_coverage
        else 0
    )
    previous_avg_coverage = (
        sum(float(item["coverage_days"]) for item in previous_coverage) / len(previous_coverage)
        if previous_coverage
        else 0
    )
    coverage_delta = _percent_change(current_avg_coverage, previous_avg_coverage)

    top_recipe = _top_recipe_for_period(db, current_start, current_end, category_ids=category_ids)
    top_recipe_delta = None
    top_recipe_detail = "Sem vendas no período"
    top_recipe_value = "-"
    if top_recipe is not None:
        current_quantity = _as_float(top_recipe.quantity)
        previous_quantity = _recipe_quantity_for_period(
            db,
            top_recipe.id,
            previous_start,
            previous_end,
        )
        top_recipe_delta = _percent_change(current_quantity, previous_quantity)
        top_recipe_value = top_recipe.name
        top_recipe_detail = f"{current_quantity:,.0f}".replace(",", ".") + f" unidades no período"

    previous_by_id = {item["ingredient_id"]: item for item in previous_coverage}
    critical = min(current_coverage, key=lambda item: float(item["coverage_days"]), default=None)
    critical_delta = None
    critical_value = "-"
    critical_detail = "Sem consumo recente"
    if critical is not None:
        previous = previous_by_id.get(str(critical["ingredient_id"]))
        previous_days = float(previous["coverage_days"]) if previous else 0
        critical_delta = _percent_change(float(critical["coverage_days"]), previous_days)
        critical_value = str(critical["name"])
        critical_detail = f"{float(critical['coverage_days']):.1f} dias de cobertura"

    return [
        *empty_kpis,
        DashboardKpi(
            id="coverage",
            label="Cobertura média",
            value=f"{current_avg_coverage:.1f} dias",
            detail="Ingredientes com consumo recente",
            trend_value=coverage_delta,
            trend_label=f"{_format_percent(coverage_delta, 'sem histórico')} vs {_month_year_label(previous_end)}",
            trend_direction=_trend_direction(coverage_delta),
        ),
        DashboardKpi(
            id="top_recipe",
            label="Receita que mais sai",
            value=top_recipe_value,
            detail=top_recipe_detail,
            trend_value=top_recipe_delta,
            trend_label=f"{_format_percent(top_recipe_delta, 'sem histórico')} vs {_month_year_label(previous_end)}",
            trend_direction=_trend_direction(top_recipe_delta),
        ),
        DashboardKpi(
            id="critical_ingredient",
            label="Ingrediente crítico",
            value=critical_value,
            detail=critical_detail,
            trend_value=critical_delta,
            trend_label=f"{_format_percent(critical_delta, 'sem histórico')} vs {_month_year_label(previous_end)}",
            trend_direction=_trend_direction(critical_delta),
        ),
    ]


def _stock_product_rows(
    db: Session,
    order,
    limit: int = 8,
    unit: Optional[str] = None,
    category_ids: Optional[list[str]] = None,
) -> list[DashboardRankItem]:
    value_expr = func.coalesce(EstoqueAtual.qtd, 0).label("value")
    query = (
        db.query(
            Ingrediente.id,
            Ingrediente.name,
            Ingrediente.unit,
            Categoria.name.label("category"),
            value_expr,
        )
        .outerjoin(Categoria, Ingrediente.category_id == Categoria.id)
        .outerjoin(EstoqueAtual, EstoqueAtual.ingrediente == Ingrediente.id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
    )
    query = _apply_ingredient_category_filters(query, category_ids)
    if unit:
        query = query.filter(Ingrediente.unit == unit)

    rows = (
        query
        .order_by(order(value_expr), Ingrediente.name)
        .limit(limit)
        .all()
    )
    return [
        DashboardRankItem(
            id=row.id,
            name=row.name,
            unit=row.unit,
            category=row.category,
            value=_as_float(row.value),
        )
        for row in rows
    ]


def _stock_category_rows(
    db: Session,
    order,
    limit: int = 8,
    unit: Optional[str] = None,
    category_ids: Optional[list[str]] = None,
) -> list[DashboardCategoryItem]:
    value_expr = func.coalesce(func.sum(EstoqueAtual.qtd), 0).label("value")
    query = (
        db.query(Categoria.id, Categoria.name, value_expr)
        .join(Ingrediente, Ingrediente.category_id == Categoria.id)
        .outerjoin(EstoqueAtual, EstoqueAtual.ingrediente == Ingrediente.id)
        .filter(Categoria.id != PRODUCTION_CATEGORY_ID)
    )
    if category_ids:
        query = query.filter(Categoria.id.in_(category_ids))
    if unit:
        query = query.filter(Ingrediente.unit == unit)

    rows = (
        query
        .group_by(Categoria.id, Categoria.name)
        .order_by(order(value_expr), Categoria.name)
        .limit(limit)
        .all()
    )
    return [
        DashboardCategoryItem(
            id=row.id,
            name=row.name,
            value=_as_float(row.value),
            unit=unit,
        )
        for row in rows
    ]


def _stock_product_groups(
    db: Session,
    order,
    limit: int = 8,
    category_ids: Optional[list[str]] = None,
) -> list[DashboardUnitRankGroup]:
    return [
        DashboardUnitRankGroup(
            unit=unit,
            items=_stock_product_rows(db, order, limit=limit, unit=unit, category_ids=category_ids),
        )
        for unit in DASHBOARD_STOCK_UNITS
    ]


def _stock_category_groups(
    db: Session,
    order,
    limit: int = 8,
    category_ids: Optional[list[str]] = None,
) -> list[DashboardUnitCategoryGroup]:
    return [
        DashboardUnitCategoryGroup(
            unit=unit,
            items=_stock_category_rows(db, order, limit=limit, unit=unit, category_ids=category_ids),
        )
        for unit in DASHBOARD_STOCK_UNITS
    ]


def _output_product_rows(
    db: Session,
    order,
    limit: int = 8,
    unit: Optional[str] = None,
    category_ids: Optional[list[str]] = None,
    days: Optional[int] = None,
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = None,
    years: Optional[list[int]] = None,
    months: Optional[list[int]] = None,
    event_dates: Optional[list[date]] = None,
) -> list[DashboardRankItem]:
    value_expr = func.coalesce(
        func.sum(Venda.quantity * ReceitaIngrediente.qty), 0
    ).label("value")
    query = (
        db.query(
            Ingrediente.id,
            Ingrediente.name,
            Ingrediente.unit,
            Categoria.name.label("category"),
            value_expr,
        )
        .outerjoin(Categoria, Ingrediente.category_id == Categoria.id)
        .outerjoin(
            ReceitaIngrediente, ReceitaIngrediente.ingredient_id == Ingrediente.id
        )
        .outerjoin(Venda, Venda.recipe_id == ReceitaIngrediente.recipe_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
    )
    query = _apply_ingredient_category_filters(query, category_ids)
    query = _apply_date_filters(
        query,
        Venda.date_time,
        db,
        days=days,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        months=months,
        event_dates=event_dates,
    )
    if unit:
        query = query.filter(Ingrediente.unit == unit)

    rows = (
        query
        .group_by(Ingrediente.id, Ingrediente.name, Ingrediente.unit, Categoria.name)
        .order_by(order(value_expr), Ingrediente.name)
        .limit(limit)
        .all()
    )
    return [
        DashboardRankItem(
            id=row.id,
            name=row.name,
            unit=row.unit,
            category=row.category,
            value=_as_float(row.value),
        )
        for row in rows
    ]


def _output_category_rows(
    db: Session,
    order,
    limit: int = 8,
    unit: Optional[str] = None,
    category_ids: Optional[list[str]] = None,
    days: Optional[int] = None,
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = None,
    years: Optional[list[int]] = None,
    months: Optional[list[int]] = None,
    event_dates: Optional[list[date]] = None,
) -> list[DashboardCategoryItem]:
    value_expr = func.coalesce(
        func.sum(Venda.quantity * ReceitaIngrediente.qty), 0
    ).label("value")
    query = (
        db.query(Categoria.id, Categoria.name, value_expr)
        .join(Ingrediente, Ingrediente.category_id == Categoria.id)
        .outerjoin(
            ReceitaIngrediente, ReceitaIngrediente.ingredient_id == Ingrediente.id
        )
        .outerjoin(Venda, Venda.recipe_id == ReceitaIngrediente.recipe_id)
        .filter(Categoria.id != PRODUCTION_CATEGORY_ID)
    )
    if category_ids:
        query = query.filter(Categoria.id.in_(category_ids))
    query = _apply_date_filters(
        query,
        Venda.date_time,
        db,
        days=days,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        months=months,
        event_dates=event_dates,
    )
    if unit:
        query = query.filter(Ingrediente.unit == unit)

    rows = (
        query
        .group_by(Categoria.id, Categoria.name)
        .order_by(order(value_expr), Categoria.name)
        .limit(limit)
        .all()
    )
    return [
        DashboardCategoryItem(
            id=row.id,
            name=row.name,
            value=_as_float(row.value),
            unit=unit,
        )
        for row in rows
    ]


def _output_product_groups(
    db: Session,
    order,
    limit: int = 8,
    category_ids: Optional[list[str]] = None,
    days: Optional[int] = None,
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = None,
    years: Optional[list[int]] = None,
    months: Optional[list[int]] = None,
    event_dates: Optional[list[date]] = None,
) -> list[DashboardUnitRankGroup]:
    return [
        DashboardUnitRankGroup(
            unit=unit,
            items=_output_product_rows(
                db,
                order,
                limit=limit,
                unit=unit,
                category_ids=category_ids,
                days=days,
                all_period=all_period,
                date_from=date_from,
                date_to=date_to,
                month_keys=month_keys,
                years=years,
                months=months,
                event_dates=event_dates,
            ),
        )
        for unit in DASHBOARD_STOCK_UNITS
    ]


def _output_category_groups(
    db: Session,
    order,
    limit: int = 8,
    category_ids: Optional[list[str]] = None,
    days: Optional[int] = None,
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = None,
    years: Optional[list[int]] = None,
    months: Optional[list[int]] = None,
    event_dates: Optional[list[date]] = None,
) -> list[DashboardUnitCategoryGroup]:
    return [
        DashboardUnitCategoryGroup(
            unit=unit,
            items=_output_category_rows(
                db,
                order,
                limit=limit,
                unit=unit,
                category_ids=category_ids,
                days=days,
                all_period=all_period,
                date_from=date_from,
                date_to=date_to,
                month_keys=month_keys,
                years=years,
                months=months,
                event_dates=event_dates,
            ),
        )
        for unit in DASHBOARD_STOCK_UNITS
    ]


def _dashboard_filters(db: Session) -> DashboardFilters:
    categories = (
        db.query(Categoria.id, Categoria.name)
        .filter(Categoria.id != PRODUCTION_CATEGORY_ID)
        .order_by(Categoria.name)
        .all()
    )
    ingredients = (
        db.query(
            Ingrediente.id,
            Ingrediente.name,
            Ingrediente.category_id,
            Categoria.name.label("category"),
        )
        .join(Categoria, Ingrediente.category_id == Categoria.id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .order_by(Categoria.name, Ingrediente.name)
        .all()
    )
    holidays = (
        db.query(FeriadoRecife.data, FeriadoRecife.nome, FeriadoRecife.tipo)
        .order_by(FeriadoRecife.data)
        .all()
    )
    year_expr = func.extract("year", Venda.date_time).label("year")
    month_expr = func.extract("month", Venda.date_time).label("month")
    month_rows = (
        db.query(
            year_expr,
            month_expr,
        )
        .group_by(year_expr, month_expr)
        .order_by(year_expr, month_expr)
        .all()
    )
    return DashboardFilters(
        categories=[
            DashboardCategoryItem(id=row.id, name=row.name, value=0)
            for row in categories
        ],
        ingredients=[
            DashboardIngredientFilter(
                id=row.id,
                name=row.name,
                category_id=row.category_id,
                category=row.category,
            )
            for row in ingredients
        ],
        holidays=[
            DashboardHolidayFilter(date=row.data, name=row.nome, type=row.tipo)
            for row in holidays
        ],
        months=[
            DashboardMonthFilter(
                key=f"{int(row.year)}-{int(row.month):02d}",
                label=_month_year_label(date(int(row.year), int(row.month), 1)),
            )
            for row in month_rows
        ],
    )


def _dashboard_alerts(
    db: Session,
    limit: int = 10,
    category_ids: Optional[list[str]] = None,
    days: int = 28,
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = None,
    years: Optional[list[int]] = None,
    months: Optional[list[int]] = None,
    event_dates: Optional[list[date]] = None,
) -> list[DashboardAlert]:
    reference_date = db.query(func.max(func.date(Venda.date_time))).scalar()
    if reference_date is None:
        return []

    consumo_expr = func.sum(Venda.quantity * ReceitaIngrediente.qty).label("output")
    query = (
        db.query(
            Ingrediente.id,
            Ingrediente.name,
            Ingrediente.unit,
            Categoria.name.label("category"),
            func.coalesce(EstoqueAtual.qtd, 0).label("current_qty"),
            consumo_expr,
        )
        .join(Categoria, Ingrediente.category_id == Categoria.id)
        .outerjoin(EstoqueAtual, EstoqueAtual.ingrediente == Ingrediente.id)
        .join(ReceitaIngrediente, ReceitaIngrediente.ingredient_id == Ingrediente.id)
        .join(Venda, Venda.recipe_id == ReceitaIngrediente.recipe_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
    )
    query = _apply_ingredient_category_filters(query, category_ids)
    query = _apply_date_filters(
        query,
        Venda.date_time,
        db,
        days=days,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        months=months,
        event_dates=event_dates,
    )
    rows = (
        query.group_by(
            Ingrediente.id,
            Ingrediente.name,
            Ingrediente.unit,
            Categoria.name,
            EstoqueAtual.qtd,
        )
        .all()
    )

    alerts: list[DashboardAlert] = []
    for row in rows:
        total_output = _as_float(row.output)
        if total_output <= 0:
            continue

        avg_daily_output = total_output / max(1, days)
        current_qty = _as_float(row.current_qty)
        coverage_days = current_qty / avg_daily_output if avg_daily_output else 999
        if coverage_days <= 3:
            severity = "Crítico"
        elif coverage_days <= 7:
            severity = "Atenção"
        else:
            severity = "Monitorar"

        alerts.append(
            DashboardAlert(
                ingredient_id=row.id,
                name=row.name,
                category=row.category,
                unit=row.unit,
                current_qty=round(current_qty, 3),
                avg_daily_output=round(avg_daily_output, 3),
                coverage_days=round(coverage_days, 1),
                suggested_qty=round(max(0, avg_daily_output * 7 - current_qty), 3),
                severity=severity,
            )
        )

    return sorted(alerts, key=lambda item: (item.coverage_days, -item.avg_daily_output))[
        :limit
    ]


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------


@app.get("/api/dashboard", response_model=DashboardResponse)
def get_dashboard(
    category_ids: Optional[list[str]] = Query(default=None),
    days: int = Query(default=90, ge=1, le=730),
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = Query(default=None),
    years: Optional[list[int]] = Query(default=None),
    month_numbers: Optional[list[int]] = Query(default=None),
    event_types: Optional[list[str]] = Query(default=None),
    db: Session = Depends(get_db),
):
    event_dates = _event_dates_for_types(db, event_types)
    top_stock_products_by_unit = _stock_product_groups(db, desc, category_ids=category_ids)
    bottom_stock_products_by_unit = _stock_product_groups(db, asc, category_ids=category_ids)
    top_stock_categories_by_unit = _stock_category_groups(db, desc, category_ids=category_ids)
    bottom_stock_categories_by_unit = _stock_category_groups(db, asc, category_ids=category_ids)
    output_kwargs = {
        "category_ids": category_ids,
        "days": days,
        "all_period": all_period,
        "date_from": date_from,
        "date_to": date_to,
        "month_keys": month_keys,
        "years": years,
        "months": month_numbers,
        "event_dates": event_dates,
    }
    top_output_products_by_unit = _output_product_groups(db, desc, **output_kwargs)
    bottom_output_products_by_unit = _output_product_groups(db, asc, **output_kwargs)
    top_output_categories_by_unit = _output_category_groups(db, desc, **output_kwargs)
    bottom_output_categories_by_unit = _output_category_groups(db, asc, **output_kwargs)

    return DashboardResponse(
        cards=DashboardCards(
            items=_dashboard_kpis(
                db,
                category_ids=category_ids,
                days=days,
                all_period=all_period,
                date_from=date_from,
                date_to=date_to,
                month_keys=month_keys,
                years=years,
                months=month_numbers,
                event_dates=event_dates,
            )
        ),
        top_stock_products_by_unit=top_stock_products_by_unit,
        bottom_stock_products_by_unit=bottom_stock_products_by_unit,
        top_stock_categories_by_unit=top_stock_categories_by_unit,
        bottom_stock_categories_by_unit=bottom_stock_categories_by_unit,
        top_output_products_by_unit=top_output_products_by_unit,
        bottom_output_products_by_unit=bottom_output_products_by_unit,
        top_output_categories_by_unit=top_output_categories_by_unit,
        bottom_output_categories_by_unit=bottom_output_categories_by_unit,
        alerts=_dashboard_alerts(
            db,
            category_ids=category_ids,
            days=days,
            all_period=all_period,
            date_from=date_from,
            date_to=date_to,
            month_keys=month_keys,
            years=years,
            months=month_numbers,
            event_dates=event_dates,
        ),
        filters=_dashboard_filters(db),
    )


@app.get("/api/dashboard/estoque-historico", response_model=list[DashboardHistoryPoint])
def get_dashboard_estoque_historico(
    ingredient_id: Optional[str] = None,
    category_id: Optional[str] = None,
    category_ids: Optional[list[str]] = Query(default=None),
    days: int = Query(default=90, ge=7, le=730),
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = Query(default=None),
    years: Optional[list[int]] = Query(default=None),
    month_numbers: Optional[list[int]] = Query(default=None),
    event_types: Optional[list[str]] = Query(default=None),
    db: Session = Depends(get_db),
):
    event_dates = _event_dates_for_types(db, event_types)
    history_reference_date = _dashboard_history_reference_date(db)
    date_expr = func.date(Estoque.date_time).label("date")
    value_expr = func.coalesce(func.sum(Estoque.quantity), 0).label("value")
    query = (
        db.query(date_expr, value_expr)
        .join(Ingrediente, Estoque.ingredient_id == Ingrediente.id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
    )
    query = _apply_date_filters(
        query,
        Estoque.date_time,
        db,
        days=days,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        months=month_numbers,
        event_dates=event_dates,
        default_reference_date=history_reference_date,
    )

    if ingredient_id:
        query = query.filter(Ingrediente.id == ingredient_id)
    elif category_id:
        query = query.filter(Ingrediente.category_id == category_id)
    else:
        query = _apply_ingredient_category_filters(query, category_ids)

    rows = query.group_by(date_expr).order_by(date_expr).all()
    return [
        DashboardHistoryPoint(date=row.date, value=_as_float(row.value))
        for row in rows
    ]


def _filtered_recipe_ids_query(
    db: Session,
    ingredient_id: Optional[str] = None,
    category_id: Optional[str] = None,
    category_ids: Optional[list[str]] = None,
):
    query = db.query(ReceitaIngrediente.recipe_id).join(
        Ingrediente, ReceitaIngrediente.ingredient_id == Ingrediente.id
    )
    if ingredient_id:
        query = query.filter(Ingrediente.id == ingredient_id)
    elif category_ids:
        query = query.filter(Ingrediente.category_id.in_(category_ids))
    elif category_id:
        query = query.filter(Ingrediente.category_id == category_id)
    else:
        query = query.filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
    return query.distinct()


@app.get("/api/dashboard/vendas-historico", response_model=list[DashboardHistoryPoint])
def get_dashboard_vendas_historico(
    ingredient_id: Optional[str] = None,
    category_id: Optional[str] = None,
    category_ids: Optional[list[str]] = Query(default=None),
    days: int = Query(default=90, ge=7, le=730),
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = Query(default=None),
    years: Optional[list[int]] = Query(default=None),
    month_numbers: Optional[list[int]] = Query(default=None),
    event_types: Optional[list[str]] = Query(default=None),
    db: Session = Depends(get_db),
):
    event_dates = _event_dates_for_types(db, event_types)
    history_reference_date = _dashboard_history_reference_date(db)
    date_expr = func.date(Venda.date_time).label("date")
    value_expr = func.coalesce(func.sum(Venda.quantity), 0).label("value")
    query = db.query(date_expr, value_expr)
    query = _apply_date_filters(
        query,
        Venda.date_time,
        db,
        days=days,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        months=month_numbers,
        event_dates=event_dates,
        default_reference_date=history_reference_date,
    )

    if ingredient_id or category_id or category_ids:
        query = query.filter(
            Venda.recipe_id.in_(
                _filtered_recipe_ids_query(
                    db,
                    ingredient_id=ingredient_id,
                    category_id=category_id,
                    category_ids=category_ids,
                )
            )
        )

    rows = query.group_by(date_expr).order_by(date_expr).all()
    return [
        DashboardHistoryPoint(date=row.date, value=_as_float(row.value))
        for row in rows
    ]


@app.get("/api/dashboard/faturamento-resumo", response_model=DashboardRevenueSummary)
def get_dashboard_faturamento_resumo(
    months: int = Query(default=12, ge=3, le=36),
    category_ids: Optional[list[str]] = Query(default=None),
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = Query(default=None),
    years: Optional[list[int]] = Query(default=None),
    month_numbers: Optional[list[int]] = Query(default=None),
    event_types: Optional[list[str]] = Query(default=None),
    db: Session = Depends(get_db),
):
    event_dates = _event_dates_for_types(db, event_types)
    reference_date = db.query(func.max(func.date(Venda.date_time))).scalar()
    if reference_date is None:
        return DashboardRevenueSummary(monthly=[], quarterly=[])

    start_date = (
        db.query(func.min(func.date(Venda.date_time))).scalar()
        if all_period
        else _shift_month_start(_month_start(reference_date), -(months - 1))
    )
    revenue_expr = func.coalesce(func.sum(Venda.quantity * Venda.unit_price), 0).label("value")

    year_expr = func.extract("year", Venda.date_time).label("year")
    month_expr = func.extract("month", Venda.date_time).label("month")
    month_rows = (
        db.query(year_expr, month_expr, revenue_expr)
        .filter(func.date(Venda.date_time) >= start_date)
    )
    month_query = _apply_date_filters(
        month_rows,
        Venda.date_time,
        db,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        months=month_numbers,
        event_dates=event_dates,
    )
    if category_ids:
        month_query = month_query.filter(
            Venda.recipe_id.in_(_filtered_recipe_ids_query(db, category_ids=category_ids))
        )
    month_rows = month_query.group_by(year_expr, month_expr).order_by(year_expr, month_expr).all()
    monthly = [
        DashboardNamedMetric(
            key=f"{int(row.year)}-{int(row.month):02d}",
            label=_month_year_label(date(int(row.year), int(row.month), 1)),
            value=_as_float(row.value),
        )
        for row in month_rows
    ]

    quarter_expr = func.extract("quarter", Venda.date_time).label("quarter")
    quarter_rows = (
        db.query(year_expr, quarter_expr, revenue_expr)
        .filter(func.date(Venda.date_time) >= start_date)
    )
    quarter_query = _apply_date_filters(
        quarter_rows,
        Venda.date_time,
        db,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        months=month_numbers,
        event_dates=event_dates,
    )
    if category_ids:
        quarter_query = quarter_query.filter(
            Venda.recipe_id.in_(_filtered_recipe_ids_query(db, category_ids=category_ids))
        )
    quarter_rows = quarter_query.group_by(year_expr, quarter_expr).order_by(year_expr, quarter_expr).all()
    quarterly = [
        DashboardNamedMetric(
            key=f"{int(row.year)}-Q{int(row.quarter)}",
            label=_quarter_label(int(row.year), int(row.quarter)),
            value=_as_float(row.value),
        )
        for row in quarter_rows
    ]

    return DashboardRevenueSummary(monthly=monthly, quarterly=quarterly)


@app.get("/api/dashboard/pedidos-semana", response_model=list[DashboardNamedMetric])
def get_dashboard_pedidos_semana(
    days: int = Query(default=90, ge=7, le=730),
    category_ids: Optional[list[str]] = Query(default=None),
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = Query(default=None),
    years: Optional[list[int]] = Query(default=None),
    month_numbers: Optional[list[int]] = Query(default=None),
    event_types: Optional[list[str]] = Query(default=None),
    db: Session = Depends(get_db),
):
    event_dates = _event_dates_for_types(db, event_types)
    dow_expr = func.extract("dow", Venda.date_time).label("dow")
    value_expr = func.coalesce(func.sum(Venda.quantity), 0).label("value")
    query = db.query(dow_expr, value_expr)
    query = _apply_date_filters(
        query,
        Venda.date_time,
        db,
        days=days,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        months=month_numbers,
        event_dates=event_dates,
    )
    if category_ids:
        query = query.filter(
            Venda.recipe_id.in_(_filtered_recipe_ids_query(db, category_ids=category_ids))
        )
    rows = query.group_by(dow_expr).all()
    by_dow = {int(row.dow): _as_float(row.value) for row in rows}
    labels = {
        0: "Dom",
        1: "Seg",
        2: "Ter",
        3: "Qua",
        4: "Qui",
        5: "Sex",
        6: "Sáb",
    }

    return [
        DashboardNamedMetric(key=str(dow), label=labels[dow], value=by_dow.get(dow, 0))
        for dow in [1, 2, 3, 4, 5, 6, 0]
    ]


@app.get("/api/dashboard/receitas-ranking", response_model=list[DashboardRecipeItem])
def get_dashboard_receitas_ranking(
    ingredient_id: Optional[str] = None,
    category_ids: Optional[list[str]] = Query(default=None),
    days: int = Query(default=90, ge=7, le=730),
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = Query(default=None),
    years: Optional[list[int]] = Query(default=None),
    month_numbers: Optional[list[int]] = Query(default=None),
    event_types: Optional[list[str]] = Query(default=None),
    limit: int = Query(default=10, ge=1, le=50),
    db: Session = Depends(get_db),
):
    event_dates = _event_dates_for_types(db, event_types)
    quantity_expr = func.coalesce(func.sum(Venda.quantity), 0).label("quantity")
    revenue_expr = func.coalesce(func.sum(Venda.quantity * Venda.unit_price), 0).label("revenue")
    query = (
        db.query(Receita.id, Receita.name, quantity_expr, revenue_expr)
        .join(Venda, Venda.recipe_id == Receita.id)
    )
    query = _apply_date_filters(
        query,
        Venda.date_time,
        db,
        days=days,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        months=month_numbers,
        event_dates=event_dates,
    )

    if ingredient_id or category_ids:
        query = query.filter(
            Receita.id.in_(
                _filtered_recipe_ids_query(
                    db,
                    ingredient_id=ingredient_id,
                    category_ids=category_ids,
                )
            )
        )

    rows = (
        query
        .group_by(Receita.id, Receita.name)
        .order_by(desc(quantity_expr), Receita.name)
        .limit(limit)
        .all()
    )
    return [
        DashboardRecipeItem(
            id=row.id,
            name=row.name,
            quantity=_as_float(row.quantity),
            revenue=_as_float(row.revenue),
        )
        for row in rows
    ]


@app.get("/api/export/dashboard")
def export_dashboard(
    format: str = Query(default="pdf"),
    theme: Optional[str] = Query(default=None),
    category_ids: Optional[list[str]] = Query(default=None),
    days: int = Query(default=90, ge=7, le=730),
    all_period: bool = False,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month_keys: Optional[list[str]] = Query(default=None),
    years: Optional[list[int]] = Query(default=None),
    month_numbers: Optional[list[int]] = Query(default=None),
    event_types: Optional[list[str]] = Query(default=None),
    db: Session = Depends(get_db),
):
    if date_from and date_to and date_from > date_to:
        raise HTTPException(status_code=400, detail="Data inicial maior que data final")
    tables, chart_data = _dashboard_export_tables(
        db,
        category_ids=category_ids,
        days=days,
        all_period=all_period,
        date_from=date_from,
        date_to=date_to,
        month_keys=month_keys,
        years=years,
        month_numbers=month_numbers,
        event_types=event_types,
    )
    return _dashboard_export_response(tables, chart_data, format, theme)


@app.get("/api/estoque", response_model=list[IngredienteOut])
def get_estoque(
    category: Optional[str] = None,
    status: Optional[str] = None,
    q: Optional[str] = None,
    db: Session = Depends(get_db),
):
    query = (
        db.query(Ingrediente)
        .outerjoin(Categoria)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .order_by(Categoria.name, Ingrediente.name)
    )
    if category:
        query = query.filter(or_(Ingrediente.category_id == category, Categoria.name == category))
    if q:
        query = query.filter(Ingrediente.name.ilike(f"%{q}%"))
    criticidade_run, criticidade_by_ingredient = _latest_criticidade_by_ingredient(db)
    items = query.all()
    if status:
        items = [
            item
            for item in items
            if _model_stock_status(item, criticidade_by_ingredient.get(item.id)) == status
        ]
    return [_ingrediente_out(item, criticidade_run, criticidade_by_ingredient) for item in items]


@app.get("/api/estoque/paginado", response_model=EstoquePaginado)
def get_estoque_paginado(
    category: Optional[str] = None,
    status: Optional[str] = None,
    q: Optional[str] = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    query = (
        db.query(Ingrediente)
        .outerjoin(Categoria)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .order_by(Categoria.name, Ingrediente.name)
    )
    if category:
        query = query.filter(or_(Ingrediente.category_id == category, Categoria.name == category))
    if q:
        query = query.filter(Ingrediente.name.ilike(f"%{q}%"))
    criticidade_run, criticidade_by_ingredient = _latest_criticidade_by_ingredient(db)
    items = query.all()
    if status:
        items = [
            item
            for item in items
            if _model_stock_status(item, criticidade_by_ingredient.get(item.id)) == status
        ]

    total = len(items)
    total_pages = max(1, math.ceil(total / page_size))
    offset = (page - 1) * page_size
    page_items = items[offset: offset + page_size]

    return EstoquePaginado(
        items=[_ingrediente_out(item, criticidade_run, criticidade_by_ingredient) for item in page_items],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
    )


@app.get("/api/export/estoque")
def export_estoque(
    format: str = Query(default="csv"),
    theme: Optional[str] = Query(default=None),
    date_from: date = Query(...),
    date_to: date = Query(...),
    db: Session = Depends(get_db),
):
    if date_from > date_to:
        raise HTTPException(status_code=400, detail="Data inicial maior que data final")

    rows = (
        db.query(
            Estoque.id,
            Estoque.date_time,
            Estoque.ingredient_id,
            Ingrediente.name.label("ingredient_name"),
            Categoria.name.label("category"),
            Ingrediente.unit,
            Estoque.quantity,
        )
        .join(Ingrediente, Ingrediente.id == Estoque.ingredient_id)
        .join(Categoria, Categoria.id == Ingrediente.category_id)
        .filter(func.date(Estoque.date_time) >= date_from)
        .filter(func.date(Estoque.date_time) <= date_to)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .order_by(Estoque.date_time.asc(), Ingrediente.name.asc())
        .all()
    )
    data = [
        {
            "id": row.id,
            "data_hora": row.date_time,
            "ingrediente_id": row.ingredient_id,
            "ingrediente": row.ingredient_name,
            "categoria": row.category,
            "unidade": row.unit,
            "quantidade": _as_float(row.quantity),
        }
        for row in rows
    ]
    return _export_response(
        data,
        f"estoque_{date_from.isoformat()}_{date_to.isoformat()}",
        format,
        "Historico de estoque",
        [
            "id",
            "data_hora",
            "ingrediente_id",
            "ingrediente",
            "categoria",
            "unidade",
            "quantidade",
        ],
        theme_id=theme,
    )


@app.patch("/api/estoque", response_model=ResultadoLote)
def update_estoque(lote: AtualizacaoLote, db: Session = Depends(get_db)):
    contagem = None
    if lote.contagem_id is not None:
        contagem = db.query(Contagem).filter(Contagem.id == lote.contagem_id).first()
        if contagem is None:
            raise HTTPException(status_code=404, detail="Contagem não encontrada")
        if contagem.status == "finalizada":
            contagem.status = "em_andamento"
            contagem.finalizada_em = None

    ids = [u.id for u in lote.updates]
    por_id = {
        i.id: i
        for i in db.query(Ingrediente)
        .filter(Ingrediente.id.in_(ids))
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .all()
    }

    count = 0
    for atualizacao in lote.updates:
        ingrediente = por_id.get(atualizacao.id)
        if ingrediente is None:
            continue

        anterior = float(ingrediente.current_qty)
        if contagem is not None:
            delta = round(atualizacao.new_qty - anterior, 3)
            snapshot = _estoque_snapshot_for_ingredient(
                db,
                ingrediente.id,
                contagem.estoque_snapshot_data,
            )
            existing_log = (
                db.query(ContagemLog)
                .filter(ContagemLog.contagem_id == contagem.id)
                .filter(ContagemLog.ingrediente_id == ingrediente.id)
                .order_by(ContagemLog.criado_em.desc(), ContagemLog.id.desc())
                .first()
            )
            if existing_log is None:
                db.add(
                    ContagemLog(
                        contagem_id=contagem.id,
                        ingrediente_id=ingrediente.id,
                        estoque_id=snapshot.id if snapshot else None,
                        estoque_data=(
                            snapshot.date_time.date()
                            if snapshot and snapshot.date_time
                            else contagem.estoque_snapshot_data
                        ),
                        estoque_quantidade=snapshot.quantity if snapshot else None,
                        category_id=ingrediente.category_id,
                        categoria=ingrediente.category,
                        quantidade_anterior=anterior,
                        quantidade_nova=atualizacao.new_qty,
                        delta=delta,
                    )
                )
            else:
                existing_log.estoque_id = snapshot.id if snapshot else None
                existing_log.estoque_data = (
                    snapshot.date_time.date()
                    if snapshot and snapshot.date_time
                    else contagem.estoque_snapshot_data
                )
                existing_log.estoque_quantidade = snapshot.quantity if snapshot else None
                existing_log.category_id = ingrediente.category_id
                existing_log.categoria = ingrediente.category
                existing_log.quantidade_anterior = anterior
                existing_log.quantidade_nova = atualizacao.new_qty
                existing_log.delta = delta
                existing_log.criado_em = _now_recife()
            if ingrediente.estoque_atual is None:
                ingrediente.estoque_atual = EstoqueAtual(
                    id=f"CUR-{ingrediente.id}",
                    qtd=atualizacao.new_qty,
                    data=contagem.data_contagem,
                )
            else:
                ingrediente.estoque_atual.qtd = atualizacao.new_qty
                ingrediente.estoque_atual.data = contagem.data_contagem
        if round(atualizacao.new_qty, 3) == round(anterior, 3):
            continue
        db.add(
            LogContagem(
                ingrediente_id=ingrediente.id,
                quantidade_anterior=anterior,
                quantidade_nova=atualizacao.new_qty,
                delta=round(atualizacao.new_qty - anterior, 3),
                sessao=lote.session_label,
            )
        )
        if contagem is None and ingrediente.estoque_atual is None:
                ingrediente.estoque_atual = EstoqueAtual(
                    id=f"CUR-{ingrediente.id}",
                    qtd=atualizacao.new_qty,
                    data=_now_recife().date(),
                )
        elif contagem is None:
            ingrediente.estoque_atual.qtd = atualizacao.new_qty
            ingrediente.estoque_atual.data = _now_recife().date()
        count += 1

    db.commit()
    return ResultadoLote(
        ok=True,
        atualizados=count,
        contagem_id=contagem.id if contagem is not None else None,
    )


@app.patch("/api/ingredientes/{ingrediente_id}", response_model=IngredienteOut)
def update_ingrediente(
    ingrediente_id: str,
    atualizacao: AtualizacaoIngrediente,
    db: Session = Depends(get_db),
):
    ingrediente = db.query(Ingrediente).filter(Ingrediente.id == ingrediente_id).first()
    if not ingrediente:
        raise HTTPException(status_code=404, detail="Ingrediente não encontrado")

    payload = atualizacao.model_dump(exclude_none=True)
    category = payload.pop("category", None)
    category_id = payload.pop("category_id", None)
    payload.pop("price", None)
    payload.pop("min_qty", None)

    for field, value in payload.items():
        setattr(ingrediente, field, value)

    category_value = category_id or category
    if category_value:
        categoria = (
            db.query(Categoria)
            .filter(or_(Categoria.id == category_value, Categoria.name == category_value))
            .first()
        )
        if categoria is None:
            raise HTTPException(status_code=400, detail="Categoria não encontrada")
        ingrediente.category_id = categoria.id

    db.commit()
    db.refresh(ingrediente)
    return ingrediente


# ---------------------------------------------------------------------------
# Fornecedores
# ---------------------------------------------------------------------------


@app.get("/api/fornecedores", response_model=FornecedorListResponse)
def get_fornecedores(db: Session = Depends(get_db)):
    rows = (
        db.query(
            Fornecedor,
            func.count(FornecedorIngrediente.ingredient_id).label("item_count"),
            func.avg(FornecedorIngrediente.price).label("avg_price"),
        )
        .outerjoin(
            FornecedorIngrediente,
            FornecedorIngrediente.supplier_id == Fornecedor.id,
        )
        .group_by(Fornecedor.id)
        .order_by(Fornecedor.name.asc())
        .all()
    )

    items = [
        FornecedorListItem(
            id=fornecedor.id,
            name=fornecedor.name,
            cnpj=fornecedor.cnpj,
            email=fornecedor.email,
            phone=fornecedor.phone,
            avg_delivery_time=fornecedor.avg_delivery_time,
            item_count=int(item_count or 0),
            avg_price=_as_float(avg_price) if avg_price is not None else None,
        )
        for fornecedor, item_count, avg_price in rows
    ]

    supplier_count = len(items)
    delivery_values = [
        item.avg_delivery_time for item in items if item.avg_delivery_time is not None
    ]
    avg_delivery_time = (
        sum(delivery_values) / len(delivery_values) if delivery_values else 0
    )
    avg_items_per_supplier = (
        sum(item.item_count for item in items) / supplier_count if supplier_count else 0
    )

    scored = [
        item
        for item in items
        if item.avg_price is not None and item.avg_delivery_time is not None
    ]
    best_supplier = None
    if scored:
        min_price = min(item.avg_price for item in scored if item.avg_price is not None)
        max_price = max(item.avg_price for item in scored if item.avg_price is not None)
        min_delivery = min(item.avg_delivery_time for item in scored if item.avg_delivery_time is not None)
        max_delivery = max(item.avg_delivery_time for item in scored if item.avg_delivery_time is not None)

        def score(item: FornecedorListItem) -> float:
            price_range = max_price - min_price
            delivery_range = max_delivery - min_delivery
            price_score = (
                ((item.avg_price or 0) - min_price) / price_range
                if price_range > 0
                else 0
            )
            delivery_score = (
                ((item.avg_delivery_time or 0) - min_delivery) / delivery_range
                if delivery_range > 0
                else 0
            )
            return price_score + delivery_score

        best_supplier = min(scored, key=score)

    return FornecedorListResponse(
        kpis=FornecedorKpis(
            supplier_count=supplier_count,
            avg_delivery_time=round(avg_delivery_time, 2),
            avg_items_per_supplier=round(avg_items_per_supplier, 2),
            best_value_supplier_id=best_supplier.id if best_supplier else None,
            best_value_supplier_name=best_supplier.name if best_supplier else None,
            best_value_detail=(
                f"R$ {best_supplier.avg_price:.2f} médio, "
                f"{best_supplier.avg_delivery_time} "
                f"{'dia' if best_supplier.avg_delivery_time == 1 else 'dias'}"
                if best_supplier
                else "Sem dados suficientes"
            ),
        ),
        items=items,
    )


@app.get("/api/export/fornecedores")
def export_fornecedores(
    format: str = Query(default="csv"),
    theme: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(
            Fornecedor,
            func.count(FornecedorIngrediente.ingredient_id).label("item_count"),
            func.avg(FornecedorIngrediente.price).label("avg_price"),
        )
        .outerjoin(
            FornecedorIngrediente,
            FornecedorIngrediente.supplier_id == Fornecedor.id,
        )
        .group_by(Fornecedor.id)
        .order_by(Fornecedor.name.asc())
        .all()
    )
    data = [
        {
            "id": fornecedor.id,
            "nome": fornecedor.name,
            "cnpj": fornecedor.cnpj,
            "email": fornecedor.email,
            "telefone": fornecedor.phone,
            "prazo_medio_entrega_dias": _as_float(fornecedor.avg_delivery_time)
            if fornecedor.avg_delivery_time is not None
            else None,
            "itens_fornecidos": int(item_count or 0),
            "preco_medio": _as_float(avg_price) if avg_price is not None else None,
        }
        for fornecedor, item_count, avg_price in rows
    ]
    return _export_response(
        data,
        "fornecedores",
        format,
        "Fornecedores cadastrados",
        [
            "id",
            "nome",
            "cnpj",
            "email",
            "telefone",
            "prazo_medio_entrega_dias",
            "itens_fornecidos",
            "preco_medio",
        ],
        theme_id=theme,
    )


@app.post("/api/fornecedores", response_model=FornecedorOut)
def create_fornecedor(payload: FornecedorCreate, db: Session = Depends(get_db)):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nome do fornecedor é obrigatório")

    last_id = (
        db.query(Fornecedor.id)
        .filter(Fornecedor.id.like("FOR%"))
        .order_by(Fornecedor.id.desc())
        .first()
    )
    next_number = 1
    if last_id:
        try:
            next_number = int(last_id[0].replace("FOR", "")) + 1
        except ValueError:
            next_number = db.query(func.count(Fornecedor.id)).scalar() + 1

    supplier_id = f"FOR{next_number:04d}"
    while db.query(Fornecedor).filter(Fornecedor.id == supplier_id).first():
        next_number += 1
        supplier_id = f"FOR{next_number:04d}"

    ingredient_ids = [item.ingredient_id for item in payload.ingredients]
    existing_ingredients = {
        ingredient.id
        for ingredient in db.query(Ingrediente)
        .filter(Ingrediente.id.in_(ingredient_ids))
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .all()
    }
    missing = sorted(set(ingredient_ids) - existing_ingredients)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Ingredientes inválidos para fornecedor: {', '.join(missing)}",
        )

    fornecedor = Fornecedor(
        id=supplier_id,
        name=name,
        cnpj=payload.cnpj,
        email=payload.email,
        phone=payload.phone,
        avg_delivery_time=payload.avg_delivery_time,
    )
    db.add(fornecedor)
    for item in payload.ingredients:
        db.add(
            FornecedorIngrediente(
                supplier_id=supplier_id,
                ingredient_id=item.ingredient_id,
                price=item.price,
                discount_percent=item.discount_percent,
                min_to_discount=item.min_to_discount,
            )
        )

    db.commit()
    db.refresh(fornecedor)
    return fornecedor


@app.get("/api/fornecedores/{fornecedor_id}", response_model=FornecedorProfileResponse)
def get_fornecedor_profile(fornecedor_id: str, db: Session = Depends(get_db)):
    fornecedor = db.query(Fornecedor).filter(Fornecedor.id == fornecedor_id).first()
    if fornecedor is None:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado")

    product_rows = (
        db.query(
            Ingrediente.id,
            Ingrediente.name,
            Categoria.name.label("category"),
            EstoqueAtual.qtd,
            Ingrediente.unit,
            FornecedorIngrediente.price,
        )
        .join(FornecedorIngrediente, FornecedorIngrediente.ingredient_id == Ingrediente.id)
        .join(Categoria, Categoria.id == Ingrediente.category_id)
        .outerjoin(EstoqueAtual, EstoqueAtual.ingrediente == Ingrediente.id)
        .filter(FornecedorIngrediente.supplier_id == fornecedor_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .order_by(Ingrediente.name.asc())
        .all()
    )
    products = [
        FornecedorProductOut(
            ingredient_id=ingredient_id,
            name=name,
            category=category,
            current_qty=_as_float(qtd),
            unit=unit,
            unit_price=_as_float(price),
        )
        for ingredient_id, name, category, qtd, unit, price in product_rows
    ]

    order_rows = (
        db.query(
            Pedido.id,
            Pedido.data_pedido,
            func.sum(Pedido.qty).label("items_qty"),
            func.sum(Pedido.valor).label("total_value"),
            Pedido.status,
        )
        .join(Ingrediente, Ingrediente.id == Pedido.ingredient_id)
        .filter(Pedido.supplier_id == fornecedor_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .group_by(Pedido.id, Pedido.data_pedido, Pedido.status)
        .order_by(Pedido.data_pedido.desc(), Pedido.id.desc())
        .all()
    )
    orders = [
        FornecedorOrderOut(
            id=order_id,
            order_date=order_date,
            items_qty=_as_float(items_qty),
            total_value=_as_float(total_value),
            status=status,
        )
        for order_id, order_date, items_qty, total_value, status in order_rows
    ]

    lead_time = (
        db.query(func.avg(Pedido.data_prevista - Pedido.data_pedido))
        .join(Ingrediente, Ingrediente.id == Pedido.ingredient_id)
        .filter(Pedido.supplier_id == fornecedor_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .scalar()
    )
    orders_count = len(orders)
    delivered_count = sum(
        1 for order in orders if order.status.lower() == "entregue"
    )
    delivery_rate = (delivered_count / orders_count) * 100 if orders_count else 0

    return FornecedorProfileResponse(
        supplier=fornecedor,
        kpis=FornecedorProfileKpis(
            avg_lead_time=round(_as_float(lead_time), 2),
            orders_count=orders_count,
            delivery_rate=round(delivery_rate, 2),
        ),
        products=products,
        orders=orders,
    )


# ---------------------------------------------------------------------------
# Vendas
# ---------------------------------------------------------------------------


def _next_prefixed_id(
    db: Session,
    column,
    prefix: str,
    reserved_ids: Optional[set[str]] = None,
) -> str:
    reserved_ids = reserved_ids if reserved_ids is not None else set()
    last_id = (
        db.query(column)
        .filter(column.like(f"{prefix}%"))
        .order_by(column.desc())
        .first()
    )
    next_number = 1
    if last_id:
        try:
            next_number = int(str(last_id[0]).replace(prefix, "")) + 1
        except ValueError:
            next_number = (db.query(func.count(column)).scalar() or 0) + 1

    candidate = f"{prefix}{next_number:012d}"
    while candidate in reserved_ids or db.query(column).filter(column == candidate).first():
        next_number += 1
        candidate = f"{prefix}{next_number:012d}"
    reserved_ids.add(candidate)
    return candidate


def _next_cliente_id(db: Session, reserved_ids: Optional[set[str]] = None) -> str:
    return _next_prefixed_id(db, Cliente.id, "CLI", reserved_ids)


def _next_venda_transacao_id(db: Session, reserved_ids: Optional[set[str]] = None) -> str:
    return _next_prefixed_id(db, VendaTransacao.id, "VTR", reserved_ids)


def _next_venda_item_id(db: Session, reserved_ids: Optional[set[str]] = None) -> str:
    return _next_prefixed_id(db, VendaItem.id, "VTI", reserved_ids)


def _next_venda_pagamento_id(db: Session, reserved_ids: Optional[set[str]] = None) -> str:
    return _next_prefixed_id(db, VendaPagamento.id, "VPG", reserved_ids)


def _next_estoque_movimento_id(db: Session, reserved_ids: Optional[set[str]] = None) -> str:
    return _next_prefixed_id(db, EstoqueMovimento.id, "MOV", reserved_ids)


def _next_venda_documento_fiscal_id(db: Session, reserved_ids: Optional[set[str]] = None) -> str:
    return _next_prefixed_id(db, VendaDocumentoFiscal.id, "VDF", reserved_ids)


def _next_historical_venda_id(db: Session, reserved_ids: Optional[set[str]] = None) -> str:
    return _next_prefixed_id(db, Venda.id, "VEN", reserved_ids)


def _money(value: float) -> float:
    return round(max(float(value or 0), 0), 2)


def _paid_total(venda: VendaTransacao) -> float:
    return _money(
        sum(
            _as_float(payment.amount)
            for payment in venda.payments
            if payment.status == "pago"
        )
    )


def _serialize_cliente(cliente: Optional[Cliente]) -> Optional[ClienteOut]:
    if cliente is None:
        return None
    return ClienteOut(
        id=cliente.id,
        name=cliente.name,
        document=cliente.document,
        email=cliente.email,
        phone=cliente.phone,
        created_at=cliente.created_at,
    )


def _serialize_fiscal_document(
    document: Optional[VendaDocumentoFiscal],
) -> Optional[VendaFiscalDocumentOut]:
    if document is None:
        return None
    return VendaFiscalDocumentOut(
        id=document.id,
        venda_id=document.venda_id,
        document_type=document.document_type,
        status=document.status,
        provider=document.provider,
        access_key=document.access_key,
        protocol=document.protocol,
        issued_at=document.issued_at,
        cancelled_at=document.cancelled_at,
        payload=document.payload,
        error_message=document.error_message,
        created_at=document.created_at,
        updated_at=document.updated_at,
    )


def _latest_fiscal_document(venda: VendaTransacao) -> Optional[VendaDocumentoFiscal]:
    if not venda.fiscal_documents:
        return None
    return sorted(
        venda.fiscal_documents,
        key=lambda item: (item.created_at or datetime.min.replace(tzinfo=RECIFE_TZ), item.id),
    )[-1]


def _serialize_venda_detail(venda: VendaTransacao) -> VendaDetailOut:
    return VendaDetailOut(
        id=venda.id,
        date_time=venda.date_time,
        customer=_serialize_cliente(venda.cliente),
        status=venda.status,
        fiscal_status=venda.fiscal_status,
        subtotal=_as_float(venda.subtotal),
        discount_total=_as_float(venda.discount_total),
        total=_as_float(venda.total),
        source=venda.source,
        notes=venda.notes,
        confirmed_at=venda.confirmed_at,
        canceled_at=venda.canceled_at,
        items=[
            VendaItemOut(
                id=item.id,
                recipe_id=item.recipe_id,
                recipe_name=item.recipe_name,
                quantity=_as_float(item.quantity),
                unit_price=_as_float(item.unit_price),
                discount_value=_as_float(item.discount_value),
                total_value=_as_float(item.total_value),
                venda_historica_id=item.venda_historica_id,
            )
            for item in sorted(venda.items, key=lambda item: item.id)
        ],
        payments=[
            VendaPagamentoOut(
                id=payment.id,
                method=payment.method,
                amount=_as_float(payment.amount),
                status=payment.status,
                paid_at=payment.paid_at,
                change_amount=_as_float(payment.change_amount),
                external_reference=payment.external_reference,
            )
            for payment in sorted(venda.payments, key=lambda payment: payment.id)
        ],
        fiscal_document=_serialize_fiscal_document(_latest_fiscal_document(venda)),
    )


def _recipe_stock_profile(db: Session, recipe_id: str) -> tuple[int, bool, Optional[float], list[str]]:
    rows = (
        db.query(
            ReceitaIngrediente.ingredient_id,
            ReceitaIngrediente.qty,
            Ingrediente.name,
            Ingrediente.unit,
            EstoqueAtual.qtd,
        )
        .join(Ingrediente, Ingrediente.id == ReceitaIngrediente.ingredient_id)
        .outerjoin(EstoqueAtual, EstoqueAtual.ingrediente == ReceitaIngrediente.ingredient_id)
        .filter(ReceitaIngrediente.recipe_id == recipe_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .all()
    )
    warnings: list[str] = []
    max_quantity: Optional[float] = None
    for row in rows:
        required = _as_float(row.qty)
        current_qty = _as_float(row.qtd)
        if required <= 0:
            warnings.append(f"{row.name}: ficha tecnica sem consumo valido")
            continue
        possible = current_qty / required
        max_quantity = possible if max_quantity is None else min(max_quantity, possible)
        if current_qty <= 0:
            warnings.append(f"{row.name}: sem estoque")

    if not rows:
        warnings.append("Produto sem ficha tecnica de ingredientes")
    available = bool(rows) and max_quantity is not None and max_quantity >= 1
    return len(rows), available, round(max_quantity, 2) if max_quantity is not None else None, warnings


def _serialize_venda_produto(db: Session, receita: Receita) -> VendaProdutoOut:
    ingredients_count, available, max_quantity, warnings = _recipe_stock_profile(
        db,
        receita.id,
    )
    return VendaProdutoOut(
        id=receita.id,
        name=receita.name,
        recipe_type=receita.recipe_type,
        sale_price=_as_float(receita.sale_price),
        yield_qty=_as_float(receita.yield_qty) if receita.yield_qty is not None else None,
        yield_unit=receita.yield_unit,
        ingredients_count=ingredients_count,
        available=available,
        max_quantity=max_quantity,
        stock_warnings=warnings,
    )


def _resolve_cliente(
    db: Session,
    customer_id: Optional[str],
    customer: Optional[ClienteCreate],
) -> Optional[Cliente]:
    if customer_id:
        cliente = db.query(Cliente).filter(Cliente.id == customer_id).first()
        if cliente is None:
            raise HTTPException(status_code=404, detail="Cliente nao encontrado")
        return cliente
    if customer is None:
        return None

    cliente = Cliente(
        id=_next_cliente_id(db),
        name=customer.name,
        document=customer.document,
        email=customer.email,
        phone=customer.phone,
    )
    db.add(cliente)
    return cliente


def _build_venda_items(
    db: Session,
    payload_items: list,
    discount_total: float,
) -> tuple[list[dict], float, float]:
    recipe_ids = {item.recipe_id for item in payload_items}
    recipes = {
        receita.id: receita
        for receita in (
            db.query(Receita)
            .filter(Receita.id.in_(recipe_ids))
            .filter(Receita.recipe_type == "PRODUTO_FINAL")
            .all()
        )
    }
    missing = sorted(recipe_ids - set(recipes.keys()))
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Produtos invalidos para venda: {', '.join(missing)}",
        )

    rows: list[dict] = []
    subtotal = 0.0
    item_discount_total = 0.0
    for item in payload_items:
        receita = recipes[item.recipe_id]
        unit_price = _as_float(item.unit_price) if item.unit_price is not None else _as_float(receita.sale_price)
        if unit_price <= 0:
            raise HTTPException(
                status_code=400,
                detail=f"Produto sem preco de venda valido: {receita.name}",
            )
        gross_total = _money(unit_price * item.quantity)
        if item.discount_value > gross_total:
            raise HTTPException(
                status_code=400,
                detail=f"Desconto maior que o item: {receita.name}",
            )
        line_total = _money(gross_total - item.discount_value)
        subtotal += gross_total
        item_discount_total += item.discount_value
        rows.append(
            {
                "recipe": receita,
                "quantity": item.quantity,
                "unit_price": unit_price,
                "discount_value": item.discount_value,
                "total_value": line_total,
            }
        )

    subtotal = _money(subtotal)
    total_before_sale_discount = _money(subtotal - item_discount_total)
    if discount_total > total_before_sale_discount:
        raise HTTPException(
            status_code=400,
            detail="Desconto total maior que o valor da venda",
        )
    total = _money(total_before_sale_discount - discount_total)
    return rows, subtotal, total


def _replace_venda_payments(
    db: Session,
    venda: VendaTransacao,
    payments: list[VendaPagamentoCreate],
) -> None:
    venda.payments.clear()
    db.flush()
    reserved_ids: set[str] = set()
    paid_at = _now_recife()
    for payment in payments:
        venda.payments.append(
            VendaPagamento(
                id=_next_venda_pagamento_id(db, reserved_ids),
                method=payment.method,
                amount=payment.amount,
                status=payment.status,
                paid_at=paid_at if payment.status == "pago" else None,
                change_amount=payment.change_amount,
                external_reference=payment.external_reference,
            )
        )


def _required_ingredients_for_venda(db: Session, venda_id: str):
    return (
        db.query(
            ReceitaIngrediente.ingredient_id,
            Ingrediente.name.label("ingredient_name"),
            Ingrediente.unit,
            func.sum(VendaItem.quantity * ReceitaIngrediente.qty).label("required_qty"),
        )
        .join(Ingrediente, Ingrediente.id == ReceitaIngrediente.ingredient_id)
        .join(VendaItem, VendaItem.recipe_id == ReceitaIngrediente.recipe_id)
        .filter(VendaItem.venda_id == venda_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .group_by(ReceitaIngrediente.ingredient_id, Ingrediente.name, Ingrediente.unit)
        .all()
    )


def _apply_venda_stock_movements(db: Session, venda: VendaTransacao) -> None:
    requirements = _required_ingredients_for_venda(db, venda.id)
    if not requirements:
        raise HTTPException(
            status_code=400,
            detail="Venda sem ficha tecnica para baixa de estoque",
        )

    ingredient_ids = [row.ingredient_id for row in requirements]
    stock_rows = (
        db.query(EstoqueAtual)
        .filter(EstoqueAtual.ingrediente.in_(ingredient_ids))
        .with_for_update()
        .all()
    )
    stock_by_ingredient = {row.ingrediente: row for row in stock_rows}

    insufficient: list[dict] = []
    for row in requirements:
        required_qty = _as_float(row.required_qty)
        stock = stock_by_ingredient.get(row.ingredient_id)
        current_qty = _as_float(stock.qtd) if stock else 0.0
        if required_qty <= 0:
            insufficient.append(
                {
                    "ingredient_id": row.ingredient_id,
                    "ingredient_name": row.ingredient_name,
                    "required_qty": required_qty,
                    "current_qty": current_qty,
                    "unit": row.unit,
                    "reason": "consumo invalido",
                }
            )
        elif current_qty + 1e-9 < required_qty:
            insufficient.append(
                {
                    "ingredient_id": row.ingredient_id,
                    "ingredient_name": row.ingredient_name,
                    "required_qty": required_qty,
                    "current_qty": current_qty,
                    "unit": row.unit,
                    "reason": "estoque insuficiente",
                }
            )

    if insufficient:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Estoque insuficiente para confirmar a venda",
                "items": insufficient,
            },
        )

    reserved_movement_ids: set[str] = set()
    movement_date = _now_recife().date()
    for row in requirements:
        stock = stock_by_ingredient[row.ingredient_id]
        required_qty = _as_float(row.required_qty)
        previous_qty = _as_float(stock.qtd)
        new_qty = round(previous_qty - required_qty, 4)
        stock.qtd = new_qty
        stock.data = movement_date
        db.add(
            EstoqueMovimento(
                id=_next_estoque_movimento_id(db, reserved_movement_ids),
                ingredient_id=row.ingredient_id,
                source_type="venda",
                source_id=venda.id,
                delta_qty=-required_qty,
                previous_qty=previous_qty,
                new_qty=new_qty,
                unit=row.unit,
                reason="Baixa por venda de balcao",
            )
        )


def _insert_historical_sales(db: Session, venda: VendaTransacao) -> None:
    reserved_historical_ids: set[str] = set()
    sale_time = venda.date_time
    if sale_time.tzinfo is not None:
        sale_time = sale_time.astimezone(RECIFE_TZ).replace(tzinfo=None)
    for item in venda.items:
        if item.venda_historica_id:
            continue
        historical_id = _next_historical_venda_id(db, reserved_historical_ids)
        db.add(
            Venda(
                id=historical_id,
                date_time=sale_time,
                recipe_id=item.recipe_id,
                quantity=item.quantity,
                unit_price=item.unit_price,
            )
        )
        item.venda_historica_id = historical_id


def _fiscal_payload(venda: VendaTransacao) -> dict:
    return {
        "document_type": "NFC-e",
        "integration_status": "prepared_only",
        "sale": {
            "id": venda.id,
            "date_time": _to_recife_export_datetime(venda.date_time),
            "subtotal": _as_float(venda.subtotal),
            "discount_total": _as_float(venda.discount_total),
            "total": _as_float(venda.total),
            "source": venda.source,
        },
        "customer": (
            {
                "id": venda.cliente.id,
                "name": venda.cliente.name,
                "document": venda.cliente.document,
                "email": venda.cliente.email,
                "phone": venda.cliente.phone,
            }
            if venda.cliente
            else None
        ),
        "items": [
            {
                "recipe_id": item.recipe_id,
                "name": item.recipe_name,
                "quantity": _as_float(item.quantity),
                "unit_price": _as_float(item.unit_price),
                "discount_value": _as_float(item.discount_value),
                "total_value": _as_float(item.total_value),
            }
            for item in venda.items
        ],
        "payments": [
            {
                "method": payment.method,
                "amount": _as_float(payment.amount),
                "status": payment.status,
                "change_amount": _as_float(payment.change_amount),
            }
            for payment in venda.payments
        ],
    }


def _prepare_fiscal_document(db: Session, venda: VendaTransacao) -> None:
    document = _latest_fiscal_document(venda)
    now = _now_recife()
    if document is None:
        document = VendaDocumentoFiscal(
            id=_next_venda_documento_fiscal_id(db),
            venda_id=venda.id,
            document_type="NFC-e",
        )
        venda.fiscal_documents.append(document)
    document.status = "pronto_para_integracao"
    document.updated_at = now
    document.payload = _fiscal_payload(venda)
    document.error_message = None
    venda.fiscal_status = document.status


def _to_recife_export_datetime(value: Optional[datetime]) -> Optional[str]:
    if value is None:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=RECIFE_TZ)
    return value.astimezone(RECIFE_TZ).isoformat()


def _get_venda_or_404(db: Session, venda_id: str) -> VendaTransacao:
    venda = db.query(VendaTransacao).filter(VendaTransacao.id == venda_id).first()
    if venda is None:
        raise HTTPException(status_code=404, detail="Venda nao encontrada")
    return venda


@app.get("/api/vendas/produtos", response_model=list[VendaProdutoOut])
def get_venda_produtos(
    q: Optional[str] = None,
    db: Session = Depends(get_db),
):
    query = (
        db.query(Receita)
        .filter(Receita.recipe_type == "PRODUTO_FINAL")
        .filter(Receita.sale_price.isnot(None))
        .filter(Receita.sale_price > 0)
        .order_by(Receita.name.asc())
    )
    if q:
        query = query.filter(Receita.name.ilike(f"%{q}%"))
    receitas = query.limit(250).all()
    return [_serialize_venda_produto(db, receita) for receita in receitas]


@app.get("/api/vendas/clientes", response_model=list[ClienteOut])
def get_clientes(
    q: Optional[str] = None,
    db: Session = Depends(get_db),
):
    query = db.query(Cliente).order_by(Cliente.name.asc())
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Cliente.name.ilike(like),
                Cliente.document.ilike(like),
                Cliente.email.ilike(like),
                Cliente.phone.ilike(like),
            )
        )
    return [_serialize_cliente(cliente) for cliente in query.limit(100).all()]


@app.post("/api/vendas/clientes", response_model=ClienteOut)
def create_cliente(payload: ClienteCreate, db: Session = Depends(get_db)):
    cliente = Cliente(
        id=_next_cliente_id(db),
        name=payload.name,
        document=payload.document,
        email=payload.email,
        phone=payload.phone,
    )
    db.add(cliente)
    db.commit()
    db.refresh(cliente)
    return _serialize_cliente(cliente)


@app.post("/api/vendas", response_model=VendaDetailOut)
def create_venda(payload: VendaCreateRequest, db: Session = Depends(get_db)):
    cliente = _resolve_cliente(db, payload.customer_id, payload.customer)
    item_rows, subtotal, total = _build_venda_items(
        db,
        payload.items,
        payload.discount_total,
    )
    now = _now_recife()
    venda = VendaTransacao(
        id=_next_venda_transacao_id(db),
        date_time=now,
        cliente=cliente,
        status="aberta",
        subtotal=subtotal,
        discount_total=payload.discount_total,
        total=total,
        source=payload.source,
        fiscal_status="pendente_preparacao",
        notes=payload.notes,
        created_at=now,
        updated_at=now,
    )
    db.add(venda)
    db.flush()

    reserved_item_ids: set[str] = set()
    for row in item_rows:
        receita = row["recipe"]
        venda.items.append(
            VendaItem(
                id=_next_venda_item_id(db, reserved_item_ids),
                recipe_id=receita.id,
                recipe_name=receita.name,
                quantity=row["quantity"],
                unit_price=row["unit_price"],
                discount_value=row["discount_value"],
                total_value=row["total_value"],
            )
        )
    if payload.payments:
        _replace_venda_payments(db, venda, payload.payments)

    db.commit()
    db.refresh(venda)
    return _serialize_venda_detail(venda)


@app.post("/api/vendas/{venda_id}/confirmar", response_model=VendaDetailOut)
def confirm_venda(
    venda_id: str,
    payload: VendaConfirmRequest,
    db: Session = Depends(get_db),
):
    venda = _get_venda_or_404(db, venda_id)
    if venda.status == "cancelada":
        raise HTTPException(status_code=400, detail="Venda cancelada nao pode ser confirmada")
    if venda.status == "paga":
        return _serialize_venda_detail(venda)

    if payload.payments:
        _replace_venda_payments(db, venda, payload.payments)
        db.flush()
    if not venda.payments:
        raise HTTPException(status_code=400, detail="Informe ao menos um pagamento")
    if _paid_total(venda) + 1e-9 < _as_float(venda.total):
        raise HTTPException(status_code=400, detail="Pagamento menor que o total da venda")

    _apply_venda_stock_movements(db, venda)
    _insert_historical_sales(db, venda)
    venda.status = "paga"
    venda.confirmed_at = _now_recife()
    venda.updated_at = venda.confirmed_at
    _prepare_fiscal_document(db, venda)

    db.commit()
    db.refresh(venda)
    return _serialize_venda_detail(venda)


@app.patch("/api/vendas/{venda_id}/cancelar", response_model=VendaDetailOut)
def cancel_venda(venda_id: str, db: Session = Depends(get_db)):
    venda = _get_venda_or_404(db, venda_id)
    if venda.status == "cancelada":
        return _serialize_venda_detail(venda)

    now = _now_recife()
    reserved_movement_ids: set[str] = set()
    movements = (
        db.query(EstoqueMovimento)
        .filter(EstoqueMovimento.source_type == "venda")
        .filter(EstoqueMovimento.source_id == venda.id)
        .all()
    )
    if movements:
        ingredient_ids = sorted({movement.ingredient_id for movement in movements})
        stock_rows = (
            db.query(EstoqueAtual)
            .filter(EstoqueAtual.ingrediente.in_(ingredient_ids))
            .with_for_update()
            .all()
        )
        stock_by_ingredient = {row.ingrediente: row for row in stock_rows}
        ingredient_units = {
            ingredient.id: ingredient.unit
            for ingredient in db.query(Ingrediente)
            .filter(Ingrediente.id.in_(ingredient_ids))
            .all()
        }
        delta_by_ingredient: dict[str, float] = {}
        for movement in movements:
            delta_by_ingredient[movement.ingredient_id] = (
                delta_by_ingredient.get(movement.ingredient_id, 0.0)
                + _as_float(movement.delta_qty)
            )
        for ingredient_id, original_delta in delta_by_ingredient.items():
            revert_qty = round(-original_delta, 4)
            stock = stock_by_ingredient.get(ingredient_id)
            previous_qty = _as_float(stock.qtd) if stock else 0.0
            new_qty = round(previous_qty + revert_qty, 4)
            if stock is None:
                stock = EstoqueAtual(
                    id=f"CUR-{ingredient_id}",
                    ingrediente=ingredient_id,
                    qtd=new_qty,
                    data=now.date(),
                )
                db.add(stock)
            else:
                stock.qtd = new_qty
                stock.data = now.date()
            db.add(
                EstoqueMovimento(
                    id=_next_estoque_movimento_id(db, reserved_movement_ids),
                    ingredient_id=ingredient_id,
                    source_type="cancelamento_venda",
                    source_id=venda.id,
                    delta_qty=revert_qty,
                    previous_qty=previous_qty,
                    new_qty=new_qty,
                    unit=ingredient_units.get(ingredient_id, ""),
                    reason="Estorno por cancelamento de venda",
                )
            )

    for item in venda.items:
        if item.venda_historica_id:
            db.query(Venda).filter(Venda.id == item.venda_historica_id).delete()
            item.venda_historica_id = None

    for payment in venda.payments:
        if payment.status == "pago":
            payment.status = "estornado"

    for document in venda.fiscal_documents:
        document.status = "cancelado"
        document.cancelled_at = now
        document.updated_at = now
    venda.status = "cancelada"
    venda.fiscal_status = "cancelado"
    venda.canceled_at = now
    venda.updated_at = now

    db.commit()
    db.refresh(venda)
    return _serialize_venda_detail(venda)


@app.get("/api/vendas", response_model=VendaPaginado)
def get_vendas(
    status: Optional[str] = None,
    q: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=100),
    db: Session = Depends(get_db),
):
    if date_from and date_to and date_from > date_to:
        raise HTTPException(status_code=400, detail="Data inicial maior que data final")
    query = db.query(VendaTransacao).outerjoin(Cliente)
    if status:
        query = query.filter(VendaTransacao.status == status)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(VendaTransacao.id.ilike(like), Cliente.name.ilike(like)))
    if date_from:
        query = query.filter(func.date(VendaTransacao.date_time) >= date_from)
    if date_to:
        query = query.filter(func.date(VendaTransacao.date_time) <= date_to)

    total = query.count()
    total_pages = max(1, math.ceil(total / page_size))
    vendas = (
        query.order_by(VendaTransacao.date_time.desc(), VendaTransacao.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    items = [
        VendaListItem(
            id=venda.id,
            date_time=venda.date_time,
            customer_name=venda.cliente.name if venda.cliente else None,
            status=venda.status,
            fiscal_status=venda.fiscal_status,
            items_count=len(venda.items),
            items_qty=sum(_as_float(item.quantity) for item in venda.items),
            total=_as_float(venda.total),
            paid_total=_paid_total(venda),
        )
        for venda in vendas
    ]
    return VendaPaginado(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
    )


@app.get("/api/vendas/{venda_id}/fiscal", response_model=VendaFiscalDocumentOut)
def get_venda_fiscal_document(venda_id: str, db: Session = Depends(get_db)):
    venda = _get_venda_or_404(db, venda_id)
    document = _latest_fiscal_document(venda)
    if document is None:
        raise HTTPException(status_code=404, detail="Documento fiscal nao preparado")
    return _serialize_fiscal_document(document)


@app.get("/api/vendas/{venda_id}", response_model=VendaDetailOut)
def get_venda_detail(venda_id: str, db: Session = Depends(get_db)):
    return _serialize_venda_detail(_get_venda_or_404(db, venda_id))


# ---------------------------------------------------------------------------
# Pedidos
# ---------------------------------------------------------------------------


def _pedido_group_key(supplier_id: str, order_date: date) -> str:
    return f"{supplier_id}_{order_date.isoformat()}"


def _discount_rate(discount_percent) -> float:
    value = _as_float(discount_percent)
    return value / 100 if value > 1 else value


def _effective_unit_price(
    price,
    discount_percent,
    min_to_discount,
    qty: float,
) -> tuple[float, bool]:
    unit_price = _as_float(price)
    min_qty = _as_float(min_to_discount)
    rate = _discount_rate(discount_percent)
    applies = rate > 0 and qty >= min_qty
    if applies:
        unit_price *= 1 - rate
    return round(unit_price, 4), applies


def _option_total(price, discount_percent, min_to_discount, qty: float) -> float:
    effective_price, _ = _effective_unit_price(
        price,
        discount_percent,
        min_to_discount,
        qty,
    )
    return round(effective_price * qty, 2)


def _next_pedido_id(db: Session, reserved_ids: Optional[set[str]] = None) -> str:
    reserved_ids = reserved_ids if reserved_ids is not None else set()
    last_id = (
        db.query(Pedido.id)
        .filter(Pedido.id.like("PED%"))
        .order_by(Pedido.id.desc())
        .first()
    )
    next_number = 1
    if last_id:
        try:
            next_number = int(last_id[0].replace("PED", "")) + 1
        except ValueError:
            next_number = db.query(func.count(Pedido.id)).scalar() + 1

    pedido_id = f"PED{next_number:012d}"
    while pedido_id in reserved_ids or db.query(Pedido).filter(Pedido.id == pedido_id).first():
        next_number += 1
        pedido_id = f"PED{next_number:012d}"
    reserved_ids.add(pedido_id)
    return pedido_id


def _add_pedidos_to_estoque_atual(
    db: Session,
    pedidos: list[Pedido],
    delivery_date: date,
) -> None:
    received_by_ingredient: dict[str, float] = {}
    for pedido in pedidos:
        received_by_ingredient[pedido.ingredient_id] = (
            received_by_ingredient.get(pedido.ingredient_id, 0.0)
            + _as_float(pedido.qty)
        )

    for ingredient_id, received_qty in received_by_ingredient.items():
        db.execute(
            text(
                """
                INSERT INTO estoque_atual (id, ingrediente, qtd, data)
                VALUES (:id, :ingrediente, :qtd, :data)
                ON CONFLICT (ingrediente) DO UPDATE
                SET qtd = estoque_atual.qtd + EXCLUDED.qtd,
                    data = EXCLUDED.data
                """
            ),
            {
                "id": f"CUR-{ingredient_id}",
                "ingrediente": ingredient_id,
                "qtd": received_qty,
                "data": delivery_date,
            },
        )


def _pedidos_without_stock_application(pedidos: list[Pedido]) -> list[Pedido]:
    return [pedido for pedido in pedidos if pedido.estoque_aplicado_em is None]


def _mark_pedidos_delivered_and_apply_stock(
    db: Session,
    pedidos: list[Pedido],
) -> None:
    delivery_time = _now_recife()
    pedidos_to_apply = _pedidos_without_stock_application(pedidos)
    _add_pedidos_to_estoque_atual(db, pedidos_to_apply, delivery_time.date())

    for pedido in pedidos:
        if pedido.estoque_aplicado_em is None:
            pedido.estoque_aplicado_em = delivery_time
        pedido.status = "entregue"


def _pedido_base_query(db: Session):
    return (
        db.query(
            Pedido.id,
            Pedido.supplier_id,
            Fornecedor.name.label("supplier_name"),
            Pedido.ingredient_id,
            Ingrediente.name.label("ingredient_name"),
            Pedido.data_pedido,
            Pedido.qty,
            Pedido.valor,
            Pedido.status,
            Pedido.data_prevista,
        )
        .join(Fornecedor, Fornecedor.id == Pedido.supplier_id)
        .join(Ingrediente, Ingrediente.id == Pedido.ingredient_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
    )


def _apply_pedido_filters(
    query,
    status: Optional[str],
    supplier_id: Optional[str],
    date_from: Optional[date],
    date_to: Optional[date],
):
    if status:
        query = query.filter(Pedido.status == status)
    if supplier_id:
        query = query.filter(Pedido.supplier_id == supplier_id)
    if date_from:
        query = query.filter(Pedido.data_pedido >= date_from)
    if date_to:
        query = query.filter(Pedido.data_pedido <= date_to)
    return query


def _pedido_group_query(db: Session):
    return (
        db.query(
            Pedido.supplier_id,
            Fornecedor.name.label("supplier_name"),
            Pedido.data_pedido,
            func.max(Pedido.data_prevista).label("data_prevista"),
            func.count(func.distinct(Pedido.ingredient_id)).label("ingredients_count"),
            func.sum(Pedido.qty).label("items_qty"),
            func.sum(Pedido.valor).label("total_value"),
            func.sum(
                case((Pedido.status == "em_transito", 1), else_=0)
            ).label("transit_count"),
        )
        .join(Fornecedor, Fornecedor.id == Pedido.supplier_id)
        .join(Ingrediente, Ingrediente.id == Pedido.ingredient_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .group_by(Pedido.supplier_id, Fornecedor.name, Pedido.data_pedido)
    )


def _serialize_pedido_group(row) -> PedidoGroupOut:
    return PedidoGroupOut(
        group_key=_pedido_group_key(row.supplier_id, row.data_pedido),
        supplier_id=row.supplier_id,
        supplier_name=row.supplier_name,
        order_date=row.data_pedido,
        expected_date=row.data_prevista,
        status="em_transito" if int(row.transit_count or 0) > 0 else "entregue",
        ingredients_count=int(row.ingredients_count or 0),
        items_qty=_as_float(row.items_qty),
        total_value=_as_float(row.total_value),
    )


def _get_pedido_group(
    db: Session,
    supplier_id: str,
    order_date: date,
    status: Optional[str] = None,
) -> Optional[PedidoGroupOut]:
    row = _apply_pedido_filters(
        _pedido_group_query(db),
        status,
        supplier_id,
        order_date,
        order_date,
    ).first()
    return _serialize_pedido_group(row) if row else None


def _serialize_pedido(row) -> PedidoOut:
    return PedidoOut(
        id=row.id,
        supplier_id=row.supplier_id,
        supplier_name=row.supplier_name,
        ingredient_id=row.ingredient_id,
        ingredient_name=row.ingredient_name,
        order_date=row.data_pedido,
        items_qty=_as_float(row.qty),
        total_value=_as_float(row.valor),
        status=row.status,
        expected_date=row.data_prevista,
    )


def _send_pedido_emails(
    email_groups: dict[str, dict],
    order_date: date,
) -> list[PedidoEmailResult]:
    settings_error = None
    try:
        settings = get_smtp_settings()
    except Exception as exc:
        logger.exception("Configuracao SMTP invalida")
        settings = None
        settings_error = str(exc)
    results: list[PedidoEmailResult] = []

    for supplier_id, group in sorted(
        email_groups.items(),
        key=lambda item: item[1]["supplier_name"],
    ):
        supplier_name = group["supplier_name"]
        supplier_email = group.get("email")

        if not supplier_email:
            results.append(
                PedidoEmailResult(
                    supplier_id=supplier_id,
                    supplier_name=supplier_name,
                    email=supplier_email,
                    status="missing_email",
                    message="Fornecedor sem email cadastrado.",
                )
            )
            continue

        if settings_error is not None:
            results.append(
                PedidoEmailResult(
                    supplier_id=supplier_id,
                    supplier_name=supplier_name,
                    email=supplier_email,
                    status="failed",
                    message=f"Configuracao SMTP invalida: {settings_error}",
                )
            )
            continue

        if settings is None:
            results.append(
                PedidoEmailResult(
                    supplier_id=supplier_id,
                    supplier_name=supplier_name,
                    email=supplier_email,
                    status="disabled",
                    message="SMTP nao configurado.",
                )
            )
            continue

        order_email = OrderEmail(
            supplier_name=supplier_name,
            supplier_email=supplier_email,
            order_date=order_date,
            expected_date=group["expected_date"],
            items=group["items"],
        )
        try:
            send_order_email(order_email, settings)
        except Exception as exc:
            logger.exception(
                "Falha ao enviar email do pedido para fornecedor %s",
                supplier_id,
            )
            results.append(
                PedidoEmailResult(
                    supplier_id=supplier_id,
                    supplier_name=supplier_name,
                    email=supplier_email,
                    status="failed",
                    message=f"Falha ao enviar email: {exc}",
                )
            )
            continue

        results.append(
            PedidoEmailResult(
                supplier_id=supplier_id,
                supplier_name=supplier_name,
                email=supplier_email,
                status="sent",
                message="Email enviado com sucesso.",
            )
        )

    return results


# ---------------------------------------------------------------------------
# Centro de Compras Autônomo Maestro
# ---------------------------------------------------------------------------


def _purchase_recent_usage_window(db: Session, days: int, reference_date: date) -> dict[str, float]:
    horizon_days = max(1, days)
    end_date = reference_date - timedelta(days=1)
    start_date = reference_date - timedelta(days=horizon_days)
    if start_date > end_date:
        return {}
    rows = (
        db.query(
            ReceitaIngrediente.ingredient_id,
            func.sum(Venda.quantity * ReceitaIngrediente.qty).label("usage_qty"),
        )
        .select_from(Venda)
        .join(Receita, Receita.id == Venda.recipe_id)
        .join(ReceitaIngrediente, ReceitaIngrediente.recipe_id == Receita.id)
        .join(Ingrediente, Ingrediente.id == ReceitaIngrediente.ingredient_id)
        .filter(func.date(Venda.date_time) >= start_date)
        .filter(func.date(Venda.date_time) <= end_date)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .group_by(ReceitaIngrediente.ingredient_id)
        .all()
    )
    return {
        ingredient_id: max(0.0, _as_float(usage_qty))
        for ingredient_id, usage_qty in rows
    }


def _purchase_in_transit_map(db: Session) -> dict[str, float]:
    rows = (
        db.query(Pedido.ingredient_id, func.sum(Pedido.qty))
        .join(Ingrediente, Ingrediente.id == Pedido.ingredient_id)
        .filter(Pedido.status == "em_transito")
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .group_by(Pedido.ingredient_id)
        .all()
    )
    return {ingredient_id: _as_float(qty) for ingredient_id, qty in rows}


def _latest_abt_criticality_by_ingredient(db: Session) -> dict[str, str]:
    try:
        with db.begin_nested():
            rows = db.execute(
                text(
                    """
                    SELECT ingredient_id, y_nivel_criticidade
                    FROM (
                        SELECT
                            ingredient_id,
                            y_nivel_criticidade,
                            ROW_NUMBER() OVER (
                                PARTITION BY ingredient_id
                                ORDER BY "date" DESC
                            ) AS row_number
                        FROM ml.abt_reposicao
                        WHERE y_nivel_criticidade IS NOT NULL
                    ) latest
                    WHERE row_number = 1
                    """
                )
            ).all()
    except Exception:
        logger.exception("Falha ao carregar criticidade de ml.abt_reposicao")
        return {}
    return {
        str(ingredient_id): str(criticality)
        for ingredient_id, criticality in rows
        if ingredient_id and criticality
    }


def _purchase_display_criticality(label: str) -> str:
    normalized = (
        normalize("NFKD", label or "")
        .encode("ascii", "ignore")
        .decode("ascii")
        .strip()
        .lower()
    )
    labels = {
        "emergencial": "Emergencial",
        "critico": "Crítico",
        "atencao": "Atenção",
        "alerta de compra": "Alerta de compra",
        "ok": "OK",
    }
    return labels.get(normalized, label or "OK")


def _resolve_purchase_criticality(
    item: Optional[CriticalityReportItem],
    abt_criticality: Optional[str],
    stock_position: float,
    forecast_qty: float,
) -> tuple[str, str]:
    if item is not None and item.criticidade_predita:
        return _purchase_display_criticality(item.criticidade_predita), "model_report"
    if abt_criticality:
        return _purchase_display_criticality(abt_criticality), "abt_reposicao"
    return (
        _purchase_display_criticality(
            _purchase_criticality_label(None, stock_position, forecast_qty)
        ),
        "operational_rule",
    )


def _purchase_criticality_label(item: Optional[CriticalityReportItem], stock_position: float, forecast_qty: float) -> str:
    if item is not None and item.criticidade_predita:
        return item.criticidade_predita
    if stock_position <= 0:
        return "Crítico"
    if forecast_qty > 0 and stock_position < forecast_qty:
        return "Alerta de compra"
    return "OK"


def _purchase_status_is_critical(label: str) -> bool:
    normalized = (
        normalize("NFKD", label or "")
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )
    return any(
        token in normalized
        for token in ("emergencial", "critico", "atencao", "alerta", "zerado", "ruptura")
    )


def _purchase_option_score(
    effective_unit_price: float,
    delivery_days: int,
    delay_risk: float,
    discount_percent: float,
) -> float:
    return round(
        effective_unit_price
        + (delivery_days * 0.12)
        + (delay_risk * 8.0)
        - (_as_float(discount_percent) * 0.015),
        4,
    )


def _purchase_option_reason(option: PurchasePlanSupplierOption) -> str:
    fragments = [
        f"R$ {_as_float(option.effective_unit_price):.2f}/un",
        f"{int(option.delivery_time_days or 0)} dias",
    ]
    if _as_float(option.discount_percent) > 0:
        fragments.append(f"{_as_float(option.discount_percent):.1f}% desc.")
    if _as_float(option.delay_risk) > 0:
        fragments.append(f"risco {(_as_float(option.delay_risk) * 100):.0f}%")
    return ", ".join(fragments)


def _sync_purchase_plan_quotes(db: Session, plan: PurchasePlan) -> None:
    selected_totals: dict[str, dict] = {}
    for item in plan.items:
        if not item.selected_supplier_id or _as_float(item.approved_qty) <= 0:
            continue
        option = _find_item_option(item, item.selected_supplier_id)
        if option is None:
            continue
        supplier = (
            db.query(Fornecedor.email)
            .filter(Fornecedor.id == option.supplier_id)
            .first()
        )
        payload = selected_totals.setdefault(
            option.supplier_id,
            {
                "supplier_name": option.supplier_name,
                "email": supplier.email if supplier else None,
                "total": 0.0,
            },
        )
        payload["total"] += _as_float(item.approved_qty) * _as_float(option.effective_unit_price)

    existing = {quote.supplier_id: quote for quote in plan.quotes}
    for supplier_id, payload in selected_totals.items():
        quote = existing.get(supplier_id)
        if quote is None:
            quote = SupplierQuote(
                supplier_id=supplier_id,
                supplier_name=payload["supplier_name"],
                email=payload.get("email"),
                channel="email",
                status="rascunho",
            )
            plan.quotes.append(quote)
        quote.supplier_name = payload["supplier_name"]
        quote.email = payload.get("email")
        quote.total_estimated = round(payload["total"], 2)

    for quote in list(plan.quotes):
        if quote.supplier_id not in selected_totals and quote.status == "rascunho":
            plan.quotes.remove(quote)


def _recalculate_purchase_plan(plan: PurchasePlan) -> None:
    items = list(plan.items)
    plan.total_estimated = round(sum(_as_float(item.estimated_total) for item in items), 2)
    plan.approved_total = round(
        sum(_as_float(item.approved_qty) * _as_float(item.estimated_unit_price) for item in items),
        2,
    )
    plan.critical_items_count = sum(1 for item in items if _purchase_status_is_critical(item.criticality))
    coverages = [
        min(_as_float(item.coverage_days), 90.0)
        for item in items
        if _as_float(item.avg_daily_usage) > 0
    ]
    plan.avg_coverage_days = round(sum(coverages) / len(coverages), 2) if coverages else 0
    best_total = 0.0
    selected_total = 0.0
    for item in items:
        qty = _as_float(item.approved_qty)
        if qty <= 0 or not item.options:
            continue
        best_total += min(_as_float(option.effective_unit_price) * qty for option in item.options)
        selected_total += _as_float(item.estimated_unit_price) * qty
    plan.savings_potential = round(max(0.0, selected_total - best_total), 2)
    plan.updated_at = _now_recife()


def _serialize_purchase_option(option: PurchasePlanSupplierOption) -> PurchasePlanSupplierOptionOut:
    return PurchasePlanSupplierOptionOut(
        id=option.id,
        supplier_id=option.supplier_id,
        supplier_name=option.supplier_name,
        unit_price=_as_float(option.unit_price),
        discount_percent=_as_float(option.discount_percent),
        min_to_discount=_as_float(option.min_to_discount),
        effective_unit_price=_as_float(option.effective_unit_price),
        delivery_time_days=int(option.delivery_time_days or 0),
        delay_risk=_as_float(option.delay_risk),
        score=_as_float(option.score),
        recommended=bool(option.recommended),
        reason=option.reason,
    )


def _serialize_purchase_item(item: PurchasePlanItem) -> PurchasePlanItemOut:
    options = sorted(item.options, key=lambda option: (option.score, option.supplier_name))
    return PurchasePlanItemOut(
        id=item.id,
        ingredient_id=item.ingredient_id,
        ingredient_name=item.ingredient_name,
        category=item.category,
        unit=item.unit,
        current_qty=_as_float(item.current_qty),
        avg_daily_usage=_as_float(item.avg_daily_usage),
        forecast_qty=_as_float(item.forecast_qty),
        in_transit_qty=_as_float(item.in_transit_qty),
        recommended_qty=_as_float(item.recommended_qty),
        approved_qty=_as_float(item.approved_qty),
        selected_supplier_id=item.selected_supplier_id,
        selected_supplier_name=item.selected_supplier_name,
        estimated_unit_price=_as_float(item.estimated_unit_price),
        estimated_total=_as_float(item.estimated_total),
        coverage_days=_as_float(item.coverage_days),
        criticality=item.criticality,
        criticality_source=item.criticality_source,
        justification=item.justification,
        note=item.note,
        options=[_serialize_purchase_option(option) for option in options],
    )


def _serialize_supplier_quote(quote: SupplierQuote) -> SupplierQuoteOut:
    return SupplierQuoteOut(
        id=quote.id,
        supplier_id=quote.supplier_id,
        supplier_name=quote.supplier_name,
        email=quote.email,
        channel=quote.channel,
        status=quote.status,
        sent_at=quote.sent_at,
        responded_at=quote.responded_at,
        approved_at=quote.approved_at,
        total_estimated=_as_float(quote.total_estimated),
        notes=quote.notes,
    )


def _serialize_purchase_plan(plan: PurchasePlan) -> PurchasePlanOut:
    items = sorted(
        plan.items,
        key=lambda item: (
            0 if _purchase_status_is_critical(item.criticality) else 1,
            -_as_float(item.recommended_qty),
            item.ingredient_name,
        ),
    )
    quotes = sorted(plan.quotes, key=lambda quote: quote.supplier_name)
    return PurchasePlanOut(
        id=plan.id,
        created_at=plan.created_at,
        updated_at=plan.updated_at,
        status=plan.status,
        source=plan.source,
        horizon_days=plan.horizon_days,
        date_from=plan.date_from,
        date_to=plan.date_to,
        contagem_id=plan.contagem_id,
        total_estimated=_as_float(plan.total_estimated),
        approved_total=_as_float(plan.approved_total),
        critical_items_count=int(plan.critical_items_count or 0),
        avg_coverage_days=_as_float(plan.avg_coverage_days),
        savings_potential=_as_float(plan.savings_potential),
        items=[_serialize_purchase_item(item) for item in items],
        quotes=[_serialize_supplier_quote(quote) for quote in quotes],
    )


def _build_purchase_plan(payload: PurchasePlanGenerateRequest, db: Session) -> PurchasePlan:
    if payload.contagem_id is not None:
        contagem = db.query(Contagem).filter(Contagem.id == payload.contagem_id).first()
        if contagem is None:
            raise HTTPException(status_code=404, detail="Contagem não encontrada")
        if contagem.status != "finalizada":
            raise HTTPException(status_code=400, detail="A contagem precisa estar finalizada")

    today = _now_recife().date()
    horizon_days = max(1, min(30, payload.horizon_days or 7))
    date_from = payload.date_from or today
    date_to = payload.date_to or (date_from + timedelta(days=horizon_days - 1))
    if date_from > date_to:
        raise HTTPException(status_code=400, detail="Data inicial maior que data final")

    usage_by_ingredient = _purchase_recent_usage_window(db, horizon_days, date_from)
    in_transit_by_ingredient = _purchase_in_transit_map(db)
    criticidade_run, criticidade_by_ingredient = _latest_criticidade_by_ingredient(db)
    abt_criticality_by_ingredient = _latest_abt_criticality_by_ingredient(db)
    source = (
        "contagem"
        if payload.contagem_id
        else (
            "model_report"
            if criticidade_run
            else ("abt_reposicao" if abt_criticality_by_ingredient else "operational_rule")
        )
    )

    rows = (
        db.query(Ingrediente, Categoria.name.label("category_name"), EstoqueAtual.qtd.label("current_qty"))
        .join(Categoria, Categoria.id == Ingrediente.category_id)
        .outerjoin(EstoqueAtual, EstoqueAtual.ingrediente == Ingrediente.id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .order_by(Categoria.name.asc(), Ingrediente.name.asc())
        .all()
    )

    plan = PurchasePlan(
        status="rascunho",
        source=source,
        horizon_days=horizon_days,
        date_from=date_from,
        date_to=date_to,
        contagem_id=payload.contagem_id,
    )
    db.add(plan)

    quote_totals: dict[str, dict] = {}
    for ingredient, category_name, current_qty_raw in rows:
        current_qty = _as_float(current_qty_raw)
        recent_usage_qty = max(0.0, _as_float(usage_by_ingredient.get(ingredient.id)))
        avg_daily_usage = round(recent_usage_qty / horizon_days, 4) if horizon_days > 0 else 0.0
        forecast_qty = round(recent_usage_qty, 4)
        in_transit_qty = _as_float(in_transit_by_ingredient.get(ingredient.id))
        stock_position = current_qty + in_transit_qty
        criticality_item = criticidade_by_ingredient.get(ingredient.id)
        criticality, criticality_source = _resolve_purchase_criticality(
            criticality_item,
            abt_criticality_by_ingredient.get(ingredient.id),
            stock_position,
            forecast_qty,
        )
        recommended_qty = max(0.0, forecast_qty)
        recommended_qty = round(recommended_qty, 2)
        if recommended_qty <= 0 and not _purchase_status_is_critical(criticality):
            continue

        coverage_before = (
            stock_position / avg_daily_usage
            if avg_daily_usage > 0
            else (90.0 if stock_position > 0 else 0.0)
        )
        supplier_rows = (
            db.query(
                FornecedorIngrediente,
                Fornecedor.name.label("supplier_name"),
                Fornecedor.email.label("supplier_email"),
                Fornecedor.avg_delivery_time,
            )
            .join(Fornecedor, Fornecedor.id == FornecedorIngrediente.supplier_id)
            .filter(FornecedorIngrediente.ingredient_id == ingredient.id)
            .order_by(Fornecedor.name.asc())
            .all()
        )
        option_payloads: list[dict] = []
        for supplier_option, supplier_name, supplier_email, avg_delivery_time in supplier_rows:
            delivery_days = int(avg_delivery_time or 0)
            effective_price, _ = _effective_unit_price(
                supplier_option.price,
                supplier_option.discount_percent,
                supplier_option.min_to_discount,
                recommended_qty,
            )
            delay_risk = 0.0
            if avg_daily_usage > 0 and coverage_before < delivery_days:
                delay_risk = min(1.0, (delivery_days - coverage_before) / max(1, delivery_days))
            score = _purchase_option_score(
                effective_price,
                delivery_days,
                delay_risk,
                supplier_option.discount_percent,
            )
            option_payloads.append(
                {
                    "supplier_id": supplier_option.supplier_id,
                    "supplier_name": supplier_name,
                    "email": supplier_email,
                    "unit_price": _as_float(supplier_option.price),
                    "discount_percent": _as_float(supplier_option.discount_percent),
                    "min_to_discount": _as_float(supplier_option.min_to_discount),
                    "effective_unit_price": effective_price,
                    "delivery_time_days": delivery_days,
                    "delay_risk": round(delay_risk, 4),
                    "score": score,
                }
            )
        option_payloads.sort(key=lambda option: (option["score"], option["supplier_name"]))
        selected = option_payloads[0] if option_payloads else None
        selected_unit_price = _as_float(selected["effective_unit_price"]) if selected else 0.0
        coverage_after = (
            (stock_position + recommended_qty) / avg_daily_usage
            if avg_daily_usage > 0
            else (90.0 if stock_position + recommended_qty > 0 else 0.0)
        )
        justification = (
            f"Sugestao baseada no consumo dos {horizon_days} dias anteriores: "
            f"{forecast_qty:.2f} {ingredient.unit}. "
            f"Esse total estima a necessidade dos proximos {horizon_days} dias; "
            f"estoque atual + transito em {stock_position:.2f} {ingredient.unit}."
        )
        item = PurchasePlanItem(
            ingredient_id=ingredient.id,
            ingredient_name=ingredient.name,
            category=category_name,
            unit=ingredient.unit,
            current_qty=round(current_qty, 4),
            avg_daily_usage=round(avg_daily_usage, 4),
            forecast_qty=forecast_qty,
            in_transit_qty=round(in_transit_qty, 4),
            recommended_qty=recommended_qty,
            approved_qty=recommended_qty,
            selected_supplier_id=selected["supplier_id"] if selected else None,
            selected_supplier_name=selected["supplier_name"] if selected else None,
            estimated_unit_price=selected_unit_price,
            estimated_total=round(selected_unit_price * recommended_qty, 2),
            coverage_days=round(min(coverage_after, 999.0), 2),
            criticality=criticality,
            criticality_source=criticality_source,
            justification=justification,
        )
        plan.items.append(item)
        for index, option in enumerate(option_payloads):
            supplier_option = PurchasePlanSupplierOption(
                supplier_id=option["supplier_id"],
                supplier_name=option["supplier_name"],
                unit_price=option["unit_price"],
                discount_percent=option["discount_percent"],
                min_to_discount=option["min_to_discount"],
                effective_unit_price=option["effective_unit_price"],
                delivery_time_days=option["delivery_time_days"],
                delay_risk=option["delay_risk"],
                score=option["score"],
                recommended=1 if index == 0 else 0,
            )
            supplier_option.reason = _purchase_option_reason(supplier_option)
            item.options.append(supplier_option)
        if selected:
            quote = quote_totals.setdefault(
                selected["supplier_id"],
                {
                    "supplier_name": selected["supplier_name"],
                    "email": selected.get("email"),
                    "total": 0.0,
                },
            )
            quote["total"] += selected_unit_price * recommended_qty

    for supplier_id, quote in quote_totals.items():
        plan.quotes.append(
            SupplierQuote(
                supplier_id=supplier_id,
                supplier_name=quote["supplier_name"],
                email=quote.get("email"),
                status="rascunho",
                channel="email",
                total_estimated=round(quote["total"], 2),
            )
        )
    _sync_purchase_plan_quotes(db, plan)
    _recalculate_purchase_plan(plan)
    db.commit()
    db.refresh(plan)
    return plan


def _get_purchase_plan_or_404(db: Session, plan_id: int) -> PurchasePlan:
    plan = db.query(PurchasePlan).filter(PurchasePlan.id == plan_id).first()
    if plan is None:
        raise HTTPException(status_code=404, detail="Plano de compra não encontrado")
    return plan


def _find_item_option(item: PurchasePlanItem, supplier_id: Optional[str]) -> Optional[PurchasePlanSupplierOption]:
    if not supplier_id:
        return None
    return next((option for option in item.options if option.supplier_id == supplier_id), None)


def _purchase_simulation(plan: PurchasePlan, payload: PurchasePlanSimulationRequest) -> PurchasePlanSimulationOut:
    overrides = {item.ingredient_id: item for item in payload.items}
    approved_total = 0.0
    best_total = 0.0
    coverage_values: list[float] = []
    rupture_risk_items = 0
    critical_items_count = 0
    notes: list[str] = []

    for item in plan.items:
        override = overrides.get(item.ingredient_id)
        approved_qty = _as_float(override.approved_qty) if override else _as_float(item.approved_qty)
        selected_option = (
            _find_item_option(item, override.selected_supplier_id)
            if override and override.selected_supplier_id
            else _find_item_option(item, item.selected_supplier_id)
        )
        unit_price = _as_float(selected_option.effective_unit_price) if selected_option else _as_float(item.estimated_unit_price)
        approved_total += approved_qty * unit_price
        if item.options and approved_qty > 0:
            best_total += min(_as_float(option.effective_unit_price) * approved_qty for option in item.options)
        avg_usage = _as_float(item.avg_daily_usage)
        stock_after = _as_float(item.current_qty) + _as_float(item.in_transit_qty) + approved_qty
        coverage = stock_after / avg_usage if avg_usage > 0 else (90.0 if stock_after > 0 else 0.0)
        if avg_usage > 0:
            coverage_values.append(min(coverage, 90.0))
        lead_time = int(selected_option.delivery_time_days or 0) if selected_option else 0
        if avg_usage > 0 and coverage < lead_time:
            rupture_risk_items += 1
            notes.append(f"{item.ingredient_name}: cobertura ({coverage:.1f} dias) menor que prazo ({lead_time} dias).")
        if _purchase_status_is_critical(item.criticality) or coverage < plan.horizon_days:
            critical_items_count += 1

    projected_coverage = round(sum(coverage_values) / len(coverage_values), 2) if coverage_values else 0
    return PurchasePlanSimulationOut(
        total_estimated=_as_float(plan.total_estimated),
        approved_total=round(approved_total, 2),
        projected_coverage_days=projected_coverage,
        rupture_risk_items=rupture_risk_items,
        critical_items_count=critical_items_count,
        savings_potential=round(max(0.0, approved_total - best_total), 2),
        notes=notes[:8],
    )


@app.post("/api/compras/planos/gerar", response_model=PurchasePlanOut)
def generate_purchase_plan(payload: PurchasePlanGenerateRequest, db: Session = Depends(get_db)):
    return _serialize_purchase_plan(_build_purchase_plan(payload, db))


@app.get("/api/compras/planos/latest", response_model=Optional[PurchasePlanOut])
def get_latest_purchase_plan(db: Session = Depends(get_db)):
    plan = db.query(PurchasePlan).order_by(PurchasePlan.created_at.desc(), PurchasePlan.id.desc()).first()
    return _serialize_purchase_plan(plan) if plan else None


@app.get("/api/compras/planos/{plan_id}", response_model=PurchasePlanOut)
def get_purchase_plan(plan_id: int, db: Session = Depends(get_db)):
    return _serialize_purchase_plan(_get_purchase_plan_or_404(db, plan_id))


@app.patch("/api/compras/planos/{plan_id}/items/{ingredient_id}", response_model=PurchasePlanOut)
def update_purchase_plan_item(
    plan_id: int,
    ingredient_id: str,
    payload: PurchasePlanItemUpdateRequest,
    db: Session = Depends(get_db),
):
    plan = _get_purchase_plan_or_404(db, plan_id)
    if plan.status == "aprovado":
        raise HTTPException(status_code=409, detail="Plano aprovado não pode ser alterado")
    item = next((plan_item for plan_item in plan.items if plan_item.ingredient_id == ingredient_id), None)
    if item is None:
        raise HTTPException(status_code=404, detail="Item do plano não encontrado")
    if payload.approved_qty is not None:
        item.approved_qty = round(payload.approved_qty, 4)
    if payload.selected_supplier_id is not None:
        option = _find_item_option(item, payload.selected_supplier_id)
        if option is None:
            raise HTTPException(status_code=400, detail="Fornecedor inválido para o item")
        for item_option in item.options:
            item_option.recommended = 1 if item_option.supplier_id == option.supplier_id else 0
        item.selected_supplier_id = option.supplier_id
        item.selected_supplier_name = option.supplier_name
        item.estimated_unit_price = option.effective_unit_price
    if payload.note is not None:
        item.note = payload.note
    item.estimated_total = round(_as_float(item.approved_qty) * _as_float(item.estimated_unit_price), 2)
    avg_usage = _as_float(item.avg_daily_usage)
    stock_after = _as_float(item.current_qty) + _as_float(item.in_transit_qty) + _as_float(item.approved_qty)
    item.coverage_days = round(min(stock_after / avg_usage if avg_usage > 0 else 90.0, 999.0), 2)
    plan.status = "em_revisao" if plan.status == "rascunho" else plan.status
    _sync_purchase_plan_quotes(db, plan)
    _recalculate_purchase_plan(plan)
    db.commit()
    db.refresh(plan)
    return _serialize_purchase_plan(plan)


@app.delete("/api/compras/planos/{plan_id}/items/{ingredient_id}", response_model=PurchasePlanOut)
def delete_purchase_plan_item(
    plan_id: int,
    ingredient_id: str,
    db: Session = Depends(get_db),
):
    plan = _get_purchase_plan_or_404(db, plan_id)
    if plan.status == "aprovado":
        raise HTTPException(status_code=409, detail="Plano aprovado não pode ser alterado")
    item = next((plan_item for plan_item in plan.items if plan_item.ingredient_id == ingredient_id), None)
    if item is None:
        raise HTTPException(status_code=404, detail="Item do plano não encontrado")

    plan.items.remove(item)
    plan.status = "em_revisao" if plan.status == "rascunho" else plan.status
    _sync_purchase_plan_quotes(db, plan)
    _recalculate_purchase_plan(plan)
    db.commit()
    db.refresh(plan)
    return _serialize_purchase_plan(plan)


@app.post("/api/compras/planos/{plan_id}/simular", response_model=PurchasePlanSimulationOut)
def simulate_purchase_plan(
    plan_id: int,
    payload: PurchasePlanSimulationRequest,
    db: Session = Depends(get_db),
):
    return _purchase_simulation(_get_purchase_plan_or_404(db, plan_id), payload)


def _purchase_plan_email_groups(plan: PurchasePlan) -> dict[str, dict]:
    groups: dict[str, dict] = {}
    for item in plan.items:
        if _as_float(item.approved_qty) <= 0 or not item.selected_supplier_id:
            continue
        option = _find_item_option(item, item.selected_supplier_id)
        if option is None:
            continue
        quote = next((quote for quote in plan.quotes if quote.supplier_id == option.supplier_id), None)
        expected_date = _now_recife().date() + timedelta(days=int(option.delivery_time_days or 0))
        group = groups.setdefault(
            option.supplier_id,
            {
                "supplier_name": option.supplier_name,
                "email": quote.email if quote else None,
                "expected_date": expected_date,
                "items": [],
            },
        )
        group["expected_date"] = max(group["expected_date"], expected_date)
        group["items"].append(
            OrderEmailItem(
                name=f"Plano #{plan.id} - {item.ingredient_name}",
                qty=_as_float(item.approved_qty),
                unit=item.unit or "",
                unit_price=_as_float(option.effective_unit_price),
                total_value=round(_as_float(option.effective_unit_price) * _as_float(item.approved_qty), 2),
            )
        )
    return groups


@app.post("/api/compras/planos/{plan_id}/cotacoes/enviar", response_model=PurchasePlanOut)
def send_purchase_plan_quotes(plan_id: int, db: Session = Depends(get_db)):
    plan = _get_purchase_plan_or_404(db, plan_id)
    _sync_purchase_plan_quotes(db, plan)
    email_results = _send_pedido_emails(_purchase_plan_email_groups(plan), _now_recife().date())
    by_supplier = {result.supplier_id: result for result in email_results}
    sent_any = False
    for quote in plan.quotes:
        result = by_supplier.get(quote.supplier_id)
        if result is None:
            continue
        quote.status = "enviada" if result.status == "sent" else result.status
        quote.notes = result.message
        if result.status == "sent":
            quote.sent_at = _now_recife()
            sent_any = True
    if sent_any:
        plan.status = "cotado"
    _recalculate_purchase_plan(plan)
    db.commit()
    db.refresh(plan)
    return _serialize_purchase_plan(plan)


@app.post("/api/compras/planos/{plan_id}/aprovar", response_model=PedidoCreateResponse)
def approve_purchase_plan(plan_id: int, db: Session = Depends(get_db)):
    plan = _get_purchase_plan_or_404(db, plan_id)
    _sync_purchase_plan_quotes(db, plan)
    items = [
        PedidoCreateItem(supplier_id=item.selected_supplier_id, ingredient_id=item.ingredient_id, qty=_as_float(item.approved_qty))
        for item in plan.items
        if item.selected_supplier_id and _as_float(item.approved_qty) > 0
    ]
    if not items:
        raise HTTPException(status_code=400, detail="Plano sem itens aprovados")
    response = create_pedidos(PedidoCreateRequest(items=items), db)
    plan.status = "aprovado"
    now = _now_recife()
    selected_suppliers = {item.supplier_id for item in items}
    for quote in plan.quotes:
        if quote.supplier_id in selected_suppliers:
            quote.status = "aprovada"
            quote.approved_at = now
    _recalculate_purchase_plan(plan)
    db.commit()
    return response


@app.get("/api/export/compras/planos/{plan_id}")
def export_purchase_plan(
    plan_id: int,
    format: str = Query(default="excel"),
    theme: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    plan = _get_purchase_plan_or_404(db, plan_id)
    rows = [
        {
            "plano_id": plan.id,
            "status": plan.status,
            "origem": plan.source,
            "ingrediente_id": item.ingredient_id,
            "ingrediente": item.ingredient_name,
            "categoria": item.category,
            "unidade": item.unit,
            "estoque_atual": _as_float(item.current_qty),
            "consumo_dia": _as_float(item.avg_daily_usage),
            "previsao_consumo": _as_float(item.forecast_qty),
            "em_transito": _as_float(item.in_transit_qty),
            "qtd_recomendada": _as_float(item.recommended_qty),
            "qtd_aprovada": _as_float(item.approved_qty),
            "fornecedor": item.selected_supplier_name,
            "preco_unitario": _as_float(item.estimated_unit_price),
            "total": _as_float(item.estimated_total),
            "cobertura_dias": _as_float(item.coverage_days),
            "criticidade": item.criticality,
            "origem_criticidade": item.criticality_source,
            "justificativa": item.justification,
        }
        for item in plan.items
    ]
    return _export_response(
        rows,
        f"plano_compra_{plan.id}",
        format,
        f"Plano de compra Maestro #{plan.id}",
        [
            "plano_id",
            "status",
            "origem",
            "ingrediente_id",
            "ingrediente",
            "categoria",
            "unidade",
            "estoque_atual",
            "consumo_dia",
            "previsao_consumo",
            "em_transito",
            "qtd_recomendada",
            "qtd_aprovada",
            "fornecedor",
            "preco_unitario",
            "total",
            "cobertura_dias",
            "criticidade",
            "justificativa",
        ],
        theme_id=theme,
    )


@app.post("/api/pedidos/recomendacao", response_model=PedidoRecommendationResponse)
def recommend_pedidos(
    payload: PedidoRecommendationRequest,
    db: Session = Depends(get_db),
):
    today = date.today()
    response_items: list[PedidoRecommendationItem] = []
    grouped: dict[str, dict] = {}

    for item in payload.items:
        ingredient = (
            db.query(Ingrediente)
            .filter(Ingrediente.id == item.ingredient_id)
            .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
            .first()
        )
        if ingredient is None:
            raise HTTPException(
                status_code=400,
                detail=f"Ingrediente inválido: {item.ingredient_id}",
            )

        option_rows = (
            db.query(
                FornecedorIngrediente.supplier_id,
                Fornecedor.name.label("supplier_name"),
                Fornecedor.avg_delivery_time,
                FornecedorIngrediente.price,
                FornecedorIngrediente.discount_percent,
                FornecedorIngrediente.min_to_discount,
            )
            .join(Fornecedor, Fornecedor.id == FornecedorIngrediente.supplier_id)
            .filter(FornecedorIngrediente.ingredient_id == item.ingredient_id)
            .order_by(Fornecedor.name.asc())
            .all()
        )

        raw_options = []
        for row in option_rows:
            delivery_days = int(row.avg_delivery_time or 0)
            effective_unit_price, discount_applied = _effective_unit_price(
                row.price,
                row.discount_percent,
                row.min_to_discount,
                item.qty,
            )
            raw_options.append(
                {
                    "supplier_id": row.supplier_id,
                    "supplier_name": row.supplier_name,
                    "unit_price": _as_float(row.price),
                    "discount_percent": _as_float(row.discount_percent),
                    "min_to_discount": _as_float(row.min_to_discount),
                    "discount_applied": discount_applied,
                    "effective_unit_price": effective_unit_price,
                    "total_value": round(effective_unit_price * item.qty, 2),
                    "delivery_time_days": delivery_days,
                    "expected_date": today + timedelta(days=delivery_days),
                }
            )

        recommended = None
        if raw_options:
            recommended = min(
                raw_options,
                key=lambda option: (
                    option["total_value"],
                    option["delivery_time_days"],
                    -option["discount_percent"],
                    option["supplier_name"],
                ),
            )

        options: list[SupplierOption] = []
        for option in raw_options:
            detractors: list[str] = []
            if recommended and option["supplier_id"] != recommended["supplier_id"]:
                if option["total_value"] > recommended["total_value"] + 0.005:
                    detractors.append("mais caro")
                if option["delivery_time_days"] > recommended["delivery_time_days"]:
                    detractors.append("entrega mais lenta")
                if option["discount_percent"] < recommended["discount_percent"]:
                    detractors.append("desconto menor")
                if recommended["discount_applied"] and not option["discount_applied"]:
                    detractors.append("não aplica desconto")
            options.append(
                SupplierOption(
                    **option,
                    detractors=detractors,
                    recommended=bool(
                        recommended
                        and option["supplier_id"] == recommended["supplier_id"]
                    ),
                )
            )

        options.sort(
            key=lambda option: (
                option.total_value,
                option.delivery_time_days,
                -option.discount_percent,
                option.supplier_name,
            )
        )

        response_items.append(
            PedidoRecommendationItem(
                ingredient_id=ingredient.id,
                ingredient_name=ingredient.name,
                category=ingredient.category,
                unit=ingredient.unit,
                qty=item.qty,
                recommended_supplier_id=(
                    recommended["supplier_id"] if recommended else None
                ),
                options=options,
            )
        )

        if recommended:
            group = grouped.setdefault(
                recommended["supplier_id"],
                {
                    "supplier_id": recommended["supplier_id"],
                    "supplier_name": recommended["supplier_name"],
                    "expected_date": recommended["expected_date"],
                    "total_value": 0.0,
                    "items": [],
                },
            )
            group["expected_date"] = max(
                group["expected_date"],
                recommended["expected_date"],
            )
            group["total_value"] += recommended["total_value"]
            group["items"].append(
                RecommendedOrderItem(
                    ingredient_id=ingredient.id,
                    ingredient_name=ingredient.name,
                    qty=item.qty,
                    unit=ingredient.unit,
                    total_value=recommended["total_value"],
                    expected_date=recommended["expected_date"],
                )
            )

    groups = [
        RecommendedOrderGroup(
            supplier_id=group["supplier_id"],
            supplier_name=group["supplier_name"],
            expected_date=group["expected_date"],
            total_value=round(group["total_value"], 2),
            items=group["items"],
        )
        for group in grouped.values()
    ]
    groups.sort(key=lambda group: (group.expected_date, group.supplier_name))

    return PedidoRecommendationResponse(items=response_items, groups=groups)


@app.post("/api/pedidos", response_model=PedidoCreateResponse)
def create_pedidos(payload: PedidoCreateRequest, db: Session = Depends(get_db)):
    today = date.today()
    aggregated: dict[tuple[str, str], float] = {}
    for item in payload.items:
        key = (item.supplier_id, item.ingredient_id)
        aggregated[key] = aggregated.get(key, 0.0) + item.qty

    created = 0
    updated = 0
    touched_groups: set[tuple[str, date]] = set()
    reserved_pedido_ids: set[str] = set()
    email_groups: dict[str, dict] = {}

    for (supplier_id, ingredient_id), qty in aggregated.items():
        option = (
            db.query(
                FornecedorIngrediente,
                Fornecedor.avg_delivery_time.label("avg_delivery_time"),
                Fornecedor.name.label("supplier_name"),
                Fornecedor.email.label("supplier_email"),
                Ingrediente.name.label("ingredient_name"),
                Ingrediente.unit.label("ingredient_unit"),
            )
            .join(Fornecedor, Fornecedor.id == FornecedorIngrediente.supplier_id)
            .join(Ingrediente, Ingrediente.id == FornecedorIngrediente.ingredient_id)
            .filter(FornecedorIngrediente.supplier_id == supplier_id)
            .filter(FornecedorIngrediente.ingredient_id == ingredient_id)
            .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
            .first()
        )
        if option is None:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Fornecedor inválido para ingrediente: "
                    f"{supplier_id}/{ingredient_id}"
                ),
            )

        (
            fornecedor_ingrediente,
            avg_delivery_time,
            supplier_name,
            supplier_email,
            ingredient_name,
            ingredient_unit,
        ) = option
        delivery_days = int(avg_delivery_time or 0)
        expected_date = today + timedelta(days=delivery_days)
        email_unit_price, _ = _effective_unit_price(
            fornecedor_ingrediente.price,
            fornecedor_ingrediente.discount_percent,
            fornecedor_ingrediente.min_to_discount,
            qty,
        )
        email_group = email_groups.setdefault(
            supplier_id,
            {
                "supplier_name": supplier_name,
                "email": supplier_email,
                "expected_date": expected_date,
                "items": [],
            },
        )
        email_group["expected_date"] = max(email_group["expected_date"], expected_date)
        email_group["items"].append(
            OrderEmailItem(
                name=ingredient_name,
                qty=qty,
                unit=ingredient_unit,
                unit_price=email_unit_price,
                total_value=round(email_unit_price * qty, 2),
            )
        )

        existing = (
            db.query(Pedido)
            .filter(Pedido.supplier_id == supplier_id)
            .filter(Pedido.ingredient_id == ingredient_id)
            .filter(Pedido.data_pedido == today)
            .filter(Pedido.status == "em_transito")
            .first()
        )

        if existing:
            new_qty = _as_float(existing.qty) + qty
            existing.qty = new_qty
            existing.valor = _option_total(
                fornecedor_ingrediente.price,
                fornecedor_ingrediente.discount_percent,
                fornecedor_ingrediente.min_to_discount,
                new_qty,
            )
            existing.data_prevista = expected_date
            existing.status = "em_transito"
            updated += 1
        else:
            db.add(
                Pedido(
                    id=_next_pedido_id(db, reserved_pedido_ids),
                    supplier_id=supplier_id,
                    ingredient_id=ingredient_id,
                    qty=qty,
                    valor=_option_total(
                        fornecedor_ingrediente.price,
                        fornecedor_ingrediente.discount_percent,
                        fornecedor_ingrediente.min_to_discount,
                        qty,
                    ),
                    data_pedido=today,
                    status="em_transito",
                    data_prevista=expected_date,
                )
            )
            created += 1

        touched_groups.add((supplier_id, today))

    db.commit()
    email_results = _send_pedido_emails(email_groups, today)

    groups = [
        group
        for supplier_id, order_date in sorted(touched_groups)
        if (group := _get_pedido_group(db, supplier_id, order_date)) is not None
    ]

    return PedidoCreateResponse(
        groups=groups,
        created=created,
        updated=updated,
        email_results=email_results,
    )


@app.get("/api/pedidos", response_model=PedidoPaginado)
def get_pedidos(
    status: Optional[str] = None,
    supplier_id: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=100),
    db: Session = Depends(get_db),
):
    query = _apply_pedido_filters(
        _pedido_group_query(db),
        status,
        supplier_id,
        date_from,
        date_to,
    )
    query = query.filter(Pedido.status != "em_transito")
    total = query.count()
    total_pages = max(1, math.ceil(total / page_size))
    rows = (
        query.order_by(Pedido.data_pedido.desc(), Fornecedor.name.asc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return PedidoPaginado(
        items=[_serialize_pedido_group(row) for row in rows],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
    )


@app.get("/api/export/pedidos")
def export_pedidos(
    format: str = Query(default="csv"),
    theme: Optional[str] = Query(default=None),
    date_from: date = Query(...),
    date_to: date = Query(...),
    db: Session = Depends(get_db),
):
    if date_from > date_to:
        raise HTTPException(status_code=400, detail="Data inicial maior que data final")

    rows = (
        db.query(
            Pedido.id,
            Pedido.data_pedido,
            Pedido.supplier_id,
            Fornecedor.name.label("supplier_name"),
            Pedido.ingredient_id,
            Ingrediente.name.label("ingredient_name"),
            Categoria.name.label("category"),
            Ingrediente.unit,
            Pedido.qty,
            Pedido.valor,
            Pedido.status,
            Pedido.data_prevista,
        )
        .join(Fornecedor, Fornecedor.id == Pedido.supplier_id)
        .join(Ingrediente, Ingrediente.id == Pedido.ingredient_id)
        .join(Categoria, Categoria.id == Ingrediente.category_id)
        .filter(Pedido.data_pedido >= date_from)
        .filter(Pedido.data_pedido <= date_to)
        .order_by(Pedido.data_pedido.asc(), Fornecedor.name.asc(), Ingrediente.name.asc())
        .all()
    )
    data = [
        {
            "pedido_id": row.id,
            "data_pedido": row.data_pedido,
            "fornecedor_id": row.supplier_id,
            "fornecedor": row.supplier_name,
            "ingrediente_id": row.ingredient_id,
            "ingrediente": row.ingredient_name,
            "categoria": row.category,
            "unidade": row.unit,
            "quantidade": _as_float(row.qty),
            "valor_total": _as_float(row.valor),
            "status": row.status,
            "data_prevista": row.data_prevista,
        }
        for row in rows
    ]
    return _export_response(
        data,
        f"pedidos_{date_from.isoformat()}_{date_to.isoformat()}",
        format,
        "Historico de pedidos",
        [
            "pedido_id",
            "data_pedido",
            "fornecedor_id",
            "fornecedor",
            "ingrediente_id",
            "ingrediente",
            "categoria",
            "unidade",
            "quantidade",
            "valor_total",
            "status",
            "data_prevista",
        ],
        theme_id=theme,
    )


@app.get("/api/pedidos/em-transito", response_model=list[PedidoGroupOut])
def get_pedidos_em_transito(
    supplier_id: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
):
    rows = (
        _apply_pedido_filters(
            _pedido_group_query(db),
            "em_transito",
            supplier_id,
            date_from,
            date_to,
        )
        .order_by(func.max(Pedido.data_prevista).asc(), Pedido.data_pedido.desc())
        .all()
    )
    return [_serialize_pedido_group(row) for row in rows]


@app.get(
    "/api/pedidos/grupos/{supplier_id}/{order_date}",
    response_model=PedidoDetailResponse,
)
def get_pedido_group_detail(
    supplier_id: str,
    order_date: date,
    db: Session = Depends(get_db),
):
    rows = (
        db.query(
            Pedido.id,
            Pedido.supplier_id,
            Fornecedor.name.label("supplier_name"),
            Pedido.data_pedido,
            Pedido.data_prevista,
            Pedido.status,
            Pedido.estoque_aplicado_em,
            Pedido.ingredient_id,
            Ingrediente.name.label("ingredient_name"),
            Categoria.name.label("category"),
            Ingrediente.unit,
            Pedido.qty,
            Pedido.valor,
        )
        .join(Fornecedor, Fornecedor.id == Pedido.supplier_id)
        .join(Ingrediente, Ingrediente.id == Pedido.ingredient_id)
        .join(Categoria, Categoria.id == Ingrediente.category_id)
        .filter(Pedido.supplier_id == supplier_id)
        .filter(Pedido.data_pedido == order_date)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .order_by(Ingrediente.name.asc())
        .all()
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")

    items = [
        PedidoDetailItem(
            ingredient_id=row.ingredient_id,
            ingredient_name=row.ingredient_name,
            category=row.category,
            unit=row.unit,
            qty=_as_float(row.qty),
            unit_price=(
                _as_float(row.valor) / _as_float(row.qty)
                if _as_float(row.qty) > 0
                else 0
            ),
            total_value=_as_float(row.valor),
        )
        for row in rows
    ]
    first = rows[0]
    expected_date = max(row.data_prevista for row in rows)
    stock_applied_at = (
        max(row.estoque_aplicado_em for row in rows)
        if all(row.estoque_aplicado_em is not None for row in rows)
        else None
    )
    status = (
        "em_transito"
        if any(row.status == "em_transito" for row in rows)
        else "entregue"
    )
    group_key = _pedido_group_key(supplier_id, order_date)

    return PedidoDetailResponse(
        id=group_key,
        group_key=group_key,
        supplier_id=first.supplier_id,
        supplier_name=first.supplier_name,
        order_date=first.data_pedido,
        expected_date=expected_date,
        status=status,
        stock_applied_at=stock_applied_at,
        items_qty=sum(item.qty for item in items),
        total_value=sum(item.total_value for item in items),
        items=items,
    )


@app.patch(
    "/api/pedidos/grupos/{supplier_id}/{order_date}/entregar",
    response_model=PedidoDetailResponse,
)
def mark_pedido_group_delivered(
    supplier_id: str,
    order_date: date,
    db: Session = Depends(get_db),
):
    pedidos = (
        db.query(Pedido)
        .join(Ingrediente, Ingrediente.id == Pedido.ingredient_id)
        .filter(Pedido.supplier_id == supplier_id)
        .filter(Pedido.data_pedido == order_date)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .all()
    )
    if not pedidos:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")

    _mark_pedidos_delivered_and_apply_stock(db, pedidos)

    db.commit()
    return get_pedido_group_detail(supplier_id, order_date, db)


@app.patch("/api/pedidos/{pedido_id}/entregar", response_model=PedidoDetailResponse)
def mark_pedido_delivered(
    pedido_id: str,
    db: Session = Depends(get_db),
):
    pedidos = (
        db.query(Pedido)
        .join(Ingrediente, Ingrediente.id == Pedido.ingredient_id)
        .filter(Pedido.id == pedido_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .all()
    )
    if not pedidos:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")

    _mark_pedidos_delivered_and_apply_stock(db, pedidos)

    db.commit()
    return get_pedido_detail(pedido_id, db)


@app.get("/api/pedidos/{pedido_id}", response_model=PedidoDetailResponse)
def get_pedido_detail(pedido_id: str, db: Session = Depends(get_db)):
    rows = (
        db.query(
            Pedido.id,
            Pedido.supplier_id,
            Fornecedor.name.label("supplier_name"),
            Pedido.data_pedido,
            Pedido.data_prevista,
            Pedido.status,
            Pedido.estoque_aplicado_em,
            Pedido.ingredient_id,
            Ingrediente.name.label("ingredient_name"),
            Categoria.name.label("category"),
            Ingrediente.unit,
            Pedido.qty,
            Pedido.valor,
        )
        .join(Fornecedor, Fornecedor.id == Pedido.supplier_id)
        .join(Ingrediente, Ingrediente.id == Pedido.ingredient_id)
        .join(Categoria, Categoria.id == Ingrediente.category_id)
        .filter(Pedido.id == pedido_id)
        .filter(Ingrediente.category_id != PRODUCTION_CATEGORY_ID)
        .order_by(Ingrediente.name.asc())
        .all()
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")

    items = [
        PedidoDetailItem(
            ingredient_id=row.ingredient_id,
            ingredient_name=row.ingredient_name,
            category=row.category,
            unit=row.unit,
            qty=_as_float(row.qty),
            unit_price=(
                _as_float(row.valor) / _as_float(row.qty)
                if _as_float(row.qty) > 0
                else 0
            ),
            total_value=_as_float(row.valor),
        )
        for row in rows
    ]
    first = rows[0]

    return PedidoDetailResponse(
        id=first.id,
        group_key=_pedido_group_key(first.supplier_id, first.data_pedido),
        supplier_id=first.supplier_id,
        supplier_name=first.supplier_name,
        order_date=first.data_pedido,
        expected_date=first.data_prevista,
        status=first.status,
        stock_applied_at=first.estoque_aplicado_em,
        items_qty=sum(item.qty for item in items),
        total_value=sum(item.total_value for item in items),
        items=items,
    )


# ---------------------------------------------------------------------------
# Log de contagem
# ---------------------------------------------------------------------------

def _serializa_log(e: LogContagem) -> LogContagemOut:
    return LogContagemOut(
        id=e.id,
        ingrediente_id=e.ingrediente_id,
        ingrediente_nome=e.ingrediente.name,
        unit=e.ingrediente.unit,
        quantidade_anterior=float(e.quantidade_anterior),
        quantidade_nova=float(e.quantidade_nova),
        delta=float(e.delta),
        sessao=e.sessao,
        criado_em=e.criado_em,
    )


@app.get("/api/log", response_model=list[LogContagemOut])
def get_log(
    limit: int = Query(default=200, le=1000),
    db: Session = Depends(get_db),
):
    entries = (
        db.query(LogContagem)
        .join(Ingrediente)
        .order_by(LogContagem.criado_em.desc())
        .limit(limit)
        .all()
    )
    return [_serializa_log(e) for e in entries]


@app.get("/api/log/{ingrediente_id}", response_model=list[LogContagemOut])
def get_log_ingrediente(ingrediente_id: str, db: Session = Depends(get_db)):
    if not db.query(Ingrediente).filter(Ingrediente.id == ingrediente_id).first():
        raise HTTPException(status_code=404, detail="Ingrediente não encontrado")

    entries = (
        db.query(LogContagem)
        .filter(LogContagem.ingrediente_id == ingrediente_id)
        .join(Ingrediente)
        .order_by(LogContagem.criado_em.desc())
        .all()
    )
    return [_serializa_log(e) for e in entries]


# ---------------------------------------------------------------------------
# Agente
# ---------------------------------------------------------------------------

def _normalize_agent_question(value: str) -> str:
    ascii_text = normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_text.lower()).strip()


def _is_purchase_needed_question(message: str) -> bool:
    normalized = _normalize_agent_question(message)
    has_item = any(term in normalized for term in ("item", "itens", "ingrediente", "insumo"))
    has_purchase = any(term in normalized for term in ("compra", "comprar", "repor", "reposicao"))
    has_need = any(term in normalized for term in ("precis", "necess", "alerta", "critico"))
    return has_item and has_purchase and has_need


def _purchase_needed_fallback_rows() -> tuple[list[dict], int]:
    with engine.connect() as conn:
        total = int(conn.execute(text(PURCHASE_NEEDED_FALLBACK_COUNT_SQL)).scalar() or 0)
        result = conn.execute(
            text(PURCHASE_NEEDED_FALLBACK_SQL),
            {"limit": AGENT_ROWS_PREVIEW_LIMIT},
        )
        return [dict(row._mapping) for row in result], total


def _agent_answer(ctx: dict, row_count: int) -> str:
    if ctx.get("answer"):
        return str(ctx["answer"])

    if not ctx.get("is_valid"):
        return (
            ctx.get("error_message")
            or "Nao consegui responder essa pergunta com os dados disponiveis."
        )

    if row_count == 0:
        return "Nao encontrei registros para essa pergunta."
    if row_count == 1:
        return "Encontrei 1 registro relacionado a sua pergunta."
    return f"Encontrei {row_count} registros relacionados a sua pergunta."


@app.post("/api/agent/chat", response_model=AgentChatResponse)
def chat_with_agent(payload: AgentChatRequest, response: Response):
    response.headers["Cache-Control"] = "no-store"
    message = payload.message.strip()
    if not message:
        raise HTTPException(status_code=400, detail="Mensagem vazia")

    session_id = payload.session_id or f"agent-{uuid4().hex}"

    try:
        ctx = perguntar(message, session_id=session_id)
    except Exception as exc:
        ctx = {
            "is_valid": False,
            "error_type": "erro_execucao",
            "error_message": f"Erro ao chamar o agente: {exc}",
            "rows": [],
        }

    rows = ctx.get("rows") or []
    fallback_row_count: Optional[int] = None
    if not rows and _is_purchase_needed_question(message):
        fallback_rows, fallback_total = _purchase_needed_fallback_rows()
        if fallback_rows:
            ctx = {
                **ctx,
                "is_valid": True,
                "error_type": None,
                "error_message": None,
                "answer": (
                    "Encontrei itens com indicacao de compra na base analitica "
                    "mais recente de reposicao."
                ),
                "rows": fallback_rows,
            }
            rows = fallback_rows
            fallback_row_count = fallback_total

    preview_rows = rows[:AGENT_ROWS_PREVIEW_LIMIT]
    columns = list(preview_rows[0].keys()) if preview_rows else []
    row_count = fallback_row_count if fallback_row_count is not None else len(rows)

    return AgentChatResponse(
        session_id=session_id,
        answer=_agent_answer(ctx, row_count),
        rows=preview_rows,
        columns=columns,
        row_count=row_count,
        is_valid=bool(ctx.get("is_valid")),
        error_type=ctx.get("error_type"),
    )


@app.get("/health")
def health():
    return {"status": "ok"}
