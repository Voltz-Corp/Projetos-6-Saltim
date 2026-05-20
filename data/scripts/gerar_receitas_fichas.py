#!/usr/bin/env python3
"""
Gera data/New/receitas.csv e data/New/fichas_tecnicas.csv a partir das bases antigas,
mapeando para ingredientes.csv atualizado.
"""
from __future__ import annotations

import csv
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
NEW = DATA / "New"
PRODUCTION_CATEGORY_ID = "CAT0015"
PRODUCTION_CATEGORY_NAME = "Produção"

# --- Receitas promovidas de Produção órfã para Produto Final ---
PROMOTE_TO_PF: set[int] = {19, 31, 33, 34, 35, 43, 44, 45, 46, 61}

# Consolidação: várias receitas antigas → um único PF
CONSOLIDATE_GROUPS: dict[str, dict] = {
    "BANNOFFE_DE_COLHER": {
        "old_ids": {36, 37, 38, 39},
        "name": "BANNOFFE DE COLHER",
        "primary_old_id": 39,
    },
    "LEMON_BAR": {
        "old_ids": {40, 41, 42},
        "name": "LEMON BAR",
        "primary_old_id": 42,
    },
}

# Produção: saída = ingrediente semi-pronto (old_recipe_id → output_ingredient_id)
OLD_RECIPE_OUTPUT: dict[int, str] = {
    1: "ING0077",   # CREAM CHEESE AERADO
    2: "ING0196",   # GELEIA DE FRUTAS VERMELHAS
    3: "ING0140",   # MIX DE NUTS
    4: "ING0216",   # RABANADA
    5: "ING0186",   # CREME INGLÊS CÍTRICO
    6: "ING0080",   # CREME DE RICOTA
    7: "ING0119",   # AIOLI DE LIMÃO SICILIANO
    8: "ING0200",   # HOMMUS DE BETERRABA
    9: "ING0141",   # MIX DE SEMENTES
    10: "ING0219",  # TOMATE CONFIT
    11: "ING0120",  # KIMCHI
    12: "ING0082",  # FONDUTA DE QUEIJOS
    13: "ING0158",  # CHICKEN KATSU
    14: "ING0184",  # COLESLAW
    15: "ING0213",  # PICLES DE PEPINO
    16: "ING0198",  # GRANOLA
    17: "ING0208",  # OVERNIGHT OATS
    18: "ING0144",  # PRALINÉ DE AMENDOIM
    20: "ING0123",  # MOLHO PONZU
    21: "ING0162",  # GEMA CURADA
    22: "ING0187",  # CRISPY DE COUVE
    23: "ING0193",  # FAROFA DE PÃO
    24: "ING0170",  # BANANA BREAD
    25: "ING0209",  # PANQUECA DE BANANA
    26: "ING0182",  # CARAMELO
    27: "ING0173",  # BISCOITO CHAMPAGNE
    28: "ING0180",  # BROWNIE
    29: "ING0176",  # BLONDIE
    30: "ING0032",  # GANACHE DE CHOCOLATE
    32: "ING0197",  # GLACÊ DE LARANJA
    47: "ING0021",  # XAROPE DE LIMÃO SICILIANO
    48: "ING0013",  # CONCENTRADO DE CAPIM SANTO
    49: "ING0022",  # XAROPE FERMENTADO DE GENGIBRE
    50: "ING0020",  # XAROPE FRUTAS VERMELHAS E HIBISCO
    51: "ING0014",  # CORDIAL DE CHÁ VERDE
    52: "ING0155",  # CHARQUE ACEBOLADA
    53: "ING0153",  # BATERÁ DE SALMÃO
    54: "ING0201",  # MASSA PARA COXINHA
    55: "ING0185",  # COXINHA LOW CARB
    56: "ING0118",  # AIOLI DE GERGELIM PRETO
    57: "ING0212",  # PASTEL DE SALMÃO
    58: "ING0210",  # PARMÊ DE BERINJELA KATSU
    59: "ING0179",  # BOLINHO DE CENOURA
    60: "ING0177",  # BOLINHA DE PÃO DE QUEIJO
}

# Overrides explícitos: old_ingredient_id → new ingredient_id
OLD_ING_OVERRIDES: dict[int, str] = {
    1: "ING0076",
    6: "ING0133",
    7: "ING0130",
    11: "ING0135",
    12: "ING0215",
    13: "ING0163",
    14: "ING0088",
    21: "ING0097",
    22: "ING0078",
    23: "ING0076",
    24: "ING0238",
    36: "ING0148",
    46: "ING0105",
    49: "ING0006",
    57: "ING0232",
    63: "ING0119",
    68: "ING0149",
    70: "ING0061",
    72: "ING0133",
    73: "ING0145",
    74: "ING0148",
    75: "ING0135",
    78: "ING0008",
    79: "ING0146",
    81: "ING0009",
    83: "ING0130",
    84: "ING0167",
    85: "ING0126",
    87: "ING0227",
    88: "ING0163",
    89: "ING0238",
    90: "ING0126",
    92: "ING0109",
    97: "ING0034",
    99: "ING0089",
    100: "ING0088",
    101: "ING0068",
    103: "ING0069",
    104: "ING0163",
    107: "ING0031",
    108: "ING0027",
    110: "ING0030",
    111: "ING0087",
    113: "ING0031",
    114: "ING0066",
    115: "ING0043",
    116: "ING0242",
    117: "ING0069",
    118: "ING0197",
    119: "ING0028",
    121: "ING0224",
    122: "ING0175",
    125: "ING0228",
    127: "ING0024",
    128: "ING0159",
    131: "ING0016",
    132: "ING0066",
    133: "ING0089",
    134: "ING0061",
    135: "ING0223",
    137: "ING0034",
    138: "ING0075",
    139: "ING0189",
    140: "ING0182",
    142: "ING0089",
    145: "ING0194",
    146: "ING0079",
    147: "ING0159",
    148: "ING0171",
    149: "ING0206",
    150: "ING0202",
    152: "ING0081",
    153: "ING0028",
    156: "ING0087",
    157: "ING0086",
    158: "ING0142",
    159: "ING0030",
    160: "ING0143",
    161: "ING0029",
    162: "ING0030",
    165: "ING0053",
    167: "ING0156",
    173: "ING0232",
    181: "ING0233",
    186: "ING0093",
    192: "ING0094",
    195: "ING0186",
    196: "ING0040",
    203: "ING0092",
    204: "ING0116",
    208: "ING0165",
    212: "ING0164",
    218: "ING0084",
    219: "ING0038",
    221: "ING0208",
    222: "ING0196",
    223: "ING0144",
    225: "ING0107",
    226: "ING0131",
    227: "ING0123",
    228: "ING0115",
    230: "ING0120",
    236: "ING0240",
    238: "ING0122",
    239: "ING0172",
    242: "ING0205",
    244: "ING0218",
    245: "ING0144",
    246: "ING0209",
    248: "ING0173",
    250: "ING0180",
    251: "ING0181",
    252: "ING0176",
    253: "ING0186",
    254: "ING0021",
    257: "ING0023",
    258: "ING0013",
    259: "ING0019",
    260: "ING0022",
    261: "ING0020",
    262: "ING0014",
    266: "ING0217",
    267: "ING0048",
    269: "ING0190",
    270: "ING0155",
    277: "ING0152",
    278: "ING0188",
    279: "ING0165",
    280: "ING0083",
    281: "ING0178",
    282: "ING0220",
    283: "ING0124",
    284: "ING0185",
    285: "ING0118",
    286: "ING0212",
    287: "ING0169",
    288: "ING0154",
    289: "ING0123",
    297: "ING0063",
    298: "ING0210",
    299: "ING0179",
    302: "ING0177",
    303: "ING0084",
    307: "ING0198",
    308: "ING0008",
}

# Aliases por nome normalizado
NAME_ALIASES: dict[str, str] = {
    "CREAM CHEESE PHILADELFIA": "ING0076",
    "CREAM CHEESE EM TEMPERATURA AMB": "ING0076",
    "RICOTA FORMAGGIO": "ING0097",
    "CREME DE LEITE NESTLE": "ING0078",
    "CASTANHA DE CAJU INTEIRA": "ING0133",
    "CASTANHAS DE CAJU": "ING0133",
    "AMENDOIM DESCASCADO SEM SAL": "ING0130",
    "AMENDOIM DESCASCADO S SAL": "ING0130",
    "AMENDOIM SEM SAL TORRADO": "ING0131",
    "FITA DE COCO DESIDRATADO": "ING0135",
    "FITAS DE COCO DESIDRATADO": "ING0135",
    "PAO DE CAIXA PACHAMAMA": "ING0215",
    "OVO DE GALINHA": "ING0163",
    "OVOS DE GALINHA": "ING0163",
    "OVOS": "ING0163",
    "LEITE DE VACA": "ING0088",
    "LEITE INTEGRAL DE VACA": "ING0088",
    "LEITE NINHO": "ING0087",
    "LEITE EM PO INTEGRAL": "ING0087",
    "SAL": "ING0238",
    "SAL DE COZINHA FINO": "ING0238",
    "ACUCAR": "ING0006",
    "AÇUCAR": "ING0006",
    "PEPITA DE GIRASSOL": "ING0148",
    "SEMENTES DE GIRASSOL": "ING0148",
    "SEMENTES DE ABOBORA": "ING0145",
    "SEMENTES DE MOSTARDA": "ING0149",
    "SEMENTES DE CHIA": "ING0146",
    "CEBOLINHO": "ING0105",
    "PIMENTA GOCHUJANG": "ING0232",
    "PIMENTA EM PASTA GOCHUJANG": "ING0232",
    "MOLHO AIOLI DE LIMAO SICILIANO": "ING0119",
    "AVEIA INTEGRAL EM FLOCOS": "ING0061",
    "AVEIA EM FLOCOS FINOS": "ING0061",
    "MEL DE VERI": "ING0008",
    "MEL DE FLOR DA AROEIRA": "ING0008",
    "XAROPE DE ACUCAR DEMERARA": "ING0009",
    "TOFU FIRME PEGADA VIVA": "ING0167",
    "MOLHO SHOYU SAKURA FOOD SERVICE": "ING0126",
    "GENGIBRE FRESCO": "ING0227",
    "COUVE FOLHA": "ING0109",
    "BANANA PRATA MADURA": "ING0034",
    "BANANA MADURA": "ING0034",
    "MANTEIGA DERRETIDA": "ING0089",
    "MANTEIGA GELADA EM CUBOS": "ING0089",
    "MANTEIGA EM CUBOS": "ING0089",
    "FERMENTO BIO SECO": "ING0068",
    "FERMENTO EM PO QUIMICO": "ING0069",
    "FERMENTO QUIMICO LIDER": "ING0069",
    "CHOCOLATE EM BARRA MEIO AMARGO": "ING0031",
    "CHOCOLATE MEIO AMARGO EM BARRA": "ING0031",
    "CACAU EM PO 100": "ING0027",
    "CHOCOLATE EM BARRA BRANCO": "ING0030",
    "CHOCOLATE BRANCO EM BARRA": "ING0030",
    "CHOCOLATE AMARGO EM BARRA": "ING0028",
    "CHOCLATE AMARGO EM BARRA": "ING0028",
    "CHOCOLATE 70 AMARGO EM BARRA": "ING0029",
    "FARINHA DE TRIGO SEM FERMENTO": "ING0066",
    "FARINHA DE TRIGO FINNA SEM FERMENTO": "ING0066",
    "SUMO DE LARANJA BAHIA": "ING0043",
    "OLEO DE CANOLA OU GIRASSOL": "ING0242",
    "GLACE DE LARANJA BAHIA": "ING0197",
    "CAPIM SANTO FRESCO": "ING0224",
    "BISCOITO NEGRESCO TRITURADO": "ING0175",
    "MATCHA CULINARIO": "ING0228",
    "AGUA GELADA": "ING0024",
    "ESSENCIA DE BAUNILHA LIQUIDA": "ING0016",
    "CANELA EM PO": "ING0223",
    "CREME DE CHANTILLY": "ING0075",
    "CRUMBLE AMANTEIGADO": "ING0189",
    "CARAMELO AMANTEIGADO": "ING0182",
    "GELATINA EM PO INCOLOR 50ML DE AGUA": "ING0194",
    "CREME DE LEITE FRESCO GELADO": "ING0079",
    "BASE DE BISCOITO AMANTEIGADO": "ING0171",
    "MOUSSE DE LIMAO SICILIANO": "ING0206",
    "MOUSSE DE LIMCAO SICILIANO": "ING0206",
    "MERENGUE MACARICADO": "ING0202",
    "DOCE DE LEITE ITAMBE LATAO": "ING0081",
    "LEITE CONDENSADO ITAMBE": "ING0086",
    "PASTA DE PISTACHE CONCENTRADO ZERO": "ING0142",
    "PISTACHE QUEBRADO": "ING0143",
    "HIBISCO INTEIRO SECO": "ING0053",
    "CHARQUE DESFIADA DESSALGADA": "ING0156",
    "QUEIJO PARMESAO": "ING0093",
    "MUSSARELA MOLFINO": "ING0092",
    "QUEIJO BOURSIN DIPOTENZA": "ING0094",
    "QUEIJO GORGONZOLA": "ING0083",
    "CREME INGLES CITRICO": "ING0186",
    "CREME INGLES CITRICO": "ING0186",
    "MORANGOS FRESCOS": "ING0040",
    "PRESUNTO CRU SERRANO": "ING0165",
    "PRESUNTO SERRANO LAMINADO": "ING0165",
    "PRESUNTO COZIDO DE PERU": "ING0164",
    "IOGURTE CREGO": "ING0084",
    "MANGA TOMMY CONGELADA": "ING0038",
    "OVERNIGHT OATS CHIA": "ING0208",
    "GELEIA DE FRUTAS VERMELHAS": "ING0196",
    "GELÉIA DE FRUTAS VERMELHAS": "ING0196",
    "PRALINE AMENDOIM": "ING0144",
    "PRALINE DE AMENDOIM": "ING0144",
    "CENOURA EM FIOS": "ING0107",
    "MOLHO PONZU DE GENGIBRE": "ING0123",
    "KIM CHI": "ING0120",
    "AZEITE": "ING0240",
    "MOLHO DE TOMATES ASSADOS": "ING0122",
    "BATATAS ASSADAS": "ING0172",
    "MIX DE VEGETAIS ESCALF": "ING0205",
    "SORVETE DE CREME APENINE": "ING0218",
    "SORVETE DE CHOCOLATE APENINE": "ING0217",
    "CALDA DE CARAMELO AMANTEIGADO": "ING0181",
    "BROWNIE PORCAO": "ING0180",
    "BLONDIE PORCAO": "ING0176",
    "AGUA COM GAS 300ML": "ING0023",
    "XAROPE FERM DE GENGIBRE": "ING0022",
    "XAROPE F V HIBISCOS": "ING0020",
    "GUARNICAO CAPIM SANTO LIMAO": "ING0019",
    "CUSCUZ SOLTINHO": "ING0190",
    "BACON FATIADO": "ING0152",
    "BOLINHAS DE RICOTA MARINADA": "ING0178",
    "TORRADINHAS DE CIABATTA": "ING0220",
    "COXINHAS LOW CARB": "ING0185",
    "PORCAO DE PASTEIS DE SALMAO": "ING0212",
    "ARROZ PARBOILIZADO COZIDA": "ING0169",
    "CAMARAO TAILON 21 25": "ING0154",
    "CAMARAO 36 40": "ING0154",
    "PASTA TEMPERO ALHO SAL PIMENTA": "ING0233",
    "PASTA TEMPERO": "ING0233",
    "PARME DE FRANGO KATSU": "ING0211",
    "PARMÊ DE FRANGO KATSU": "ING0211",
    "ESPAGUETE AL DENTE": "ING0063",
    "PARME DE BERINGELA KATSU": "ING0210",
    "PARMÊ DE BERINGELA KATSU": "ING0210",
    "BOLINHO DE CENOURA": "ING0179",
    "BOLINHAS DE PAO DE QUEIJO": "ING0177",
    "IOGURTE POLILAC": "ING0084",
    "FOLHAS DE RUCULA": "ING0116",
    "GRANOLA DA CASA": "ING0198",
    "SALADA COLESLAW": "ING0184",
    "CREME DE RICOTA": "ING0080",
    "HOMMUS DE BETERRABA": "ING0200",
    "FONDUTA DE QUEIJOS": "ING0082",
    "RABANADA": "ING0216",
    "CHICKEN KATSU": "ING0158",
    "PICLES DE PEPINO": "ING0213",
    "BANANA BREAD": "ING0170",
    "PANQUECAS DE BANANA": "ING0209",
    "BISCOITO CHAMPAGNE": "ING0173",
    "MOLHO BECHAMEL": "ING0121",
    "MOLHO BÉCHAMEL": "ING0121",
    "CREAM CHEESE AERADO": "ING0077",
    "GELÉIA DE FRUTAS VERMELHAS": "ING0196",
    "MIX DE NUTS E FRUTAS SECAS": "ING0140",
    "RABANADA DE PAO DE CAIXA": "ING0216",
    "CREME INGLES CITRICO": "ING0186",
    "AIOLI DE LIMAO SICILIANO": "ING0119",
    "TOMATES CEREJA CONFIT": "ING0219",
    "CHICKEN KATSU PRODUCAO": "ING0158",
    "GRANOLA": "ING0198",
    "PRALINÉ DE AMENDOIM": "ING0144",
    "TOFU KATSU PRODUCAO": "ING0167",
    "CRISPY DE COUVE FOLHA": "ING0187",
    "FAROFA DE PAO": "ING0193",
    "PANQUECAS DE BANANA": "ING0209",
    "BROWNIE SALTIM": "ING0180",
    "BLONDIE SALTIM": "ING0176",
    "GANACHE DE CHOCOLATE AMARGO": "ING0032",
    "GLACÊ DE LARANJA BAHIA": "ING0197",
    "GLACE DE LARANJA BAHIA": "ING0197",
    "XAROPE DE GENGIBRE FERMENTADO": "ING0022",
    "XAROPE DE FRUTAS VERMELHAS HIBISCO": "ING0020",
    "XAROPE FV HIBISCOS": "ING0020",
    "CHARQUE ACEBOLADA": "ING0155",
    "BATERÁ DE SALMÃO": "ING0153",
    "BATERA DE SALMAO": "ING0153",
    "MASSA PARA COXINHA": "ING0201",
    "COXINHAS DE FRANGO LOW CARB PRODUCAO": "ING0185",
    "AIOLI DE GERGELIM PRETO": "ING0118",
    "MINI PASTEIS DE SALMAO": "ING0212",
    "PARME DE BERINJELA KATSU PRODUCAO": "ING0210",
    "PRODUCAO DO PAO DE QUEIJO": "ING0177",
    "GEMAS CURADAS": "ING0162",
    "GEMA CURADA": "ING0162",
    "MIX DE SEMENTES": "ING0141",
    "MOLHO PONZU": "ING0123",
    "CORDIAL DE CHA VERDE": "ING0014",
    "XAROPE DE LIMAO SICILIANO": "ING0021",
    "CONCENTRADO DE CAPIM SANTO": "ING0013",
    "MERENGUE": "ING0202",
    "MOUSSE DE LIMAO PARA LEMON BAR": "ING0206",
    "BASE DE BISCOITO": "ING0171",
    "CREME INGLÊS CÍTRICO": "ING0186",
    "CREME INGLES CITRICO": "ING0186",
    "MOLHO AIOLI DE LIMÃO SICILIANO": "ING0119",
    "AIOLI DE LIMÃO SICILIANO": "ING0119",
    "MOLHO DE TOMATE ASSADO": "ING0122",
    "TOMATE CONFIT": "ING0219",
    "COLESLAW": "ING0184",
    "CARAMELO": "ING0182",
    "CHOCOLATE SHAKE": "ING0217",
    "SORVETE DE CREME": "ING0218",
    "SORVETE DE CHOCOLATE": "ING0217",
    "CUMARU EM SEMENTE": "ING0225",
    "PORÇÃO DE CHIPS": "ING0203",
    "MIX DE CHIPS": "ING0203",
    "PORÇÃO DE PÃO DE QUEIJO": "ING0177",
    "BOLINHO DE CENOURA PRODUTO FINAL": "ING0179",
    "PARMÊ DE FRANGO EXECUTIVO PRODUTO": "ING0211",
    "PARMÊ DE BERINGELA MELANZANE EXECUTIVO PRODUTO": "ING0210",
}


def normalize_name(value: str | None) -> str:
    if not value:
        return ""
    text = str(value).upper().strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"\s*-\s*", " ", text)
    text = re.sub(r"[^A-Z0-9 ]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def load_new_ingredients() -> dict[str, dict]:
    ingredients: dict[str, dict] = {}
    with open(NEW / "ingredientes.csv", newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            ingredients[row["id"]] = {
                "name": row["name"],
                "unit": row["unit"],
                "norm": normalize_name(row["name"]),
            }
    return ingredients


def build_old_ingredient_map(ingredients: dict[str, dict]) -> dict[int, str]:
    norm_to_id = {d["norm"]: iid for iid, d in ingredients.items()}
    aliases = dict(NAME_ALIASES)
    for iid, data in ingredients.items():
        aliases[data["norm"]] = iid

    wb = openpyxl.load_workbook(DATA / "lista_ingredientes.xlsx", data_only=True)
    mapping: dict[int, str] = {}

    for row in wb["Sheet1"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        old_id = int(row[0])
        if old_id in OLD_ING_OVERRIDES:
            mapping[old_id] = OLD_ING_OVERRIDES[old_id]
            continue

        name = row[1]
        norm = normalize_name(name)
        target = aliases.get(norm) or norm_to_id.get(norm)

        if not target:
            best = None
            best_len = 0
            for iid, data in ingredients.items():
                if norm in data["norm"] or data["norm"] in norm:
                    if len(data["norm"]) > best_len:
                        best_len = len(data["norm"])
                        best = iid
            target = best

        if not target:
            raise ValueError(f"Ingrediente antigo sem mapeamento: {old_id} - {name}")
        mapping[old_id] = target

    return mapping


def load_old_recipes() -> dict[int, dict]:
    wb = openpyxl.load_workbook(DATA / "lista_receitas.xlsx", data_only=True)
    recipes: dict[int, dict] = {}
    for row in wb["Sheet1"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        old_id = int(row[0])
        recipes[old_id] = {
            "name": str(row[1]).strip().upper(),
            "tipo": str(row[2]).strip(),
            "custo": row[3],
            "preco": row[4],
            "rendimento": row[5],
        }
    return recipes


def clean_recipe_name(name: str) -> str:
    display = name.strip().upper()
    for suffix in (" (PRODUÇÃO)", " PRODUÇÃO", "- PRODUÇÃO", " (PRODUTO FINAL)"):
        display = display.replace(suffix, "")
    return display.strip()


def build_recipes(old_recipes: dict[int, dict], ingredients: dict[str, dict]) -> tuple[list[dict], dict[int, str]]:
    """Retorna lista de receitas e mapa old_recipe_id -> new_recipe_id."""
    recipe_rows: list[dict] = []
    old_to_new: dict[int, str] = {}

    consolidated_keys: dict[str, str] = {}

    def next_rec_id(n: int) -> str:
        return f"REC{n:04d}"

    # 1) PRODUÇÃO
    prod_counter = 0
    for old_id in sorted(OLD_RECIPE_OUTPUT):
        if old_id not in old_recipes:
            raise ValueError(f"Receita produção {old_id} ausente em lista_receitas")
        meta = old_recipes[old_id]
        output_id = OLD_RECIPE_OUTPUT[old_id]
        prod_counter += 1
        rec_id = next_rec_id(prod_counter)
        old_to_new[old_id] = rec_id
        recipe_rows.append(
            {
                "id": rec_id,
                "name": clean_recipe_name(meta["name"]),
                "type": "PRODUCAO",
                "yield_qty": meta["rendimento"],
                "yield_unit": ingredients[output_id]["unit"],
                "output_ingredient_id": output_id,
                "sale_price": "",
            }
        )

    # 2) PRODUTO FINAL — cardápio original (62-110)
    pf_counter = prod_counter
    for old_id in sorted(old_recipes):
        meta = old_recipes[old_id]
        if meta["tipo"].lower() != "produto final":
            continue
        pf_counter += 1
        rec_id = next_rec_id(pf_counter)
        old_to_new[old_id] = rec_id
        recipe_rows.append(
            {
                "id": rec_id,
                "name": meta["name"],
                "type": "PRODUTO_FINAL",
                "yield_qty": 1,
                "yield_unit": "UND",
                "output_ingredient_id": "",
                "sale_price": meta["preco"],
            }
        )

    # 3) PRODUTO FINAL — consolidados
    for group_key, group in CONSOLIDATE_GROUPS.items():
        pf_counter += 1
        rec_id = next_rec_id(pf_counter)
        consolidated_keys[group_key] = rec_id
        primary = old_recipes[group["primary_old_id"]]
        for oid in group["old_ids"]:
            old_to_new[oid] = rec_id
        recipe_rows.append(
            {
                "id": rec_id,
                "name": group["name"],
                "type": "PRODUTO_FINAL",
                "yield_qty": 1,
                "yield_unit": "UND",
                "output_ingredient_id": "",
                "sale_price": primary["preco"],
            }
        )

    # 4) PRODUTO FINAL — promovidos (exceto já consolidados)
    consolidated_old_ids = {oid for g in CONSOLIDATE_GROUPS.values() for oid in g["old_ids"]}
    for old_id in sorted(PROMOTE_TO_PF):
        if old_id in consolidated_old_ids:
            continue
        if old_id not in old_recipes:
            raise ValueError(f"Receita promovida {old_id} ausente")
        meta = old_recipes[old_id]
        pf_counter += 1
        rec_id = next_rec_id(pf_counter)
        old_to_new[old_id] = rec_id
        recipe_rows.append(
            {
                "id": rec_id,
                "name": clean_recipe_name(meta["name"]),
                "type": "PRODUTO_FINAL",
                "yield_qty": 1,
                "yield_unit": "UND",
                "output_ingredient_id": "",
                "sale_price": meta["preco"],
            }
        )

    # Garantir nomes únicos: em conflito produção vs PF, renomeia a PRODUÇÃO.
    by_name: dict[str, list[dict]] = defaultdict(list)
    for recipe in recipe_rows:
        by_name[recipe["name"]].append(recipe)

    for name, group in by_name.items():
        if len(group) < 2:
            continue
        producoes = [r for r in group if r["type"] == "PRODUCAO"]
        finais = [r for r in group if r["type"] == "PRODUTO_FINAL"]
        for prod in producoes:
            if prod["output_ingredient_id"]:
                alt = ingredients[prod["output_ingredient_id"]]["name"]
                if alt != name and all(r["name"] != alt for r in recipe_rows if r["id"] != prod["id"]):
                    prod["name"] = alt
                    continue
            prod["name"] = f"{name} - PREPARO"

    return recipe_rows, old_to_new


def build_fichas(
    old_to_new: dict[int, str],
    old_ing_map: dict[int, str],
    ingredients: dict[str, dict],
    recipes_by_id: dict[str, dict],
) -> list[dict]:
    wb = openpyxl.load_workbook(DATA / "fichas_tecnicas_ids.xlsx", data_only=True)
    aggregated: dict[tuple[str, str], float] = {}

    for row in wb["Sheet1"].iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        if isinstance(row[0], str) and row[0] == "ID_Receita":
            continue

        old_rec = int(float(row[0]))
        old_ing = int(float(row[1]))
        qty = row[2]

        if qty is None:
            continue
        qty = float(qty)
        if qty <= 0:
            continue

        recipe_id = old_to_new.get(old_rec)
        if not recipe_id:
            raise ValueError(f"Receita antiga {old_rec} sem mapeamento para REC")

        ingredient_id = old_ing_map.get(old_ing)
        if not ingredient_id:
            raise ValueError(f"Ingrediente antigo {old_ing} sem mapeamento para ING")

        recipe = recipes_by_id[recipe_id]
        output_id = recipe.get("output_ingredient_id") or ""
        # Base antiga às vezes registra a porção final como "insumo" da própria produção.
        if output_id and ingredient_id == output_id:
            continue

        key = (recipe_id, ingredient_id)
        aggregated[key] = aggregated.get(key, 0.0) + qty

    fichas: list[dict] = []
    for (recipe_id, ingredient_id), qty in sorted(aggregated.items()):
        fichas.append(
            {
                "recipe_id": recipe_id,
                "ingredient_id": ingredient_id,
                "qty": round(qty, 4),
                "unit": ingredients[ingredient_id]["unit"],
            }
        )
    return fichas


def validate(recipes: list[dict], fichas: list[dict], ingredients: dict[str, dict]) -> None:
    recipe_ids = {r["id"] for r in recipes}
    ing_ids = set(ingredients)

    for ficha in fichas:
        assert ficha["recipe_id"] in recipe_ids, f"recipe_id inválido: {ficha['recipe_id']}"
        assert ficha["ingredient_id"] in ing_ids, f"ingredient_id inválido: {ficha['ingredient_id']}"

    outputs: dict[str, str] = {}
    for recipe in recipes:
        if recipe["type"] == "PRODUCAO":
            out = recipe["output_ingredient_id"]
            assert out in ing_ids, f"output inválido em {recipe['id']}"
            assert out not in outputs.values() or outputs.get(out) == recipe["id"], (
                f"output duplicado {out}"
            )
            outputs[recipe["id"]] = out
        else:
            assert recipe["sale_price"] not in ("", None), f"PF sem preço: {recipe['id']}"
            assert float(recipe["sale_price"]) > 0, f"PF preço inválido: {recipe['id']}"

    prod_outputs = [r["output_ingredient_id"] for r in recipes if r["type"] == "PRODUCAO"]
    assert len(prod_outputs) == len(set(prod_outputs)), "output_ingredient_id duplicado entre produções"

    for recipe in recipes:
        if recipe["type"] != "PRODUCAO":
            continue
        rec_fichas = [f for f in fichas if f["recipe_id"] == recipe["id"]]
        assert rec_fichas, f"Produção sem ficha: {recipe['id']} {recipe['name']}"


def ensure_production_category() -> None:
    """Garante CAT0015 Produção em categorias.csv."""
    path = NEW / "categorias.csv"
    rows: list[dict[str, str]] = []
    found = False
    with open(path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        fieldnames = reader.fieldnames or ["id", "name"]
        for row in reader:
            if row["id"] == PRODUCTION_CATEGORY_ID:
                row["name"] = PRODUCTION_CATEGORY_NAME
                found = True
            rows.append(row)
    if not found:
        rows.append({"id": PRODUCTION_CATEGORY_ID, "name": PRODUCTION_CATEGORY_NAME})
    write_csv(path, fieldnames, rows)


def sync_production_ingredient_categories(recipes: list[dict]) -> int:
    """
    Atribui CAT0015 (Produção) aos ingredientes que são saída de receitas PRODUCAO.
    Alinha ingredientes.csv com receitas.csv (type=PRODUCAO ↔ output_ingredient_id).
    """
    output_ids = {
        r["output_ingredient_id"]
        for r in recipes
        if r["type"] == "PRODUCAO" and r.get("output_ingredient_id")
    }
    path = NEW / "ingredientes.csv"
    rows: list[dict[str, str]] = []
    updated = 0
    with open(path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        fieldnames = reader.fieldnames or ["id", "name", "unit", "category_id", "current_qty"]
        for row in reader:
            if row["id"] in output_ids and row["category_id"] != PRODUCTION_CATEGORY_ID:
                row["category_id"] = PRODUCTION_CATEGORY_ID
                updated += 1
            rows.append(row)
    write_csv(path, fieldnames, rows)
    return updated


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    ingredients = load_new_ingredients()
    old_ing_map = build_old_ingredient_map(ingredients)
    old_recipes = load_old_recipes()

    recipes, old_to_new = build_recipes(old_recipes, ingredients)
    recipes_by_id = {r["id"]: r for r in recipes}

    fichas = build_fichas(old_to_new, old_ing_map, ingredients, recipes_by_id)
    validate(recipes, fichas, ingredients)

    write_csv(
        NEW / "receitas.csv",
        ["id", "name", "type", "yield_qty", "yield_unit", "output_ingredient_id", "sale_price"],
        recipes,
    )
    write_csv(
        NEW / "fichas_tecnicas.csv",
        ["recipe_id", "ingredient_id", "qty", "unit"],
        fichas,
    )

    ensure_production_category()
    cat_updates = sync_production_ingredient_categories(recipes)
    production_outputs = sum(1 for r in recipes if r["type"] == "PRODUCAO")

    prod = sum(1 for r in recipes if r["type"] == "PRODUCAO")
    pf = sum(1 for r in recipes if r["type"] == "PRODUTO_FINAL")
    print(f"Ingredientes mapeados: {len(old_ing_map)}")
    print(f"Receitas geradas: {len(recipes)} ({prod} PRODUCAO, {pf} PRODUTO_FINAL)")
    print(f"Fichas técnicas: {len(fichas)}")
    print(
        f"Categoria {PRODUCTION_CATEGORY_ID} ({PRODUCTION_CATEGORY_NAME}): "
        f"{production_outputs} ingredientes; {cat_updates} atualizados em ingredientes.csv"
    )
    print(f"Escritos: {NEW / 'receitas.csv'} e {NEW / 'fichas_tecnicas.csv'}")


if __name__ == "__main__":
    main()
